"""
monitor_backlinks.py
====================
Monitors all published backlinks from campaign Excel files.
Checks each URL, marks Active / Missing / Dead, and saves a report.

Usage:
    python monitor_backlinks.py
    python monitor_backlinks.py --batch 25      # check 25 at a time (default)
    python monitor_backlinks.py --timeout 15    # seconds per request
"""

import os
import sys
import time
import random
import logging
import argparse
import requests
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
]


def load_all_published_links() -> list[dict]:
    """Read all published links from every Excel file in data/."""
    try:
        import openpyxl
    except ImportError:
        logger.error("pip install openpyxl")
        sys.exit(1)

    links = []
    if not os.path.isdir(DATA_DIR):
        logger.error(f"Data directory not found: {DATA_DIR}")
        return links

    for fname in sorted(os.listdir(DATA_DIR)):
        if not fname.endswith(".xlsx") or "monitor" in fname.lower():
            continue
        fpath = os.path.join(DATA_DIR, fname)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            if "Published Links" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Published Links"]
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                d = dict(zip(headers, row))
                url = d.get("Live URL") or d.get("Article URL") or d.get("live_url") or ""
                if url and str(url).startswith("http"):
                    links.append({
                        "source_file": fname,
                        "platform": d.get("Platform") or d.get("platform") or "",
                        "da": d.get("DA") or d.get("da") or 0,
                        "title": d.get("Title") or d.get("title") or "",
                        "url": str(url).strip(),
                        "anchor": d.get("Anchor Text") or d.get("anchor") or "",
                        "date_posted": str(d.get("Date") or d.get("date") or ""),
                    })
            wb.close()
        except Exception as e:
            logger.warning(f"Could not read {fname}: {e}")

    # Deduplicate by URL
    seen = set()
    unique = []
    for link in links:
        if link["url"] not in seen:
            seen.add(link["url"])
            unique.append(link)
    return unique


def check_url(entry: dict, timeout: int = 15) -> dict:
    """Check a single URL and return its status."""
    url = entry["url"]
    headers = {"User-Agent": random.choice(USER_AGENTS)}
    result = entry.copy()
    result["checked_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    try:
        r = requests.get(url, headers=headers, timeout=timeout,
                         allow_redirects=True)
        code = r.status_code
        if code == 200:
            # Check if our target domain is still in the page
            content = r.text.lower()
            has_link = "scotle.org" in content or "scotle" in content
            if has_link:
                result["status"] = "Active"
                result["http_code"] = code
                result["action"] = ""
            else:
                result["status"] = "Missing — Link Removed"
                result["http_code"] = code
                result["action"] = "Re-outreach"
        elif code in (301, 302, 303, 307, 308):
            result["status"] = "Redirect"
            result["http_code"] = code
            result["action"] = "Verify redirect target"
        elif code == 404:
            result["status"] = "Dead — 404"
            result["http_code"] = code
            result["action"] = "Re-post to platform"
        elif code in (403, 401):
            result["status"] = "Blocked — Auth Required"
            result["http_code"] = code
            result["action"] = "Check Manually"
        else:
            result["status"] = f"Error — HTTP {code}"
            result["http_code"] = code
            result["action"] = "Check Manually"
    except requests.exceptions.ConnectionError:
        result["status"] = "Dead — Connection Failed"
        result["http_code"] = 0
        result["action"] = "Re-post to platform"
    except requests.exceptions.Timeout:
        result["status"] = "Timeout"
        result["http_code"] = 0
        result["action"] = "Check Manually"
    except Exception as e:
        result["status"] = f"Error — {str(e)[:50]}"
        result["http_code"] = 0
        result["action"] = "Check Manually"

    return result


def check_batch(links: list[dict], batch_size: int = 25, timeout: int = 15,
                max_workers: int = 8) -> list[dict]:
    """Check all links in batches using threading."""
    results = []
    total = len(links)

    for batch_start in range(0, total, batch_size):
        batch = links[batch_start: batch_start + batch_size]
        batch_num = batch_start // batch_size + 1
        total_batches = (total + batch_size - 1) // batch_size
        logger.info(f"Batch {batch_num}/{total_batches} — checking {len(batch)} links ...")

        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = {ex.submit(check_url, entry, timeout): entry for entry in batch}
            for future in as_completed(futures):
                try:
                    results.append(future.result())
                except Exception as e:
                    entry = futures[future]
                    entry["status"] = f"Error — {e}"
                    entry["http_code"] = 0
                    entry["action"] = "Check Manually"
                    entry["checked_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                    results.append(entry)

        if batch_start + batch_size < total:
            delay = random.uniform(3, 8)
            logger.info(f"Pausing {delay:.1f}s before next batch ...")
            time.sleep(delay)

    return results


def save_report(results: list[dict]) -> str:
    """Save monitoring report to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("pip install openpyxl")
        return ""

    wb = Workbook()
    thin = Side(style="thin", color="DDDDDD")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_COLORS = {
        "Active":                   ("C8E6C9", "1B5E20"),
        "Missing — Link Removed":   ("FFF9C4", "F57F17"),
        "Dead — 404":               ("FFCDD2", "B71C1C"),
        "Dead — Connection Failed": ("FFCDD2", "B71C1C"),
        "Redirect":                 ("E3F2FD", "0D47A1"),
        "Timeout":                  ("FFE0B2", "E65100"),
        "Blocked — Auth Required":  ("F3E5F5", "4A148C"),
    }

    def row_colors(status: str):
        for key, (bg, fg) in STATUS_COLORS.items():
            if status.startswith(key[:10]):
                return bg, fg
        return "F5F5F5", "333333"

    # ── Sheet 1: All Results ──────────────────────────────────────
    ws = wb.active
    ws.title = "Monitor Results"
    headers = ["#", "Platform", "DA", "Status", "HTTP",
               "URL", "Anchor Text", "Date Posted", "Checked At", "Action"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1A237E")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for i, r in enumerate(sorted(results, key=lambda x: (
            0 if "Active" in x.get("status","") else
            1 if "Missing" in x.get("status","") else
            2 if "Dead" in x.get("status","") else 3,
            -(r.get("da") or 0)
    )), 1):
        status = r.get("status", "Unknown")
        bg, fg = row_colors(status)
        row_data = [
            i,
            r.get("platform", ""),
            r.get("da", 0),
            status,
            r.get("http_code", 0),
            r.get("url", ""),
            r.get("anchor", ""),
            r.get("date_posted", ""),
            r.get("checked_at", ""),
            r.get("action", ""),
        ]
        ws.append(row_data)
        row_idx = i + 1
        fill = PatternFill("solid", fgColor=bg)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row_idx, col)
            c.fill = fill
            c.border = brd
            c.alignment = Alignment(vertical="center")
        # Status cell bold
        sc = ws.cell(row_idx, 4)
        sc.font = Font(bold=True, color=fg)
        sc.alignment = Alignment(horizontal="center")
        # URL as hyperlink
        url = r.get("url", "")
        if url:
            uc = ws.cell(row_idx, 6, url)
            uc.hyperlink = url
            uc.font = Font(color="1558D6", underline="single")

    col_widths = [5, 22, 6, 26, 7, 58, 24, 18, 18, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Action Required ──────────────────────────────────
    wa = wb.create_sheet("Action Required")
    action_items = [r for r in results if r.get("action") and r.get("status") != "Active"]
    wa.append(["#", "Platform", "DA", "Status", "URL", "Action", "Checked At"])
    for col in range(1, 8):
        c = wa.cell(1, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="B71C1C")
        c.alignment = Alignment(horizontal="center")
    wa.freeze_panes = "A2"

    for i, r in enumerate(sorted(action_items, key=lambda x: -(x.get("da") or 0)), 1):
        bg, fg = row_colors(r.get("status", ""))
        wa.append([i, r.get("platform",""), r.get("da",0), r.get("status",""),
                   r.get("url",""), r.get("action",""), r.get("checked_at","")])
        row_idx = i + 1
        fill = PatternFill("solid", fgColor=bg)
        for col in range(1, 8):
            wa.cell(row_idx, col).fill = fill
            wa.cell(row_idx, col).border = brd
        url = r.get("url","")
        if url:
            c = wa.cell(row_idx, 5, url)
            c.hyperlink = url
            c.font = Font(color="1558D6", underline="single")
    for i, w in enumerate([5, 22, 6, 26, 58, 22, 18], 1):
        wa.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Summary ──────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 20

    ws2.merge_cells("A1:B1")
    hc = ws2.cell(1, 1, "Backlink Monitor Report — scotle.org")
    hc.font = Font(bold=True, size=15, color="1A237E")
    hc.alignment = Alignment(horizontal="center")
    hc.fill = PatternFill("solid", fgColor="E8EAF6")
    ws2.row_dimensions[1].height = 28

    total       = len(results)
    active      = sum(1 for r in results if "Active" in r.get("status",""))
    missing     = sum(1 for r in results if "Missing" in r.get("status",""))
    dead        = sum(1 for r in results if "Dead" in r.get("status",""))
    timeout_c   = sum(1 for r in results if "Timeout" in r.get("status",""))
    other       = total - active - missing - dead - timeout_c

    row = 3
    ws2.cell(row, 1, "Generated:").font = Font(bold=True)
    ws2.cell(row, 2, datetime.now().strftime("%Y-%m-%d %H:%M"))

    row += 2
    ws2.cell(row, 1, "RESULTS SUMMARY").font = Font(bold=True, size=12, color="1A237E")
    stats = [
        ("Total Links Checked",       total),
        ("Active (link still live)",  active),
        ("Missing (page up, no link)",missing),
        ("Dead (404 / connection)",   dead),
        ("Timeout",                   timeout_c),
        ("Other",                     other),
        ("",                          ""),
        ("Active Rate",               f"{active/max(total,1)*100:.1f}%"),
        ("Links Needing Action",      missing + dead),
    ]
    for label, val in stats:
        row += 1
        c1 = ws2.cell(row, 1, label)
        c2 = ws2.cell(row, 2, val)
        c1.font = Font(bold=bool(label))
        if label == "Active (link still live)":
            c2.font = Font(bold=True, color="1B5E20")
        elif label in ("Missing (page up, no link)", "Dead (404 / connection)"):
            c2.font = Font(bold=True, color="B71C1C")

    row += 2
    ws2.cell(row, 1, "NEXT STEPS").font = Font(bold=True, size=12, color="B71C1C")
    row += 1
    ws2.cell(row, 1, "Missing links:").font = Font(bold=True)
    ws2.cell(row, 2, "Re-outreach to site owners to restore link")
    row += 1
    ws2.cell(row, 1, "Dead pages:").font = Font(bold=True)
    ws2.cell(row, 2, "Re-post article to same platform with new URL")
    row += 1
    ws2.cell(row, 1, "Timeouts:").font = Font(bold=True)
    ws2.cell(row, 2, "Run monitor again — may be temporary downtime")

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(DATA_DIR, f"Scotle_BacklinkMonitor_{ts}.xlsx")
    os.makedirs(DATA_DIR, exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Backlink Monitor — scotle.org")
    parser.add_argument("--batch",   type=int, default=25,  help="Links per batch (default 25)")
    parser.add_argument("--timeout", type=int, default=15,  help="Seconds per request (default 15)")
    parser.add_argument("--workers", type=int, default=8,   help="Parallel workers (default 8)")
    args = parser.parse_args()

    logger.info("=" * 60)
    logger.info("  Backlink Monitor — scotle.org")
    logger.info("=" * 60)

    links = load_all_published_links()
    if not links:
        logger.error("No published links found in data/ folder.")
        sys.exit(1)

    logger.info(f"  Found {len(links)} unique published links to check")
    logger.info(f"  Batch size: {args.batch}  |  Timeout: {args.timeout}s  |  Workers: {args.workers}")
    logger.info("")

    results = check_batch(links, batch_size=args.batch,
                          timeout=args.timeout, max_workers=args.workers)

    # Summary
    active  = sum(1 for r in results if "Active"   in r.get("status",""))
    missing = sum(1 for r in results if "Missing"  in r.get("status",""))
    dead    = sum(1 for r in results if "Dead"     in r.get("status",""))
    other   = len(results) - active - missing - dead

    logger.info("")
    logger.info("=" * 60)
    logger.info(f"  ✓ Active  : {active}")
    logger.info(f"  ⚠ Missing : {missing}  (link removed — re-outreach needed)")
    logger.info(f"  ✗ Dead    : {dead}  (page gone — re-post needed)")
    logger.info(f"  ? Other   : {other}")
    logger.info("=" * 60)

    if missing + dead > 0:
        logger.info(f"\n  Links needing action: {missing + dead}")

    out = save_report(results)
    if out:
        logger.info(f"\n  Report saved: {out}")
    else:
        logger.warning("Could not save Excel report.")


if __name__ == "__main__":
    main()
