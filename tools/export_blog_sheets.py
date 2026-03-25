"""
export_blog_sheets.py
=====================
Exports blog listing data into a Google Sheets-ready Excel file
with multiple sheets, formatting, filters, and color coding.

Upload the output file directly to Google Sheets:
    1. Go to sheets.google.com
    2. File > Import > Upload
    3. Select the .xlsx file
    4. Choose "Replace spreadsheet"

Usage:
    python export_blog_sheets.py
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "output"
INPUT_CSV = os.path.join(OUTPUT_DIR, "blog_listing_sites.csv")
OUTPUT_XLSX = os.path.join(OUTPUT_DIR, "Blog_Listing_Sites_GoogleSheets.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)


def style_header(ws, num_cols):
    """Apply header styling."""
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_rows(ws, num_rows, num_cols, da_col=None):
    """Apply row styling with DA-based color coding."""
    green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid")
    orange_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    dofollow_font = Font(name="Arial", size=10, color="137333")
    nofollow_font = Font(name="Arial", size=10, color="C5221F")
    normal_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="1A73E8", underline="single")

    thin_border = Border(
        bottom=Side(style="thin", color="E8EAED"),
        right=Side(style="thin", color="E8EAED"),
    )

    for row in range(2, num_rows + 2):
        # Alternate row colors
        row_fill = white_fill if row % 2 == 0 else alt_fill

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        # Color code DA column
        if da_col:
            da_cell = ws.cell(row=row, column=da_col)
            try:
                da_val = int(float(str(da_cell.value or 0)))
                if da_val >= 80:
                    da_cell.fill = green_fill
                    da_cell.font = Font(name="Arial", size=10, bold=True, color="137333")
                elif da_val >= 50:
                    da_cell.fill = yellow_fill
                    da_cell.font = Font(name="Arial", size=10, color="7C4A03")
                else:
                    da_cell.fill = orange_fill
                    da_cell.font = Font(name="Arial", size=10, color="C5221F")
            except (ValueError, TypeError):
                pass

        # Color code Link Type column
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            header = ws.cell(row=1, column=col).value
            if header == "Link Type":
                if str(cell.value).strip() == "DoFollow":
                    cell.font = dofollow_font
                    cell.value = "DoFollow"
                elif str(cell.value).strip() == "NoFollow":
                    cell.font = nofollow_font
                    cell.value = "NoFollow"
            elif header == "URL":
                cell.font = link_font


def auto_width(ws, num_cols):
    """Auto-adjust column widths."""
    for col in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def add_autofilter(ws, num_cols):
    """Add auto-filter to header row."""
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col}1"


def create_summary_sheet(writer, df):
    """Create a summary/dashboard sheet."""
    da = pd.to_numeric(df["Domain Authority"], errors="coerce").fillna(0)

    rows = [
        ("Total Sites", len(df), "100%"),
        ("", "", ""),
        ("BY LINK TYPE", "", ""),
        ("DoFollow Sites", len(df[df["Link Type"] == "DoFollow"]), f"{len(df[df['Link Type'] == 'DoFollow'])/len(df)*100:.0f}%"),
        ("NoFollow Sites", len(df[df["Link Type"] == "NoFollow"]), f"{len(df[df['Link Type'] == 'NoFollow'])/len(df)*100:.0f}%"),
        ("", "", ""),
        ("BY DOMAIN AUTHORITY", "", ""),
        ("DA 90+ (Highest Authority)", int((da >= 90).sum()), f"{(da >= 90).sum()/len(df)*100:.0f}%"),
        ("DA 80-89 (Very High)", int(((da >= 80) & (da < 90)).sum()), f"{((da >= 80) & (da < 90)).sum()/len(df)*100:.0f}%"),
        ("DA 70-79 (High)", int(((da >= 70) & (da < 80)).sum()), f"{((da >= 70) & (da < 80)).sum()/len(df)*100:.0f}%"),
        ("DA 50-69 (Medium)", int(((da >= 50) & (da < 70)).sum()), f"{((da >= 50) & (da < 70)).sum()/len(df)*100:.0f}%"),
        ("DA 40-49 (Low-Medium)", int(((da >= 40) & (da < 50)).sum()), f"{((da >= 40) & (da < 50)).sum()/len(df)*100:.0f}%"),
        ("DA < 40 (Low)", int((da < 40).sum()), f"{(da < 40).sum()/len(df)*100:.0f}%"),
        ("", "", ""),
        ("BY TYPE", "", ""),
        ("Web 2.0 (Create Blog)", len(df[df["Type"] == "Web 2.0"]), ""),
        ("Article Directory", len(df[df["Type"] == "Article Directory"]), ""),
        ("Guest Post", len(df[df["Type"] == "Guest Post"]), ""),
        ("Blog Platform", len(df[df["Type"] == "Blog Platform"]), ""),
        ("Blog Directory", len(df[df["Type"] == "Blog Directory"]), ""),
        ("Content Curation", len(df[df["Type"] == "Content Curation"]), ""),
        ("Social Bookmarking", len(df[df["Type"] == "Social Bookmarking"]), ""),
        ("Social Platform", len(df[df["Type"] == "Social Platform"]), ""),
        ("Q&A Platform", len(df[df["Type"] == "Q&A Platform"]), ""),
        ("", "", ""),
        ("BY COUNTRY", "", ""),
        ("Global", len(df[df["Country"] == "Global"]), ""),
        ("India", len(df[df["Country"] == "India"]), ""),
        ("US", len(df[df["Country"] == "US"]), ""),
        ("Other", len(df[~df["Country"].isin(["Global", "India", "US"])]), ""),
    ]

    summary_df = pd.DataFrame(rows, columns=["Metric", "Count", "Percentage"])
    summary_df.to_excel(writer, sheet_name="Dashboard", index=False)


def main():
    print("\nExporting Blog Listing Sites for Google Sheets...\n")

    # Read data
    if not os.path.exists(INPUT_CSV):
        print(f"ERROR: {INPUT_CSV} not found. Run compile_blog_listings.py first.")
        return

    df = pd.read_csv(INPUT_CSV)
    print(f"  Loaded {len(df)} sites from {INPUT_CSV}")

    # Sort by DA
    df["_da"] = pd.to_numeric(df["Domain Authority"], errors="coerce").fillna(0)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:

        # Sheet 1: Dashboard
        create_summary_sheet(writer, df)
        print("  Sheet 1: Dashboard")

        # Sheet 2: All Sites (sorted by DA)
        all_df = df.sort_values("_da", ascending=False).drop(columns=["_da"]).reset_index(drop=True)
        all_df["S.No"] = range(1, len(all_df) + 1)
        cols = ["S.No", "Website Name", "URL", "Domain Authority", "Type", "Link Type", "Category", "Country"]
        all_df[cols].to_excel(writer, sheet_name="All Sites", index=False)
        print(f"  Sheet 2: All Sites ({len(all_df)} rows)")

        # Sheet 3: DoFollow Only (the most valuable for SEO)
        dofollow = df[df["Link Type"] == "DoFollow"].sort_values("_da", ascending=False).drop(columns=["_da"]).reset_index(drop=True)
        dofollow["S.No"] = range(1, len(dofollow) + 1)
        dofollow[cols].to_excel(writer, sheet_name="DoFollow Sites", index=False)
        print(f"  Sheet 3: DoFollow Sites ({len(dofollow)} rows)")

        # Sheet 4: High DA (70+)
        high_da = df[df["_da"] >= 70].sort_values("_da", ascending=False).drop(columns=["_da"]).reset_index(drop=True)
        high_da["S.No"] = range(1, len(high_da) + 1)
        high_da[cols].to_excel(writer, sheet_name="High DA (70+)", index=False)
        print(f"  Sheet 4: High DA 70+ ({len(high_da)} rows)")

        # Sheet 5: Web 2.0 Sites
        web20 = df[df["Type"] == "Web 2.0"].sort_values("_da", ascending=False).drop(columns=["_da"]).reset_index(drop=True)
        web20["S.No"] = range(1, len(web20) + 1)
        web20[cols].to_excel(writer, sheet_name="Web 2.0 Sites", index=False)
        print(f"  Sheet 5: Web 2.0 Sites ({len(web20)} rows)")

        # Sheet 6: Article Directories
        articles = df[df["Type"] == "Article Directory"].sort_values("_da", ascending=False).drop(columns=["_da"]).reset_index(drop=True)
        articles["S.No"] = range(1, len(articles) + 1)
        articles[cols].to_excel(writer, sheet_name="Article Directories", index=False)
        print(f"  Sheet 6: Article Directories ({len(articles)} rows)")

        # Sheet 7: Guest Post Sites
        guest = df[df["Type"] == "Guest Post"].sort_values("_da", ascending=False).drop(columns=["_da"]).reset_index(drop=True)
        guest["S.No"] = range(1, len(guest) + 1)
        guest[cols].to_excel(writer, sheet_name="Guest Post Sites", index=False)
        print(f"  Sheet 7: Guest Post Sites ({len(guest)} rows)")

        # Sheet 8: Blog Platforms
        blogs = df[df["Type"].isin(["Blog Platform", "Blog Directory"])].sort_values("_da", ascending=False).drop(columns=["_da"]).reset_index(drop=True)
        blogs["S.No"] = range(1, len(blogs) + 1)
        blogs[cols].to_excel(writer, sheet_name="Blog Platforms", index=False)
        print(f"  Sheet 8: Blog Platforms ({len(blogs)} rows)")

        # Sheet 9: India Focused
        india = df[df["Country"] == "India"].sort_values("_da", ascending=False).drop(columns=["_da"]).reset_index(drop=True)
        india["S.No"] = range(1, len(india) + 1)
        india[cols].to_excel(writer, sheet_name="India Sites", index=False)
        print(f"  Sheet 9: India Sites ({len(india)} rows)")

    # Apply formatting
    print("\n  Applying formatting...")
    wb = load_workbook(OUTPUT_XLSX)

    # Format Dashboard
    ws = wb["Dashboard"]
    style_header(ws, 3)
    auto_width(ws, 3)

    # Section headers in dashboard
    section_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    section_font = Font(name="Arial", size=11, bold=True, color="1A73E8")
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        val = str(cell.value or "")
        if val.startswith("BY ") or val.startswith("TOP "):
            for c in range(1, 4):
                ws.cell(row=cell.row, column=c).fill = section_fill
                ws.cell(row=cell.row, column=c).font = section_font

    # Format data sheets
    data_sheets = [
        "All Sites", "DoFollow Sites", "High DA (70+)",
        "Web 2.0 Sites", "Article Directories", "Guest Post Sites",
        "Blog Platforms", "India Sites"
    ]

    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        num_cols = ws.max_column
        num_rows = ws.max_row - 1
        style_header(ws, num_cols)
        style_data_rows(ws, num_rows, num_cols, da_col=4)  # DA is column 4
        auto_width(ws, num_cols)
        add_autofilter(ws, num_cols)

        # Freeze top row
        ws.freeze_panes = "A2"

    wb.save(OUTPUT_XLSX)

    print(f"\n  File saved: {OUTPUT_XLSX}")
    print(f"  Size: {os.path.getsize(OUTPUT_XLSX) / 1024:.1f} KB")

    print(f"\n  HOW TO UPLOAD TO GOOGLE SHEETS:")
    print(f"  1. Go to sheets.google.com")
    print(f"  2. Click 'Blank spreadsheet' or open existing")
    print(f"  3. File > Import > Upload")
    print(f"  4. Select: {os.path.abspath(OUTPUT_XLSX)}")
    print(f"  5. Choose 'Replace spreadsheet' > Import")
    print()

    # Auto-open
    import platform as plat
    filepath = os.path.abspath(OUTPUT_XLSX)
    try:
        if plat.system() == "Windows":
            os.startfile(filepath)
    except Exception:
        pass


if __name__ == "__main__":
    main()
