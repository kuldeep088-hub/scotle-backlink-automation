"""
Build master backlinks Excel report combining all published posts.
"""
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

OUTPUT = "data/Scotle_ALL_Backlinks_Master.xlsx"

# ---- Load all sources ----

# Source 1: mega_post.py run (20 posts - all scotle.org)
df_mega = pd.read_excel('data/Scotle_30_Blogs_Report.xlsx', sheet_name='All Published Posts')
src1 = []
for _, r in df_mega.iterrows():
    src1.append({
        'Platform': r['platform'], 'DA': r['da'],
        'Article URL': r['url'], 'Title': r['title'],
        'Anchor Text': r.get('anchor', ''), 'Word Count': r.get('word_count', 0),
        'Target URL': 'https://scotle.org/', 'Source': 'Run-3 (mega_post)'
    })

# Source 2: post_30_blogs.py run (10 posts - all scotle.org via FriendPaste)
df_30 = pd.read_excel('data/Scotle_30_Blog_Posts.xlsx', sheet_name='Published Posts')
src2 = []
for _, r in df_30.iterrows():
    src2.append({
        'Platform': r['Platform'], 'DA': r['DA'],
        'Article URL': r['Article URL'], 'Title': r['Title'],
        'Anchor Text': r.get('Anchor Text', ''), 'Word Count': r.get('Word Count', 0),
        'Target URL': 'https://scotle.org/', 'Source': 'Run-2 (post_30_blogs)'
    })

# Source 3: bulk_post.py - only scotle.org targeted posts
df_old = pd.read_excel('data/Scotle_Blog_Backlinks_Report.xlsx', sheet_name='Live Posts')
src3 = []
for _, r in df_old.iterrows():
    url = str(r.get('Target URL', ''))
    if 'scotle.org' in url:
        src3.append({
            'Platform': 'Write.as', 'DA': 69,
            'Article URL': r['Published URL'], 'Title': r['Title'],
            'Anchor Text': r.get('Anchor Text', ''), 'Word Count': r.get('Word Count', 0),
            'Target URL': 'https://scotle.org/', 'Source': 'Run-1 (bulk_post)'
        })

all_posts = src1 + src2 + src3

# Deduplicate by URL
seen = set()
unique = []
for p in all_posts:
    if p['Article URL'] not in seen:
        seen.add(p['Article URL'])
        unique.append(p)

df_all = pd.DataFrame(unique)
df_all.insert(0, 'S.No', range(1, len(df_all) + 1))

print(f"Total unique published posts: {len(df_all)}")
print(df_all.groupby('Platform')[['DA']].count().rename(columns={'DA': 'Count'}))

# ---- Platform stats ----
stats = df_all.groupby(['Platform', 'DA']).size().reset_index(name='Posts')
stats['Backlinks'] = stats['Posts'] * 3
stats = stats.sort_values('DA', ascending=False)

# ---- Write Excel ----
with pd.ExcelWriter(OUTPUT, engine='openpyxl') as writer:
    # Sheet 1: Dashboard
    rows = [
        ('SCOTLE.ORG - BACKLINK CAMPAIGN REPORT', ''),
        ('', ''),
        ('Target Website', 'https://scotle.org/'),
        ('Brand', 'Scotle High School, Jaipur'),
        ('Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('', ''),
        ('== RESULTS ==', ''),
        ('', ''),
        ('Total Blog Posts Published', len(df_all)),
        ('Total Backlinks Created', len(df_all) * 3),
        ('Platforms Used', df_all['Platform'].nunique()),
        ('Average DA', round(float(df_all['DA'].mean()), 1)),
        ('', ''),
        ('== PLATFORM BREAKDOWN ==', ''),
    ]
    for _, row in stats.iterrows():
        rows.append((
            f"  {row['Platform']} (DA {row['DA']})",
            f"{row['Posts']} posts = {row['Backlinks']} backlinks"
        ))
    dash = pd.DataFrame(rows, columns=['Metric', 'Value'])
    dash.to_excel(writer, sheet_name='Dashboard', index=False)

    # Sheet 2: All Posts
    df_all.to_excel(writer, sheet_name='All Published Posts', index=False)

    # Sheet 3: Write.as
    was = df_all[df_all['Platform'] == 'Write.as'].reset_index(drop=True)
    was.insert(0, '#', range(1, len(was) + 1))
    was.drop('S.No', axis=1, errors='ignore').to_excel(writer, sheet_name='Write.as Posts (DA 69)', index=False)

    # Sheet 4: FriendPaste
    fp = df_all[df_all['Platform'] == 'FriendPaste'].reset_index(drop=True)
    fp.insert(0, '#', range(1, len(fp) + 1))
    fp.drop('S.No', axis=1, errors='ignore').to_excel(writer, sheet_name='FriendPaste Posts (DA 42)', index=False)

    # Sheet 5: Catbox
    cb = df_all[df_all['Platform'] == 'Catbox.moe'].reset_index(drop=True)
    cb.insert(0, '#', range(1, len(cb) + 1))
    cb.drop('S.No', axis=1, errors='ignore').to_excel(writer, sheet_name='Catbox Posts (DA 55)', index=False)

    # Sheet 6: Quick URL list
    urls = df_all[['S.No', 'Platform', 'DA', 'Article URL', 'Title', 'Target URL']].copy()
    urls.to_excel(writer, sheet_name='Quick URL List', index=False)

    # Sheet 7: Stats
    stats.to_excel(writer, sheet_name='Platform Stats', index=False)

# ---- Format ----
wb = load_workbook(OUTPUT)
h_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
h_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
g_fill = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')
g_font = Font(name='Arial', size=10, bold=True, color='137333')
link_font = Font(name='Arial', size=10, color='1558D6', underline='single')
title_font = Font(name='Arial', size=14, bold=True, color='1A73E8')
y_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')

for sn in wb.sheetnames:
    ws = wb[sn]
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            cell = ws.cell(row=row, column=col)
            val = str(cell.value or '')

            if sn == 'Dashboard' and col == 1 and 'SCOTLE' in val:
                cell.font = title_font
                cell.fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')

            if header in ('Article URL', 'Target URL', 'Published URL') and val.startswith('http'):
                cell.font = link_font

            if header == 'Platform':
                if 'Write.as' in val:
                    cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                    cell.font = Font(name='Arial', size=10, color='1B5E20', bold=True)
                elif 'FriendPaste' in val:
                    cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
                    cell.font = Font(name='Arial', size=10, color='E65100', bold=True)
                elif 'Catbox' in val:
                    cell.fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
                    cell.font = Font(name='Arial', size=10, color='6A1B9A', bold=True)

            if header == 'DA':
                try:
                    da = int(cell.value)
                    if da >= 65:
                        cell.fill = g_fill
                        cell.font = g_font
                    elif da >= 50:
                        cell.fill = y_fill
                except Exception:
                    pass

    for col in range(1, ws.max_column + 1):
        ml = max((len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions[get_column_letter(col)].width = min(ml + 3, 65)
    ws.freeze_panes = 'A2'

ws_dash = wb['Dashboard']
ws_dash.column_dimensions['A'].width = 38
ws_dash.column_dimensions['B'].width = 52

wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print(f"Size: {os.path.getsize(OUTPUT) / 1024:.1f} KB")
