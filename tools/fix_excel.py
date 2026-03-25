"""
Replaces the save_excel() function in big_poster.py with an improved version.
Run once: python fix_excel.py
"""
import os

NEW_FUNC = '''def save_excel(results, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    import datetime as _dt

    wb  = Workbook()
    pub = [r for r in results if r["Status"] == "Published"]
    fai = [r for r in results if r["Status"] == "Failed"]

    DA_HIGH = "C8E6C9"; DA_MED = "DCEDC8"; DA_LOW = "FFF9C4"; DA_VLOW = "F5F5F5"
    HDR_GREEN = "0D652D"; HDR_BLUE = "1A73E8"; LINK = "1558D6"
    ROW_PUB = "E8F5E9"; ROW_FAIL = "FFEBEE"

    def da_color(da):
        if da >= 70: return DA_HIGH
        if da >= 50: return DA_MED
        if da >= 30: return DA_LOW
        return DA_VLOW

    def wc_color(wc):
        if wc >= 500: return "C8E6C9"
        if wc >= 300: return "FFF9C4"
        return "FFCCBC"

    thin = Side(style="thin", color="DDDDDD")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_hdr(ws, headers, fill_hex, txt="FFFFFF", sz=11):
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h)
            c.font = Font(bold=True, color=txt, size=sz)
            c.fill = PatternFill("solid", fgColor=fill_hex)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    def hyperlink(ws, row, col, url, label=None):
        c = ws.cell(row, col, label or url)
        if url and url != "N/A":
            c.hyperlink = url
            c.font = Font(color=LINK, underline="single")

    # ── Sheet 1: Published Links ──────────────────────────────────
    ws = wb.active
    ws.title = "Published Links"
    cols1 = ["#", "Platform", "DA", "DA Band", "Title", "Live URL",
             "Anchor Text", "Words", "Date"]
    set_hdr(ws, cols1, HDR_GREEN)
    for i, r in enumerate(pub, 1):
        da  = r["DA"]
        wc  = r.get("Word Count") or 0
        band = ("DA 70+" if da >= 70 else "DA 50-69" if da >= 50
                else "DA 30-49" if da >= 30 else "DA <30")
        ws.append([i, r["Platform"], da, band, r["Title"],
                   r["Article URL"], r["Anchor Text"], wc, r["Date"]])
        row = i + 1
        rf  = PatternFill("solid", fgColor=da_color(da))
        for col in range(1, len(cols1) + 1):
            ws.cell(row, col).fill = rf
            ws.cell(row, col).border = brd
            ws.cell(row, col).alignment = Alignment(vertical="center")
        hyperlink(ws, row, 6, r["Article URL"])
        ws.cell(row, 8).fill = PatternFill("solid", fgColor=wc_color(wc))
        ws.cell(row, 3).alignment = Alignment(horizontal="center")
        ws.cell(row, 4).alignment = Alignment(horizontal="center")
    for col, w in zip(range(1, 10), [5, 22, 6, 10, 54, 58, 22, 8, 16]):
        ws.column_dimensions[get_column_letter(col)].width = w
    # Legend row
    lr = ws.max_row + 2
    ws.cell(lr, 1, "Legend:").font = Font(bold=True)
    for off, (lbl, clr) in enumerate([
        ("DA 70+", DA_HIGH), ("DA 50-69", DA_MED), ("DA 30-49", DA_LOW),
        ("Words 500+", "C8E6C9"), ("Words <300", "FFCCBC")
    ], 2):
        c = ws.cell(lr, off, lbl)
        c.fill = PatternFill("solid", fgColor=clr)
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=9)

    # ── Sheet 2: All Results ─────────────────────────────────────
    wa = wb.create_sheet("All Results")
    cols2 = ["#", "Platform", "DA", "Status", "Title", "Article URL",
             "Words", "Error Message", "Date"]
    set_hdr(wa, cols2, HDR_BLUE)
    for i, r in enumerate(results, 1):
        is_pub = r["Status"] == "Published"
        wc  = r.get("Word Count") or 0
        url = r["Article URL"] or "N/A"
        wa.append([i, r["Platform"], r["DA"], r["Status"],
                   r["Title"], url, wc, r.get("Error", "") or "", r["Date"]])
        row = i + 1
        rf  = PatternFill("solid", fgColor=ROW_PUB if is_pub else ROW_FAIL)
        for col in range(1, len(cols2) + 1):
            wa.cell(row, col).fill = rf
            wa.cell(row, col).border = brd
            wa.cell(row, col).alignment = Alignment(vertical="center")
        wa.cell(row, 4).font = Font(bold=True,
                                    color="2E7D32" if is_pub else "C62828")
        wa.cell(row, 4).alignment = Alignment(horizontal="center")
        wa.cell(row, 3).alignment = Alignment(horizontal="center")
        if is_pub and r["Article URL"]:
            hyperlink(wa, row, 6, r["Article URL"])
    for col, w in zip(range(1, 10), [5, 22, 6, 11, 52, 58, 8, 42, 16]):
        wa.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: DA Breakdown + Bar Chart ───────────────────────
    wd = wb.create_sheet("DA Breakdown")
    wd.merge_cells("A1:D1")
    tc = wd.cell(1, 1, "Published Backlinks by DA Range")
    tc.font = Font(bold=True, size=13, color=HDR_BLUE)
    tc.alignment = Alignment(horizontal="center")
    wd.row_dimensions[1].height = 24
    wd.append([])
    wd.append(["DA Range", "Platforms", "Links", "Avg Words"])
    for col in range(1, 5):
        c = wd.cell(3, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR_BLUE)
        c.alignment = Alignment(horizontal="center")

    clr_map = {"DA 70+": DA_HIGH, "DA 50-69": DA_MED,
               "DA 30-49": DA_LOW, "DA <30": DA_VLOW}
    bands = [("DA 70+", 70, 999), ("DA 50-69", 50, 69),
             ("DA 30-49", 30, 49), ("DA <30", 0, 29)]
    for label, lo, hi in bands:
        grp = [r for r in pub if lo <= r["DA"] <= hi]
        wcs = [r.get("Word Count") or 0 for r in grp]
        avg_wc = int(sum(wcs) / len(wcs)) if wcs else 0
        wd.append([label, len(set(r["Platform"] for r in grp)), len(grp), avg_wc])
        row = wd.max_row
        for col in range(1, 5):
            wd.cell(row, col).fill = PatternFill("solid", fgColor=clr_map[label])
            wd.cell(row, col).border = brd
            wd.cell(row, col).alignment = Alignment(horizontal="center")

    chart = BarChart()
    chart.type = "col"
    chart.title = "Links Published by DA Range"
    chart.y_axis.title = "Links"
    chart.x_axis.title = "DA Range"
    chart.style = 10; chart.width = 18; chart.height = 12
    chart.add_data(Reference(wd, min_col=3, max_col=3, min_row=3, max_row=7),
                   titles_from_data=True)
    chart.set_categories(Reference(wd, min_col=1, min_row=4, max_row=7))
    wd.add_chart(chart, "F3")
    for col, w in zip(range(1, 5), [14, 14, 10, 13]):
        wd.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 4: Summary ────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 55

    ws2.merge_cells("A1:B1")
    hc = ws2.cell(1, 1, "Scotle.org Backlink Campaign Report")
    hc.font = Font(bold=True, size=16, color=HDR_BLUE)
    hc.alignment = Alignment(horizontal="center")
    hc.fill = PatternFill("solid", fgColor="E3F2FD")
    ws2.row_dimensions[1].height = 30

    row = 2
    ws2.cell(row, 1, "Generated:").font = Font(bold=True)
    ws2.cell(row, 2, _dt.datetime.now().strftime("%Y-%m-%d %H:%M"))

    row += 2
    ws2.cell(row, 1, "CAMPAIGN STATS").font = Font(bold=True, size=12, color=HDR_BLUE)
    plats_tried = len(set(r["Platform"] for r in results))
    plats_ok    = len(set(r["Platform"] for r in pub))
    total_wc    = sum(r.get("Word Count") or 0 for r in pub)
    stats = [
        ("Platforms Tried",         plats_tried),
        ("Platforms That Worked",   plats_ok),
        ("Total Links Published",   len(pub)),
        ("Total Attempts",          len(results)),
        ("Platform Success Rate",   f"{plats_ok / max(plats_tried, 1) * 100:.0f}%"),
        ("Avg DA of Published Links", f"{sum(r['DA'] for r in pub) // max(len(pub), 1)}"),
        ("Total Words Published",   f"{total_wc:,}"),
        ("Avg Words per Post",      f"{total_wc // max(len(pub), 1)}"),
    ]
    for label, value in stats:
        row += 1
        ws2.cell(row, 1, label).font = Font(bold=True)
        ws2.cell(row, 2, value)

    row += 2
    ws2.cell(row, 1, "UNLOCK MORE PLATFORMS").font = Font(bold=True, size=12, color="B71C1C")
    unlock = [
        ("Medium (DA 95)",   "medium.com/me/settings > Integration tokens"),
        ("Dev.to (DA 90)",   "dev.to/settings/extensions > API Keys"),
        ("Tumblr (DA 81)",   "api.tumblr.com/console > OAuth keys"),
        ("Hashnode (DA 79)", "hashnode.com/settings/developer > Token + Publication ID"),
    ]
    for label, where in unlock:
        row += 1
        ws2.cell(row, 1, label).font = Font(bold=True, color="B71C1C")
        ws2.cell(row, 2, where)

    row += 2
    ws2.cell(row, 1, "ALL PUBLISHED LINKS").font = Font(bold=True, size=12, color=HDR_BLUE)
    row += 1
    ws2.cell(row, 1, "Platform (DA)").font = Font(bold=True, underline="single")
    ws2.cell(row, 2, "Live URL").font    = Font(bold=True, underline="single")
    last_plat = None
    for entry in sorted(pub, key=lambda x: (-x["DA"], x["Platform"])):
        row += 1
        if entry["Platform"] != last_plat:
            ws2.cell(row, 1, f"{entry['Platform']}  (DA {entry['DA']})").font = Font(bold=True)
            last_plat = entry["Platform"]
        else:
            ws2.cell(row, 1, "")
        hyperlink(ws2, row, 2, entry["Article URL"])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    logger.info(f"Excel saved: {path}")
'''

# Read the file as bytes to handle any special characters
with open("big_poster.py", "rb") as f:
    content = f.read()

func_start = content.find(b"def save_excel(")
func_end   = content.find(b"\ndef main()", func_start)

if func_start == -1 or func_end == -1:
    print("ERROR: Could not find function boundaries")
else:
    new_content = content[:func_start] + NEW_FUNC.encode("utf-8") + b"\n" + content[func_end:]
    with open("big_poster.py", "wb") as f:
        f.write(new_content)
    print(f"Done. Replaced save_excel() — file size: {len(new_content):,} bytes")
