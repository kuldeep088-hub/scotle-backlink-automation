"""Save the 30 live blog posts to Excel."""
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_XLSX = "data/Scotle_30_Best.xlsx"
TARGET_URL  = "https://scotle.org/"

pub = [
    {"S.No": 1,  "platform": "Ubuntu Paste", "da": 90, "url": "https://paste.ubuntu.com/p/5C98fnZsxc/",             "title": "The Real Reason Children Stop Trying in Maths"},
    {"S.No": 2,  "platform": "Write.as",      "da": 69, "url": "https://write.as/dg4a9eh5gspec",                    "title": "Why Speed Tests Are Not Measuring What You Think They Are"},
    {"S.No": 3,  "platform": "Catbox.moe",    "da": 55, "url": "https://files.catbox.moe/3mctuc.html",              "title": "Why Some Children Shut Down in Maths"},
    {"S.No": 4,  "platform": "Filebin.net",   "da": 55, "url": "https://filebin.net/mlil9pk4pk37qu3l",              "title": "The Question Every Parent Should Ask About Their Childs Maths Class"},
    {"S.No": 5,  "platform": "paste.ee",      "da": 52, "url": "https://pastee.dev/p/cvs0wHva",                     "title": "Your Child Is Not Bad at Maths"},
    {"S.No": 6,  "platform": "cl1p.net",      "da": 50, "url": "https://cl1p.net/rn4hiypqd73m2q",                  "title": "A Small Classroom Moment That Explained Math Anxiety"},
    {"S.No": 7,  "platform": "Bytebin",       "da": 50, "url": "https://bytebin.lucko.me/lyjJsWBrKr",              "title": "The Real Reason Children Stop Trying in Maths"},
    {"S.No": 8,  "platform": "just-paste.it", "da": 45, "url": "https://just-paste.it/0jNSKYoSwo",                  "title": "Maths Is Not a Talent. It Is a Skill Built With the Right Environment."},
    {"S.No": 9,  "platform": "FriendPaste",   "da": 42, "url": "https://friendpaste.com/4661NCeSifuoakMhj7HqXt",   "title": "What Primary School Maths Should Actually Be Teaching"},
    {"S.No": 10, "platform": "Ubuntu Paste",  "da": 90, "url": "https://paste.ubuntu.com/p/HZgn9pxZG6/",           "title": "Your Child Is Not Bad at Maths"},
    {"S.No": 11, "platform": "Write.as",      "da": 69, "url": "https://write.as/qqigc36tsedjs",                    "title": "Why Comparing Children in Maths Is Quietly Damaging Their Confidence"},
    {"S.No": 12, "platform": "Catbox.moe",    "da": 55, "url": "https://files.catbox.moe/24ucxj.html",             "title": "Why Speed Tests Are Not Measuring What You Think They Are"},
    {"S.No": 13, "platform": "Filebin.net",   "da": 55, "url": "https://filebin.net/qs5mbo5b14v9tfdp",             "title": "Maths Is Not a Talent. It Is a Skill Built With the Right Environment."},
    {"S.No": 14, "platform": "paste.ee",      "da": 52, "url": "https://pastee.dev/p/bDqhKOji",                    "title": "Why Some Children Shut Down in Maths"},
    {"S.No": 15, "platform": "cl1p.net",      "da": 50, "url": "https://cl1p.net/hdj01uvnea04zi",                  "title": "Why Homework Is Not Fixing Your Childs Maths Problem"},
    {"S.No": 16, "platform": "Bytebin",       "da": 50, "url": "https://bytebin.lucko.me/25ZHS2lFH3",              "title": "Your Child Is Not Bad at Maths"},
    {"S.No": 17, "platform": "just-paste.it", "da": 45, "url": "https://just-paste.it/XSwrmWFgh4",                 "title": "A Small Classroom Moment That Explained Math Anxiety"},
    {"S.No": 18, "platform": "FriendPaste",   "da": 42, "url": "https://friendpaste.com/4661NCeSifuoakMhj7HpMl",  "title": "The Real Reason Children Stop Trying in Maths"},
    {"S.No": 19, "platform": "Ubuntu Paste",  "da": 90, "url": "https://paste.ubuntu.com/p/kwJnJdhHth/",           "title": "Why Some Children Shut Down in Maths"},
    {"S.No": 20, "platform": "Write.as",      "da": 69, "url": "https://write.as/m0exmxeijobvs",                   "title": "The Question Every Parent Should Ask About Their Childs Maths Class"},
    {"S.No": 21, "platform": "Catbox.moe",    "da": 55, "url": "https://files.catbox.moe/oyea9z.html",             "title": "Why Comparing Children in Maths Is Quietly Damaging Their Confidence"},
    {"S.No": 22, "platform": "Filebin.net",   "da": 55, "url": "https://filebin.net/ddbws9ajn8yguv8r",             "title": "A Small Classroom Moment That Explained Math Anxiety"},
    {"S.No": 23, "platform": "paste.ee",      "da": 52, "url": "https://pastee.dev/p/2ZrykW5d",                    "title": "Why Speed Tests Are Not Measuring What You Think They Are"},
    {"S.No": 24, "platform": "cl1p.net",      "da": 50, "url": "https://cl1p.net/egpj59xh8epa3u",                  "title": "What Primary School Maths Should Actually Be Teaching"},
    {"S.No": 25, "platform": "Bytebin",       "da": 50, "url": "https://bytebin.lucko.me/E3N0qCL7HK",             "title": "Why Some Children Shut Down in Maths"},
    {"S.No": 26, "platform": "just-paste.it", "da": 45, "url": "https://just-paste.it/ov58BI5q2N",                 "title": "Why Homework Is Not Fixing Your Childs Maths Problem"},
    {"S.No": 27, "platform": "FriendPaste",   "da": 42, "url": "https://friendpaste.com/3kAFXskTg6zxkNEkVTcTJV",  "title": "Your Child Is Not Bad at Maths"},
    {"S.No": 28, "platform": "Ubuntu Paste",  "da": 90, "url": "https://paste.ubuntu.com/p/JZKtpHryjS/",          "title": "Why Comparing Children in Maths Is Quietly Damaging Their Confidence"},
    {"S.No": 29, "platform": "Write.as",      "da": 69, "url": "https://write.as/hilix8t3hfs3m",                   "title": "A Small Classroom Moment That Explained Math Anxiety"},
    {"S.No": 30, "platform": "Catbox.moe",    "da": 55, "url": "https://files.catbox.moe/mxe5d6.html",             "title": "The Real Reason Children Stop Trying in Maths"},
]

avg_da = sum(r["da"] for r in pub) // len(pub)

summary_rows = [
    ("SCOTLE.ORG -- 30 BEST SITES BLOG REPORT", ""),
    ("", ""),
    ("Target URL",       TARGET_URL),
    ("Blog Style",       "Scotle High School -- Maths Education"),
    ("Report Date",      datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("", ""),
    ("Total Published",  30),
    ("Unique Platforms", len(set(r["platform"] for r in pub))),
    ("Average DA",       avg_da),
    ("Highest DA",       90),
    ("Backlinks Created","30 (1 per post -- CTA end only)"),
    ("", ""),
    ("ALL LIVE BACKLINKS (sorted by DA)", ""),
]
for r in sorted(pub, key=lambda x: (-x["da"], x["S.No"])):
    summary_rows.append((f"  {r['platform']}  (DA {r['da']})", r["url"]))

pub_df = pd.DataFrame(pub)

with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as w:
    pd.DataFrame(summary_rows, columns=["Metric", "Value"]).to_excel(
        w, sheet_name="Dashboard", index=False)
    pub_df.to_excel(w, sheet_name="All 30 Posts", index=False)
    pub_df[["S.No","platform","da","url","title"]].to_excel(
        w, sheet_name="Quick URL List", index=False)

wb    = load_workbook(OUTPUT_XLSX)
hfill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
lfont = Font(name="Arial", size=10, color="1565C0", underline="single")
gfill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
gfont = Font(name="Arial", size=10, bold=True, color="1B5E20")
yfill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

for sn in wb.sheetnames:
    ws = wb[sn]
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).fill = hfill
        ws.cell(row=1, column=c).font = hfont
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            h    = ws.cell(row=1, column=col).value
            cell = ws.cell(row=row, column=col)
            val  = str(cell.value or "")
            if h in ("url", "Value") and val.startswith("http"):
                cell.font = lfont
            if h == "da":
                try:
                    da = int(cell.value)
                    if da >= 65:
                        cell.fill = gfill
                        cell.font = gfont
                    elif da >= 50:
                        cell.fill = yfill
                except Exception:
                    pass
    for c in range(1, ws.max_column + 1):
        ml = max(len(str(ws.cell(row=r, column=c).value or ""))
                 for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(c)].width = min(ml + 4, 72)
    ws.freeze_panes = "A2"

wb.save(OUTPUT_XLSX)
print(f"Excel saved -> {OUTPUT_XLSX}")
print(f"Rows: {len(pub)} | Avg DA: {avg_da} | Platforms: {len(set(r['platform'] for r in pub))}")
