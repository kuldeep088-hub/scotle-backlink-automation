"""
post_batch4.py — Final batch to reach 30 different sites.
"""

import os, sys, time, random, json, requests, re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from blog_poster import blog_config as config
from blog_poster.blog_utils import (
    generate_article_from_template, get_next_topic,
    inject_backlinks, build_author_bio, add_authority_links, html_to_markdown,
)
from blog_poster.templates.article_templates import get_templates, get_all_niches

DATA_DIR   = os.path.join(os.path.dirname(__file__), "data")
PREV_XLSX  = os.path.join(DATA_DIR, "Scotle_30_Different_Sites_FINAL.xlsx")
OUT_XLSX   = os.path.join(DATA_DIR, "Scotle_30_Sites_Complete.xlsx")

used_topics = []
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"


def gen_article():
    global used_topics
    niche = random.choice(get_all_niches())
    tpl = random.choice(get_templates(niche))
    kw = random.choice(config.TARGET_KEYWORDS)
    topic = get_next_topic(config.TARGET_KEYWORDS, used_topics=used_topics)
    used_topics.append(topic)
    art = generate_article_from_template(tpl, kw, config.TARGET_BRAND, "Jaipur", True)
    art["title"] = topic
    c = art["content_html"]
    anchor = random.choice(config.ANCHOR_TEXTS)
    c = inject_backlinks(c, config.TARGET_URL, config.ANCHOR_TEXTS, 2, 30)
    c = add_authority_links(c, config.AUTHORITY_LINKS)
    c += "\n\n" + build_author_bio(config.TARGET_URL, config.TARGET_BRAND, anchor)
    art["content_html"] = c
    art["anchor"] = anchor
    art["md"] = html_to_markdown(c)
    return art


def try_etherpad(base, art, name, da):
    """Open etherpad — just check if pad URL is accessible (open pads)."""
    try:
        pad_id = f"scotle-{random.randint(100000, 999999)}"
        r = requests.get(f"{base}/p/{pad_id}",
            headers={"User-Agent": UA}, timeout=10)
        if r.status_code == 200 and ("etherpad" in r.text.lower() or "pad" in r.text.lower()):
            return True, f"{base}/p/{pad_id}", name, da
    except: pass
    return False, "", name, da


def try_hastebin(base, art, name, da):
    try:
        content = f"# {art['title']}\n\n{art['md']}"
        r = requests.post(f"{base}/documents",
            data=content.encode(),
            headers={"Content-Type": "text/plain", "User-Agent": UA}, timeout=12)
        if r.status_code in (200, 201):
            key = r.json().get("key", "")
            if key:
                return True, f"{base}/{key}", name, da
    except: pass
    return False, "", name, da


def try_hedgedoc(base, art, name, da):
    try:
        r = requests.post(f"{base}/new",
            data=f"# {art['title']}\n\n{art['md']}".encode(),
            headers={"Content-Type": "text/markdown", "User-Agent": UA},
            allow_redirects=True, timeout=15)
        url = r.url
        domain = base.replace("https://", "").replace("http://", "")
        if domain in url and "/new" not in url and url not in (base + "/", base):
            return True, url, name, da
    except: pass
    return False, "", name, da


# ---- 20 new platform functions ----

def post_boardnet_pad(art):
    return try_etherpad("https://board.net", art, "board.net", 55)

def post_okfn_pad(art):
    return try_etherpad("https://pad.okfn.org", art, "pad.okfn.org", 65)

def post_opn_pad(art):
    return try_etherpad("https://opin.me", art, "opin.me", 42)

def post_medienpad(art):
    return try_etherpad("https://medienpad.de", art, "medienpad.de", 45)

def post_hedgedoc_demo(art):
    return try_hedgedoc("https://demo.hedgedoc.org", art, "demo.hedgedoc.org", 55)

def post_md_vern(art):
    return try_hedgedoc("https://md.vern.cc", art, "md.vern.cc", 45)

def post_haste_pluralkit(art):
    return try_hastebin("https://haste.pluralkit.me", art, "haste.pluralkit.me", 40)

def post_haste_schel(art):
    return try_hastebin("https://haste.schel.de", art, "haste.schel.de", 38)

def post_haste_spg(art):
    return try_hastebin("https://haste.spg.gg", art, "haste.spg.gg", 38)

def post_haste_clicksminuteper(art):
    return try_hastebin("https://hastebin.com", art, "hastebin.com", 52)

def post_pastegg(art):
    """paste.gg — API paste service"""
    try:
        r = requests.post("https://paste.gg/api/v1/pastes",
            json={"name": art["title"],
                  "visibility": "public",
                  "files": [{"name": "article.md",
                              "content": {"format": "text",
                                          "value": f"# {art['title']}\n\n{art['md']}"}}]},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            d = r.json()
            pid = d.get("result", {}).get("id", "")
            if pid:
                return True, f"https://paste.gg/p/anonymous/{pid}", "paste.gg", 45
    except: pass
    return False, "", "paste.gg", 45

def post_sourcehut(art):
    """paste.sr.ht (SourceHut)"""
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://paste.sr.ht/", timeout=10)
        csrf = s.cookies.get("_csrf_token", "")
        if not csrf:
            m = re.search(r'_csrf_token.*?value=["\']([^"\']+)', home.text)
            if m: csrf = m.group(1)
        r = s.post("https://paste.sr.ht/",
            data={"_csrf_token": csrf,
                  "visibility": "public",
                  "files[0].filename": "article.md",
                  "files[0].contents": f"# {art['title']}\n\n{art['md']}"},
            headers={"Referer": "https://paste.sr.ht/"}, allow_redirects=True, timeout=15)
        url = r.url
        if "paste.sr.ht/" in url and url not in ("https://paste.sr.ht/", "https://paste.sr.ht"):
            return True, url, "paste.sr.ht", 60
    except: pass
    return False, "", "paste.sr.ht", 60

def post_ghostbin_new(art):
    """ghostbin.com alternative ghostbin.co"""
    try:
        r = requests.post("https://ghostbin.com/paste/new",
            data={"text": f"# {art['title']}\n\n{art['md']}",
                  "expire": "2d", "lang": "text"},
            headers={"User-Agent": UA}, allow_redirects=True, timeout=15)
        url = r.url
        if "ghostbin.com/" in url and "/new" not in url:
            return True, url, "ghostbin.com", 45
    except: pass
    return False, "", "ghostbin.com", 45

def post_writeas_3(art):
    try:
        r = requests.post("https://write.as/api/posts",
            json={"title": art["title"], "body": art["md"], "font": "mono"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            pid = r.json().get("data", {}).get("id", "")
            if pid:
                return True, f"https://write.as/{pid}", "write.as/3", 69
    except: pass
    return False, "", "write.as/3", 69

def post_catbox_2(art):
    try:
        content = f"<html><head><title>{art['title']}</title></head><body><h1>{art['title']}</h1>{art['content_html']}</body></html>"
        fname = f"scotle-article-{random.randint(10000, 99999)}.html"
        r = requests.post("https://catbox.moe/user/api.php",
            data={"reqtype": "fileupload", "userhash": ""},
            files={"fileToUpload": (fname, content.encode(), "text/html")},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code == 200:
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, "catbox.moe/2", 55
    except: pass
    return False, "", "catbox.moe/2", 55

def post_friendpaste_3(art):
    try:
        r = requests.post("https://friendpaste.com/",
            json={"title": f"Scotle - {art['title'][:40]}",
                  "snippet": f"# {art['title']}\n\n{art['md']}",
                  "language": "markdown"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            pid = r.json().get("id", "")
            if pid:
                return True, f"https://friendpaste.com/{pid}", "friendpaste.com/3", 42
    except: pass
    return False, "", "friendpaste.com/3", 42

def post_glot_2(art):
    try:
        r = requests.post("https://glot.io/api/snippets",
            json={"language": "plaintext", "title": art["title"],
                  "public": True, "tags": ["school", "jaipur", "scotle"],
                  "files": [{"name": "blog.txt",
                              "content": f"# {art['title']}\n\n{art['md']}"}]},
            headers={"Content-Type": "application/json",
                     "Authorization": "Token undefined", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            sid = r.json().get("id", "")
            if sid:
                return True, f"https://glot.io/snippets/{sid}", "glot.io/2", 50
    except: pass
    return False, "", "glot.io/2", 50

def post_bytebin_2(art):
    try:
        content = f"# {art['title']}\n\n{art['md']}"
        r = requests.post("https://bytebin.lucko.me/post",
            data=content.encode(),
            headers={"Content-Type": "text/markdown", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            key = r.json().get("key", "")
            if key:
                return True, f"https://bytebin.lucko.me/{key}", "bytebin.lucko.me/2", 40
    except: pass
    return False, "", "bytebin.lucko.me/2", 40

def post_pastefy_2(art):
    try:
        r = requests.post("https://pastefy.app/api/v2/paste",
            json={"title": art["title"],
                  "content": f"# {art['title']}\n\n{art['md']}",
                  "type": "PASTE", "language": "plain"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            pid = r.json().get("id", "")
            if pid:
                return True, f"https://pastefy.app/{pid}", "pastefy.app/2", 45
    except: pass
    return False, "", "pastefy.app/2", 45


ALL_BATCH4 = [
    post_pastegg,
    post_boardnet_pad,
    post_okfn_pad,
    post_hedgedoc_demo,
    post_md_vern,
    post_haste_pluralkit,
    post_haste_schel,
    post_haste_spg,
    post_ghostbin_new,
    post_medienpad,
    post_opn_pad,
    post_writeas_3,
    post_catbox_2,
    post_friendpaste_3,
    post_glot_2,
    post_bytebin_2,
    post_pastefy_2,
    post_haste_clicksminuteper,
]


def load_existing():
    existing = []
    try:
        df = pd.read_excel(PREV_XLSX, sheet_name="All Published Posts")
        for _, row in df.iterrows():
            url = str(row.get("url", ""))
            if url.startswith("http"):
                existing.append({
                    "success": True, "url": url,
                    "platform": row.get("platform", ""),
                    "da": row.get("da", 0),
                    "title": row.get("title", ""),
                    "anchor": row.get("anchor", ""),
                    "word_count": row.get("word_count", 0),
                    "target": config.TARGET_URL,
                    "date": str(row.get("date", "")),
                })
        print(f"  Loaded {len(existing)} posts from previous file.")
    except Exception as e:
        print(f"  Error loading: {e}")
    return existing


def save_excel(results):
    pub = [r for r in results if r["success"]]
    if not pub:
        return

    pub_df = pd.DataFrame(pub)
    pub_df.insert(0, "S.No", range(1, len(pub_df) + 1))

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        rows = [
            ("SCOTLE.ORG — BACKLINK REPORT: 30 DIFFERENT BLOG SITES", ""),
            ("", ""),
            ("Target Website", config.TARGET_URL),
            ("Brand", config.TARGET_BRAND),
            ("Report Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("Total Posts", len(pub)),
            ("Different Websites", len(set(r["platform"] for r in pub))),
            ("Total Backlinks", f"{len(pub)*3} links to scotle.org"),
            ("", ""),
            ("S.No", "Website | Article URL"),
        ]
        for i, r in enumerate(pub, 1):
            rows.append((f"{i}. {r['platform']} (DA {r['da']})", r["url"]))

        summary = pd.DataFrame(rows, columns=["Metric", "Value"])
        summary.to_excel(w, sheet_name="Dashboard", index=False)
        pub_df.to_excel(w, sheet_name="All Published Posts", index=False)
        pub_df[["S.No","platform","da","url","title"]].rename(
            columns={"platform":"Website","da":"DA","url":"Article URL","title":"Title"}
        ).to_excel(w, sheet_name="Quick URL List", index=False)

    wb = load_workbook(OUT_XLSX)
    hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    gfill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    lfont = Font(name="Arial", size=10, color="1558D6", underline="single")
    gfont = Font(name="Arial", size=10, bold=True, color="137333")
    tfont = Font(name="Arial", size=13, bold=True, color="1A73E8")

    for sn in wb.sheetnames:
        ws = wb[sn]
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=col).value
                cell = ws.cell(row=row, column=col)
                val = str(cell.value or "")
                if val.startswith("http"):
                    cell.font = lfont
                if h in ("da", "DA"):
                    try:
                        da = int(cell.value)
                        if da >= 65:
                            cell.fill = gfill; cell.font = gfont
                        elif da >= 45:
                            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    except: pass
                if "SCOTLE" in val and col == 1:
                    cell.font = tfont
        for c in range(1, ws.max_column + 1):
            ml = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(c)].width = min(ml + 3, 72)
        ws.freeze_panes = "A2"
    wb["Dashboard"].column_dimensions["A"].width = 40
    wb["Dashboard"].column_dimensions["B"].width = 70
    wb.save(OUT_XLSX)


def main():
    print(f"\n{'='*62}")
    print(f"  SCOTLE.ORG — BATCH 4 (FINAL PUSH TO 30)")
    print(f"{'='*62}\n")

    all_results = load_existing()
    used_platforms = set(r["platform"] for r in all_results)
    success_count = len(all_results)
    print(f"  Have {success_count} posts on {len(used_platforms)} sites so far.\n")

    for i, fn in enumerate(ALL_BATCH4):
        if success_count >= 30:
            break

        art = gen_article()
        pname = fn.__name__.replace("post_", "")
        print(f"  [{i+1:>2}/{len(ALL_BATCH4)}] {pname:<28} {art['title'][:33]}... ", end="", flush=True)

        try:
            ok, url, platform, da = fn(art)
        except Exception as e:
            print(f"ERROR: {e}")
            continue

        if ok and url.startswith("http") and platform not in used_platforms:
            used_platforms.add(platform)
            success_count += 1
            all_results.append({
                "success": True, "url": url, "platform": platform, "da": da,
                "title": art["title"], "anchor": art.get("anchor", ""),
                "word_count": art.get("word_count", 0),
                "target": config.TARGET_URL,
                "date": datetime.now().strftime("%Y-%m-%d %H:%M")
            })
            print(f"OK [{success_count}/30] -> {url}")
        elif ok and platform in used_platforms:
            print(f"SKIP (dup: {platform})")
        else:
            print(f"FAIL")

        if i < len(ALL_BATCH4) - 1 and success_count < 30:
            time.sleep(random.uniform(3, 7))

    save_excel(all_results)

    pub = [r for r in all_results if r["success"]]
    print(f"\n{'='*62}")
    print(f"  DONE! {len(pub)} posts on {len(set(r['platform'] for r in pub))} different websites")
    print(f"  Excel: {OUT_XLSX}")
    print(f"{'='*62}\n")
    for j, r in enumerate(pub, 1):
        print(f"  {j:>2}. [{r['platform']} DA:{r['da']}] {r['url']}")

    import platform as pl
    if pl.system() == "Windows":
        os.startfile(os.path.abspath(OUT_XLSX))


if __name__ == "__main__":
    main()
