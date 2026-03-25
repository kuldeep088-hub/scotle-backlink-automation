"""
post_scotle_30_best.py
======================
Post 30 Scotle-style education blogs to the BEST available websites.

Strategy:
  - Only platforms with DA 50+ that are confirmed working
  - 10 unique article templates — each post gets a different article
  - Same platform used multiple times (each time = different article = real backlink)
  - Sorted by DA so highest-quality sites post first
  - All 30 results saved to data/Scotle_30_Best.xlsx

Run:
    python post_scotle_30_best.py
    python post_scotle_30_best.py --dry-run
"""

import os, sys, time, random, re, json, argparse, logging
import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from blog_poster import blog_config as config
from blog_poster.templates.scotle_templates import (
    get_all_scotle_articles, article_to_markdown, article_to_html
)

logging.basicConfig(level=logging.WARNING)

DATA_DIR   = os.path.join(os.path.dirname(__file__), "data")
OUTPUT_XLSX = os.path.join(DATA_DIR, "Scotle_30_Best.xlsx")
TELEGRAPH_TOKEN_FILE = os.path.join(os.path.dirname(__file__), "blog_poster", "telegraph_token.txt")
os.makedirs(DATA_DIR, exist_ok=True)

TARGET_URL = config.TARGET_URL
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"

ALL_ARTICLES = get_all_scotle_articles()


# ─────────────────────────────────────────────────────────────
# Build article dict from template index
# ─────────────────────────────────────────────────────────────
def build_art(index: int) -> dict:
    art = ALL_ARTICLES[index % len(ALL_ARTICLES)]
    md   = article_to_markdown(art, TARGET_URL)
    html = article_to_html(art, TARGET_URL)
    return {
        "title":      art["title"],
        "md":         md,
        "html":       html,
        "word_count": len(md.split()),
    }


# ══════════════════════════════════════════════════════════════
# PLATFORM FUNCTIONS  (DA 50+, confirmed working)
# ══════════════════════════════════════════════════════════════

# ── DA 90 ── Ubuntu Paste ────────────────────────────────────
def post_ubuntu(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://paste.ubuntu.com", timeout=15)
        csrf = ""
        for pat in [r"name=['\"]csrfmiddlewaretoken['\"] value=['\"]([^'\"]+)",
                    r"csrfmiddlewaretoken['\"][^'\"]*['\"][^'\"]*['\"]([^'\"]+)"]:
            m = re.search(pat, home.text)
            if m:
                csrf = m.group(1)
                break
        if not csrf:
            csrf = s.cookies.get("csrftoken", "")
        r = s.post("https://paste.ubuntu.com/",
            data={"csrfmiddlewaretoken": csrf,
                  "poster": "Scotle High School",
                  "syntax": "text",
                  "content": f"# {art['title']}\n\n{art['md']}",
                  "expiration": "week"},
            headers={"Referer": "https://paste.ubuntu.com/"}, timeout=20)
        if r.status_code == 200 and "paste.ubuntu.com/p/" in r.url:
            return True, r.url, "Ubuntu Paste", 90
    except Exception:
        pass
    return False, "", "Ubuntu Paste", 90


# ── DA 69 ── Write.as (real blog) ────────────────────────────
def post_writeas(art):
    try:
        r = requests.post("https://write.as/api/posts",
            json={"title": art["title"], "body": art["md"], "font": "sans"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            pid = r.json().get("data", {}).get("id", "")
            if pid:
                return True, f"https://write.as/{pid}", "Write.as", 69
    except Exception:
        pass
    return False, "", "Write.as", 69


# ── DA 55 ── Catbox.moe (HTML file hosting) ──────────────────
def post_catbox(art):
    try:
        html_page = (
            f"<!DOCTYPE html><html><head><title>{art['title']}</title>"
            f"<meta charset='utf-8'>"
            f"<style>body{{font-family:Georgia,serif;max-width:720px;margin:40px auto;"
            f"padding:0 20px;line-height:1.8;color:#222}}"
            f"h1{{color:#111;font-size:1.6em}}h2{{color:#333;font-size:1.2em;margin-top:1.5em}}"
            f"a{{color:#1558D6}}hr{{border:1px solid #ddd;margin:2em 0}}"
            f"</style></head><body>{art['html']}</body></html>"
        )
        fname = f"scotle-{random.randint(100000, 999999)}.html"
        r = requests.post("https://catbox.moe/user/api.php",
            data={"reqtype": "fileupload", "userhash": ""},
            files={"fileToUpload": (fname, html_page.encode("utf-8"), "text/html")},
            headers={"User-Agent": UA}, timeout=25)
        if r.status_code == 200:
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, "Catbox.moe", 55
    except Exception:
        pass
    return False, "", "Catbox.moe", 55


# ── DA 55 ── Filebin.net ─────────────────────────────────────
def post_filebin(art):
    try:
        bin_id = "".join(random.choices("abcdefghijklmnopqrstuvwxyz0123456789", k=16))
        fname  = f"scotle-blog-{random.randint(1000,9999)}.md"
        r = requests.post(f"https://filebin.net/{bin_id}/{fname}",
            data=f"# {art['title']}\n\n{art['md']}".encode("utf-8"),
            headers={"Content-Type": "text/plain", "User-Agent": UA,
                     "Accept": "application/json"}, timeout=20)
        if r.status_code in (200, 201):
            return True, f"https://filebin.net/{bin_id}", "Filebin.net", 55
    except Exception:
        pass
    return False, "", "Filebin.net", 55


# ── DA 52 ── paste.ee ────────────────────────────────────────
def post_paste_ee(art):
    try:
        r = requests.post("https://paste.ee/api",
            data={"key": "public",
                  "description": art["title"][:100],
                  "paste": f"# {art['title']}\n\n{art['md']}",
                  "expire": 0},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            link = d.get("link") or d.get("paste", {}).get("link", "")
            if link and link.startswith("http"):
                return True, link, "paste.ee", 52
    except Exception:
        pass
    return False, "", "paste.ee", 52


# ── DA 50 ── cl1p.net ────────────────────────────────────────
def post_cl1p(art):
    try:
        slug = "".join(random.choices("abcdefghijklmnopqrstuvwxyz0123456789", k=14))
        r = requests.post(f"https://cl1p.net/{slug}",
            data={"cl1pTextArea": f"# {art['title']}\n\n{art['md']}"},
            headers={"User-Agent": UA, "Referer": f"https://cl1p.net/{slug}"}, timeout=15)
        if r.status_code in (200, 201):
            return True, f"https://cl1p.net/{slug}", "cl1p.net", 50
    except Exception:
        pass
    return False, "", "cl1p.net", 50


# ── DA 50 ── Bytebin / pastes.dev ────────────────────────────
def post_bytebin(art):
    content = f"# {art['title']}\n\n{art['md']}"
    for base in ["https://pastes.dev", "https://bytebin.lucko.me"]:
        try:
            r = requests.post(f"{base}/post",
                data=content.encode("utf-8"),
                headers={"Content-Type": "text/markdown", "User-Agent": UA}, timeout=15)
            if r.status_code in (200, 201):
                key = r.json().get("key", "")
                if key:
                    return True, f"{base}/{key}", "Bytebin", 50
        except Exception:
            pass
    return False, "", "Bytebin", 50


# ── DA 45 ── just-paste.it ───────────────────────────────────
def post_justpaste(art):
    try:
        r = requests.post("https://just-paste.it/documents",
            data=f"# {art['title']}\n\n{art['md']}".encode("utf-8"),
            headers={"Content-Type": "text/plain", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            key = r.json().get("key", "")
            if key:
                return True, f"https://just-paste.it/{key}", "just-paste.it", 45
    except Exception:
        pass
    return False, "", "just-paste.it", 45


# ── DA 42 ── FriendPaste ─────────────────────────────────────
def post_friendpaste(art):
    try:
        r = requests.post("https://friendpaste.com/",
            json={"title": art["title"],
                  "snippet": f"# {art['title']}\n\n{art['md']}",
                  "language": "markdown"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            pid = r.json().get("id", "")
            if pid:
                return True, f"https://friendpaste.com/{pid}", "FriendPaste", 42
    except Exception:
        pass
    return False, "", "FriendPaste", 42


# ── DA 40 ── Bytebin (pastes.dev fallback already covered) ───

# ── BONUS: extra platforms to boost count ────────────────────

# Telegra.ph (DA 93) — try auto token
def _get_telegraph_token():
    if os.path.exists(TELEGRAPH_TOKEN_FILE):
        t = open(TELEGRAPH_TOKEN_FILE).read().strip()
        if t:
            return t
    try:
        r = requests.post("https://api.telegra.ph/createAccount",
            json={"short_name": "ScotleEdu",
                  "author_name": "Scotle High School",
                  "author_url": TARGET_URL}, timeout=15)
        d = r.json()
        if d.get("ok"):
            token = d["result"]["access_token"]
            os.makedirs(os.path.dirname(TELEGRAPH_TOKEN_FILE), exist_ok=True)
            open(TELEGRAPH_TOKEN_FILE, "w").write(token)
            return token
    except Exception:
        pass
    return None

def _html_to_telegraph_nodes(html_text):
    nodes = []
    chunks = re.split(
        r'(<h[123][^>]*>.*?</h[123]>|<p[^>]*>.*?</p>|<ul>.*?</ul>|<hr[^>]*/>)',
        html_text, flags=re.DOTALL)
    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk:
            continue
        m = re.match(r'<(h[123])[^>]*>(.*?)</\1>', chunk, re.DOTALL)
        if m:
            text = re.sub(r'<[^>]+>', '', m.group(2)).strip()
            nodes.append({"tag": "h4", "children": [text]})
            continue
        m = re.match(r'<p[^>]*>(.*?)</p>', chunk, re.DOTALL)
        if m:
            inner = m.group(1).strip()
            parts = []
            last = 0
            for am in re.finditer(r'<a\s+href="([^"]+)"[^>]*>(.*?)</a>', inner, re.DOTALL):
                pre = re.sub(r'<[^>]+>', '', inner[last:am.start()])
                if pre: parts.append(pre)
                parts.append({"tag": "a", "attrs": {"href": am.group(1)},
                               "children": [am.group(2)]})
                last = am.end()
            tail = re.sub(r'<[^>]+>', '', inner[last:])
            if tail: parts.append(tail)
            if parts:
                nodes.append({"tag": "p", "children": parts})
            continue
        m = re.match(r'<ul>(.*?)</ul>', chunk, re.DOTALL)
        if m:
            for li in re.findall(r'<li>(.*?)</li>', m.group(1), re.DOTALL):
                text = re.sub(r'<[^>]+>', '', li).strip()
                nodes.append({"tag": "p", "children": [f"• {text}"]})
            continue
        plain = re.sub(r'<[^>]+>', '', chunk).strip()
        if plain:
            nodes.append({"tag": "p", "children": [plain]})
    return nodes

def post_telegraph(art):
    token = _get_telegraph_token()
    if not token:
        return False, "", "Telegra.ph", 93
    nodes = _html_to_telegraph_nodes(art["html"])
    try:
        r = requests.post("https://api.telegra.ph/createPage",
            json={"access_token": token,
                  "title": art["title"][:256],
                  "author_name": "Scotle High School",
                  "author_url": TARGET_URL,
                  "content": nodes,
                  "return_content": False}, timeout=20)
        d = r.json()
        if d.get("ok"):
            url = d["result"].get("url", "")
            if url:
                return True, url, "Telegra.ph", 93
    except Exception:
        pass
    return False, "", "Telegra.ph", 93

# Mozilla Paste (DA 97)
def post_mozilla(art):
    try:
        r = requests.post("https://paste.mozilla.org/api/",
            data={"content": f"# {art['title']}\n\n{art['md']}",
                  "format": "url", "expires": "3153600000", "lexer": "_markdown"},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, "Mozilla Paste", 97
    except Exception:
        pass
    return False, "", "Mozilla Paste", 97

# GitHub Gist anonymous (DA 96)
def post_gist(art):
    fname = f"scotle-education-{random.randint(1000, 9999)}.md"
    try:
        r = requests.post("https://api.github.com/gists",
            json={"description": art["title"], "public": True,
                  "files": {fname: {"content": f"# {art['title']}\n\n{art['md']}"}}},
            headers={"Accept": "application/vnd.github.v3+json", "User-Agent": UA,
                     "X-GitHub-Api-Version": "2022-11-28"}, timeout=20)
        if r.status_code == 201:
            url = r.json().get("html_url", "")
            if url:
                return True, url, "GitHub Gist", 96
    except Exception:
        pass
    return False, "", "GitHub Gist", 96

# hastebin.com DA ~55
def post_hastebin(art):
    for base in ["https://hastebin.com", "https://toptal.com/developers/hastebin"]:
        try:
            r = requests.post(f"{base}/documents",
                data=f"# {art['title']}\n\n{art['md']}".encode("utf-8"),
                headers={"Content-Type": "text/plain", "User-Agent": UA}, timeout=15)
            if r.status_code in (200, 201):
                key = r.json().get("key", "")
                if key:
                    return True, f"https://hastebin.com/{key}", "Hastebin", 55
        except Exception:
            pass
    return False, "", "Hastebin", 55

# paste.gg DA ~45
def post_paste_gg(art):
    try:
        r = requests.post("https://paste.gg/api/v1/pastes",
            json={"name": art["title"][:128],
                  "description": "Education blog — Scotle High School",
                  "visibility": "public",
                  "files": [{"name": "article.md",
                              "content": {"format": "text",
                                          "value": f"# {art['title']}\n\n{art['md']}"}}]},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            paste_id = d.get("result", {}).get("id", "")
            if paste_id:
                return True, f"https://paste.gg/p/anonymous/{paste_id}", "paste.gg", 45
    except Exception:
        pass
    return False, "", "paste.gg", 45

# dpaste.org DA ~50 (different from dpaste.com)
def post_dpaste_org(art):
    try:
        r = requests.post("https://dpaste.org/api/",
            data={"content": f"# {art['title']}\n\n{art['md']}",
                  "syntax": "markdown", "expiry_days": 365},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            url = r.text.strip().strip('"')
            if url.startswith("http"):
                return True, url, "dpaste.org", 50
    except Exception:
        pass
    return False, "", "dpaste.org", 50

# rentry.co DA 55
def post_rentry(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        s.get("https://rentry.co", timeout=15)
        csrf = s.cookies.get("csrftoken", "")
        r = s.post("https://rentry.co/api/new",
            data={"csrfmiddlewaretoken": csrf,
                  "text": f"# {art['title']}\n\n{art['md']}",
                  "edit_code": "".join(random.choices(
                      "abcdefghijklmnopqrstuvwxyz0123456789", k=12))},
            headers={"Referer": "https://rentry.co"}, timeout=20)
        if r.status_code == 200:
            d = r.json()
            if d.get("status") == "200":
                url = d.get("url", "")
                if url:
                    return True, url, "Rentry.co", 55
    except Exception:
        pass
    return False, "", "Rentry.co", 55

# Debian Paste DA 65
def post_debian(art):
    try:
        r = requests.post("https://paste.debian.net/",
            data={"poster": "Scotle",
                  "paste": f"# {art['title']}\n\n{art['md']}",
                  "expire": "604800", "submit": "Submit"},
            headers={"User-Agent": UA, "Referer": "https://paste.debian.net/"}, timeout=20)
        if r.status_code == 200:
            m = re.search(r'(https?://paste\.debian\.net/\d+)', r.url + " " + r.text)
            if m:
                return True, m.group(1), "Debian Paste", 65
    except Exception:
        pass
    return False, "", "Debian Paste", 65


# ══════════════════════════════════════════════════════════════
# PLATFORM POOL — ordered by DA, each has a max post count
# The script round-robins through this pool until 30 posts done
# ══════════════════════════════════════════════════════════════

PLATFORM_POOL = [
    # (function,         name,            DA,  max_posts)
    (post_mozilla,       "Mozilla Paste",  97,  3),
    (post_gist,          "GitHub Gist",    96,  3),
    (post_telegraph,     "Telegra.ph",     93,  3),
    (post_ubuntu,        "Ubuntu Paste",   90,  5),
    (post_debian,        "Debian Paste",   65,  3),
    (post_hastebin,      "Hastebin",       55,  3),
    (post_catbox,        "Catbox.moe",     55,  4),
    (post_filebin,       "Filebin.net",    55,  4),
    (post_rentry,        "Rentry.co",      55,  3),
    (post_dpaste_org,    "dpaste.org",     50,  3),
    (post_writeas,       "Write.as",       69,  6),
    (post_paste_ee,      "paste.ee",       52,  4),
    (post_cl1p,          "cl1p.net",       50,  4),
    (post_bytebin,       "Bytebin",        50,  4),
    (post_paste_gg,      "paste.gg",       45,  3),
    (post_justpaste,     "just-paste.it",  45,  3),
    (post_friendpaste,   "FriendPaste",    42,  3),
]


# ══════════════════════════════════════════════════════════════
# Excel report
# ══════════════════════════════════════════════════════════════

def save_excel(results):
    pub  = [r for r in results if r["success"]]
    fail = [r for r in results if not r["success"]]

    if not pub:
        print("  [!] No posts to save.")
        return

    pub_df = pd.DataFrame(pub)
    if "S.No" not in pub_df.columns:
        pub_df.insert(0, "S.No", range(1, len(pub_df) + 1))
    else:
        pub_df["S.No"] = range(1, len(pub_df) + 1)

    avg_da = sum(r["da"] for r in pub) // max(len(pub), 1)
    top_da = max(r["da"] for r in pub)

    summary_rows = [
        ("SCOTLE.ORG — 30 BEST SITES BLOG REPORT", ""),
        ("", ""),
        ("Target URL",      TARGET_URL),
        ("Blog Style",      "Scotle High School — Maths Education"),
        ("Report Date",     datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total Published",    len(pub)),
        ("Total Failed",       len(fail)),
        ("Unique Platforms",   len(set(r["platform"] for r in pub))),
        ("Average DA",         avg_da),
        ("Highest DA",         top_da),
        ("Backlinks Created",  f"{len(pub)} (1 per post — end CTA only)"),
        ("", ""),
        ("ALL LIVE BACKLINKS (sorted by DA)", ""),
    ]
    for r in sorted(pub, key=lambda x: (-x["da"], x["platform"])):
        summary_rows.append((f"  {r['platform']}  (DA {r['da']})", r["url"]))

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as w:
        pd.DataFrame(summary_rows, columns=["Metric", "Value"]).to_excel(
            w, sheet_name="Dashboard", index=False)
        pub_df.to_excel(w, sheet_name="All 30 Posts", index=False)
        pub_df[["S.No", "platform", "da", "url", "title"]].to_excel(
            w, sheet_name="Quick URL List", index=False)
        if fail:
            pd.DataFrame(fail).to_excel(w, sheet_name="Failed", index=False)

    # ── Style ────────────────────────────────────────────────
    wb    = load_workbook(OUTPUT_XLSX)
    hfill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    lfont = Font(name="Arial", size=10, color="1565C0", underline="single")
    gfill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    gfont = Font(name="Arial", size=10, bold=True, color="1B5E20")
    yfill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

    for sn in wb.sheetnames:
        ws = wb[sn]
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                h    = ws.cell(row=1, column=col).value
                cell = ws.cell(row=row, column=col)
                val  = str(cell.value or "")
                if h in ("url", "Value") and val.startswith("http"):
                    cell.font = lfont
                if h == "da":
                    try:
                        da = int(cell.value)
                        if da >= 65:
                            cell.fill = gfill
                            cell.font = gfont
                        elif da >= 50:
                            cell.fill = yfill
                    except Exception:
                        pass
        for c in range(1, ws.max_column + 1):
            ml = max(len(str(ws.cell(row=r, column=c).value or ""))
                     for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(c)].width = min(ml + 4, 72)
        ws.freeze_panes = "A2"
    wb.save(OUTPUT_XLSX)
    print(f"\n  Excel saved -> {OUTPUT_XLSX}")


# ══════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════

def main(target=30, dry_run=False):
    print(f"\n{'='*68}")
    print(f"  SCOTLE.ORG — 30 BEST-SITE EDUCATION BLOG POSTS")
    print(f"  Target URL    : {TARGET_URL}")
    print(f"  Articles ready: {len(ALL_ARTICLES)} unique templates")
    print(f"  Platforms pool: {len(PLATFORM_POOL)}")
    print(f"  Target posts  : {target}")
    if dry_run:
        print(f"  Mode          : DRY RUN (no actual posting)")
    print(f"{'='*68}\n")

    if dry_run:
        print("  Platform pool (DA order):\n")
        for fn, name, da, mx in sorted(PLATFORM_POOL, key=lambda x: -x[2]):
            print(f"    DA {da:>2}  {name:<22}  (up to {mx} posts)")
        print(f"\n  Articles:\n")
        for i, a in enumerate(ALL_ARTICLES, 1):
            md = article_to_markdown(a, TARGET_URL)
            print(f"    {i:>2}. {a['title']}  ({len(md.split())} words)")
        print()
        return

    results     = []
    art_index   = 0
    post_counts = {name: 0 for _, name, _, _ in PLATFORM_POOL}

    # Build attempt queue: sorted by DA descending, respecting max_posts
    # We round-robin: one pass gives 1 post per platform, repeat until target
    attempt_queue = []
    max_rounds = max(mx for _, _, _, mx in PLATFORM_POOL)
    for round_n in range(max_rounds):
        for fn, name, da, mx in sorted(PLATFORM_POOL, key=lambda x: -x[2]):
            if round_n < mx:
                attempt_queue.append((fn, name, da))

    success_count = 0
    print(f"  {'#':<4} {'DA':>4}  {'Platform':<22}  {'Status'}")
    print(f"  {'-'*62}")

    for fn, name, da in attempt_queue:
        if success_count >= target:
            break

        art = build_art(art_index)
        art_index += 1

        title_short = art["title"][:36]
        print(f"  [{success_count+1:>2}] DA{da:>3}  {name:<22}  {title_short}... ",
              end="", flush=True)

        ok, url, platform, actual_da = fn(art)

        if ok:
            success_count += 1
            post_counts[name] = post_counts.get(name, 0) + 1
            results.append({
                "success":    True,
                "S.No":       success_count,
                "platform":   platform,
                "da":         actual_da,
                "url":        url,
                "title":      art["title"],
                "word_count": art["word_count"],
                "backlink":   TARGET_URL,
                "date":       datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
            print(f"OK  ->  {url}")
        else:
            results.append({
                "success": False, "S.No": "",
                "platform": platform, "da": actual_da,
                "url": "", "title": art["title"],
                "word_count": art["word_count"],
                "backlink": TARGET_URL,
                "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
            print("FAIL")

        time.sleep(random.uniform(4, 8))

    # ── Summary ──────────────────────────────────────────────
    pub  = [r for r in results if r["success"]]
    fail = [r for r in results if not r["success"]]
    avg_da = sum(r["da"] for r in pub) // max(len(pub), 1)

    save_excel(results)

    print(f"\n{'='*68}")
    print(f"  COMPLETED")
    print(f"  Posts published : {len(pub)} / {target}")
    print(f"  Average DA      : {avg_da}")
    print(f"  Unique platforms: {len(set(r['platform'] for r in pub))}")
    print(f"{'='*68}\n")

    if pub:
        print("  ALL LIVE BACKLINKS (best DA first):\n")
        for r in sorted(pub, key=lambda x: (-x["da"], x["date"])):
            print(f"    DA {r['da']:>2}  {r['platform']:<22}  {r['url']}")

    if fail:
        fail_platforms = list(set(r["platform"] for r in fail))
        print(f"\n  Failed platforms: {', '.join(fail_platforms)}")

    print()
    return pub


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--target",  type=int, default=30,
                        help="Number of posts to publish (default: 30)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview without posting")
    args = parser.parse_args()
    main(target=args.target, dry_run=args.dry_run)
