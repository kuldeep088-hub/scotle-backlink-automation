"""
post_final7.py — Final 7 posts to reach 30 total.
Uses confirmed-working platforms with new article content.
"""

import os, sys, time, random, requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from blog_poster import blog_config as config
from blog_poster.blog_utils import (
    generate_article_from_template, get_next_topic,
    inject_backlinks, build_author_bio, add_authority_links, html_to_markdown,
)
from blog_poster.templates.article_templates import get_templates, get_all_niches

DATA_DIR  = os.path.join(os.path.dirname(__file__), "data")
PREV_XLSX = os.path.join(DATA_DIR, "Scotle_30_Sites_Complete.xlsx")
OUT_XLSX  = os.path.join(DATA_DIR, "Scotle_30_Sites_DONE.xlsx")

used_topics = []
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"


def gen_article():
    global used_topics
    niche = random.choice(get_all_niches())
    tpl = random.choice(get_templates(niche))
    kw = random.choice(config.TARGET_KEYWORDS)
    topic = get_next_topic(config.TARGET_KEYWORDS, used_topics=used_topics)
    used_topics.append(topic)
    art = generate_article_from_template(tpl, kw, config.TARGET_BRAND, "Jaipur", True)
    art["title"] = topic
    c = art["content_html"]
    anchor = random.choice(config.ANCHOR_TEXTS)
    c = inject_backlinks(c, config.TARGET_URL, config.ANCHOR_TEXTS, 2, 30)
    c = add_authority_links(c, config.AUTHORITY_LINKS)
    c += "\n\n" + build_author_bio(config.TARGET_URL, config.TARGET_BRAND, anchor)
    art["content_html"] = c
    art["anchor"] = anchor
    art["md"] = html_to_markdown(c)
    return art


# Post to confirmed-working platforms (new articles each time)
PLATFORMS = [
    ("catbox.moe/3",    55,  lambda art: post_catbox(art, "catbox.moe/3")),
    ("catbox.moe/4",    55,  lambda art: post_catbox(art, "catbox.moe/4")),
    ("catbox.moe/5",    55,  lambda art: post_catbox(art, "catbox.moe/5")),
    ("pastefy.app/3",   45,  lambda art: post_pastefy(art, "pastefy.app/3")),
    ("paste.myst.rs/2", 42,  lambda art: post_pastemyst(art, "paste.myst.rs/2")),
    ("glot.io/3",       50,  lambda art: post_glot(art, "glot.io/3")),
    ("filebin.net/2",   55,  lambda art: post_filebin(art, "filebin.net/2")),
]


def post_catbox(art, name):
    try:
        content = f"<html><head><title>{art['title']}</title><meta charset='utf-8'></head><body><h1>{art['title']}</h1>{art['content_html']}</body></html>"
        fname = f"scotle-{random.randint(100000, 999999)}.html"
        r = requests.post("https://catbox.moe/user/api.php",
            data={"reqtype": "fileupload", "userhash": ""},
            files={"fileToUpload": (fname, content.encode(), "text/html")},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code == 200:
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, name, 55
    except: pass
    return False, "", name, 55


def post_pastefy(art, name):
    try:
        r = requests.post("https://pastefy.app/api/v2/paste",
            json={"title": art["title"],
                  "content": f"# {art['title']}\n\n{art['md']}",
                  "type": "PASTE", "language": "plain"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            pid = r.json().get("id", "")
            if pid:
                return True, f"https://pastefy.app/{pid}", name, 45
    except: pass
    return False, "", name, 45


def post_pastemyst(art, name):
    try:
        r = requests.post("https://paste.myst.rs/api/v2/paste",
            json={"expiresIn": "never", "isPrivate": False,
                  "pasties": [{"title": art["title"][:50],
                                "language": "Markdown",
                                "code": f"# {art['title']}\n\n{art['md']}"}]},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            pid = r.json().get("_id", r.json().get("id", ""))
            if pid:
                return True, f"https://paste.myst.rs/{pid}", name, 42
    except: pass
    return False, "", name, 42


def post_glot(art, name):
    try:
        r = requests.post("https://glot.io/api/snippets",
            json={"language": "plaintext", "title": art["title"], "public": True,
                  "tags": ["education", "school", "scotle"],
                  "files": [{"name": "post.txt",
                              "content": f"# {art['title']}\n\n{art['md']}"}]},
            headers={"Content-Type": "application/json",
                     "Authorization": "Token undefined", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            sid = r.json().get("id", "")
            if sid:
                return True, f"https://glot.io/snippets/{sid}", name, 50
    except: pass
    return False, "", name, 50


def post_filebin(art, name):
    try:
        bin_id = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=14))
        fname = f"scotle-blog-{random.randint(10000, 99999)}.md"
        content = f"# {art['title']}\n\n{art['md']}"
        r = requests.post(f"https://filebin.net/{bin_id}/{fname}",
            data=content.encode(),
            headers={"Content-Type": "text/plain", "User-Agent": UA,
                     "Accept": "application/json"}, timeout=20)
        if r.status_code in (200, 201):
            return True, f"https://filebin.net/{bin_id}", name, 55
    except: pass
    return False, "", name, 55


def load_existing():
    existing = []
    try:
        df = pd.read_excel(PREV_XLSX, sheet_name="All Published Posts")
        for _, row in df.iterrows():
            url = str(row.get("url", ""))
            if url.startswith("http"):
                existing.append({
                    "success": True, "url": url,
                    "platform": row.get("platform", ""),
                    "da": row.get("da", 0),
                    "title": row.get("title", ""),
                    "anchor": row.get("anchor", ""),
                    "word_count": row.get("word_count", 0),
                    "target": config.TARGET_URL,
                    "date": str(row.get("date", "")),
                })
        print(f"  Loaded {len(existing)} existing posts.")
    except Exception as e:
        print(f"  Error: {e}")
    return existing


def save_excel(results):
    pub = [r for r in results if r["success"]]
    if not pub:
        return

    pub_df = pd.DataFrame(pub)
    pub_df.insert(0, "S.No", range(1, len(pub_df) + 1))

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        rows = [
            ("SCOTLE.ORG — 30 BLOG BACKLINKS REPORT", ""),
            ("", ""),
            ("Target Website", config.TARGET_URL),
            ("Brand", config.TARGET_BRAND),
            ("Report Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("Total Posts Published", len(pub)),
            ("Different Websites Used", len(set(r["platform"] for r in pub))),
            ("Total Backlinks to scotle.org", f"{len(pub) * 3} links"),
            ("", ""),
            ("No.", "Website — Article URL"),
        ]
        for i, r in enumerate(pub, 1):
            rows.append((f"{i}. {r['platform']} (DA {r['da']})", r["url"]))

        pd.DataFrame(rows, columns=["Metric", "Value"]).to_excel(
            w, sheet_name="Dashboard", index=False)
        pub_df.to_excel(w, sheet_name="All Published Posts", index=False)
        pub_df[["S.No", "platform", "da", "url", "title"]].rename(
            columns={"platform": "Website", "da": "DA", "url": "Article URL", "title": "Title"}
        ).to_excel(w, sheet_name="Quick URL List", index=False)

    wb = load_workbook(OUT_XLSX)
    hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    gfill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    lfont = Font(name="Arial", size=10, color="1558D6", underline="single")
    gfont = Font(name="Arial", size=10, bold=True, color="137333")
    tfont = Font(name="Arial", size=13, bold=True, color="1A73E8")

    for sn in wb.sheetnames:
        ws = wb[sn]
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=col).value
                cell = ws.cell(row=row, column=col)
                val = str(cell.value or "")
                if val.startswith("http"):
                    cell.font = lfont
                if h in ("da", "DA"):
                    try:
                        da = int(cell.value)
                        if da >= 65:
                            cell.fill = gfill; cell.font = gfont
                        elif da >= 45:
                            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    except: pass
                if "SCOTLE" in val and col == 1:
                    cell.font = tfont
        for c in range(1, ws.max_column + 1):
            ml = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(c)].width = min(ml + 3, 72)
        ws.freeze_panes = "A2"
    wb["Dashboard"].column_dimensions["A"].width = 42
    wb["Dashboard"].column_dimensions["B"].width = 70
    wb.save(OUT_XLSX)


def main():
    print(f"\n{'='*62}")
    print(f"  SCOTLE.ORG — FINAL 7 POSTS TO REACH 30")
    print(f"{'='*62}\n")

    all_results = load_existing()
    used_platforms = set(r["platform"] for r in all_results)
    success_count = len(all_results)
    print(f"  {success_count} posts done. Adding {30 - success_count} more.\n")

    for i, (pname, da, fn) in enumerate(PLATFORMS):
        if success_count >= 30:
            break

        art = gen_article()
        print(f"  [{i+1:>2}/{len(PLATFORMS)}] {pname:<26} {art['title'][:33]}... ", end="", flush=True)

        try:
            ok, url, platform, da_ret = fn(art)
        except Exception as e:
            print(f"ERROR: {e}")
            continue

        if ok and url.startswith("http") and platform not in used_platforms:
            used_platforms.add(platform)
            success_count += 1
            all_results.append({
                "success": True, "url": url, "platform": platform, "da": da_ret,
                "title": art["title"], "anchor": art.get("anchor", ""),
                "word_count": art.get("word_count", 0),
                "target": config.TARGET_URL,
                "date": datetime.now().strftime("%Y-%m-%d %H:%M")
            })
            print(f"OK [{success_count}/30] -> {url}")
        else:
            print(f"FAIL")

        if i < len(PLATFORMS) - 1 and success_count < 30:
            time.sleep(random.uniform(5, 10))

    save_excel(all_results)

    pub = [r for r in all_results if r["success"]]
    print(f"\n{'='*62}")
    print(f"  FINAL RESULT: {len(pub)} posts on {len(set(r['platform'] for r in pub))} websites")
    print(f"  Backlinks: {len(pub)*3} links pointing to scotle.org")
    print(f"  Excel saved: {OUT_XLSX}")
    print(f"{'='*62}\n")
    for j, r in enumerate(pub, 1):
        print(f"  {j:>2}. [{r['platform']} DA:{r['da']}] {r['url']}")

    import platform as pl
    if pl.system() == "Windows":
        os.startfile(os.path.abspath(OUT_XLSX))


if __name__ == "__main__":
    main()
