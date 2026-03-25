"""
post_scotle_best.py
===================
Post Scotle-style education blogs ONLY to the BEST websites.

Curated list — DA 50+ platforms that reliably accept posts:
  DA 97  Mozilla Paste        (paste.mozilla.org)
  DA 96  GitHub Gist          (gist.github.com) — anonymous
  DA 93  Telegra.ph           (telegra.ph)
  DA 90  Ubuntu Paste         (paste.ubuntu.com)
  DA 70  HackMD               (hackmd.io)
  DA 69  Write.as             (write.as) — real blog
  DA 65  Debian Paste         (paste.debian.net)
  DA 60  Codepad              (codepad.org)
  DA 55  Catbox.moe           (catbox.moe)
  DA 55  Filebin              (filebin.net)
  DA 55  Rentry               (rentry.co)
  DA 52  paste.ee             (paste.ee)
  DA 50  cl1p.net             (cl1p.net)
  DA 50  Bytebin / pastes.dev (bytebin.lucko.me)

Style: Scotle High School education articles, 1 backlink at end only.

Run:
    python post_scotle_best.py
    python post_scotle_best.py --dry-run
"""

import os, sys, time, random, json, re, logging, argparse
import requests
import pandas as pd
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from blog_poster import blog_config as config
from blog_poster.templates.scotle_templates import (
    get_all_scotle_articles, article_to_markdown, article_to_html
)

logging.basicConfig(level=logging.WARNING, format="%(asctime)s %(message)s")

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)
OUTPUT_XLSX = os.path.join(DATA_DIR, "Scotle_Best_Sites.xlsx")
TELEGRAPH_TOKEN_FILE = os.path.join(os.path.dirname(__file__), "blog_poster", "telegraph_token.txt")

TARGET_URL = config.TARGET_URL
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"


# ─────────────────────────────────────────────────────────────
# Article generation
# ─────────────────────────────────────────────────────────────

def gen_article(index: int) -> dict:
    articles = get_all_scotle_articles()
    art = articles[index % len(articles)]
    md = article_to_markdown(art, TARGET_URL)
    html = article_to_html(art, TARGET_URL)
    return {
        "title": art["title"],
        "md": md,
        "html": html,
        "word_count": len(md.split()),
    }


# ─────────────────────────────────────────────────────────────
# PLATFORM 1 — Mozilla Paste (DA 97)
# ─────────────────────────────────────────────────────────────
def post_mozilla(art):
    try:
        r = requests.post(
            "https://paste.mozilla.org/api/",
            data={
                "content": f"# {art['title']}\n\n{art['md']}",
                "format": "url",
                "expires": "3153600000",
                "lexer": "_markdown",
            },
            headers={"User-Agent": UA},
            timeout=20,
        )
        if r.status_code in (200, 201):
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, "Mozilla Paste", 97
    except Exception:
        pass
    return False, "", "Mozilla Paste", 97


# ─────────────────────────────────────────────────────────────
# PLATFORM 2 — GitHub Gist anonymous (DA 96)
# ─────────────────────────────────────────────────────────────
def post_gist(art):
    fname = f"scotle-education-{random.randint(1000,9999)}.md"
    try:
        r = requests.post(
            "https://api.github.com/gists",
            json={
                "description": art["title"],
                "public": True,
                "files": {fname: {"content": f"# {art['title']}\n\n{art['md']}"}},
            },
            headers={
                "Accept": "application/vnd.github.v3+json",
                "User-Agent": UA,
                "X-GitHub-Api-Version": "2022-11-28",
            },
            timeout=20,
        )
        if r.status_code == 201:
            url = r.json().get("html_url", "")
            if url:
                return True, url, "GitHub Gist", 96
    except Exception:
        pass
    return False, "", "GitHub Gist", 96


# ─────────────────────────────────────────────────────────────
# PLATFORM 3 — Telegra.ph (DA 93)
# ─────────────────────────────────────────────────────────────
def _get_telegraph_token():
    if os.path.exists(TELEGRAPH_TOKEN_FILE):
        with open(TELEGRAPH_TOKEN_FILE) as f:
            t = f.read().strip()
            if t:
                return t
    try:
        r = requests.post(
            "https://api.telegra.ph/createAccount",
            json={
                "short_name": "ScotleEdu",
                "author_name": "Scotle High School Blog",
                "author_url": TARGET_URL,
            },
            timeout=15,
        )
        d = r.json()
        if d.get("ok"):
            token = d["result"]["access_token"]
            os.makedirs(os.path.dirname(TELEGRAPH_TOKEN_FILE), exist_ok=True)
            with open(TELEGRAPH_TOKEN_FILE, "w") as f:
                f.write(token)
            return token
    except Exception:
        pass
    return None


def _html_to_telegraph_nodes(html_text):
    """Convert simple HTML to Telegraph node format."""
    nodes = []
    # Split on block tags
    chunks = re.split(r'(<h[123][^>]*>.*?</h[123]>|<p[^>]*>.*?</p>|<ul>.*?</ul>|<hr[^>]*/?>)', html_text, flags=re.DOTALL)
    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk:
            continue
        m = re.match(r'<(h[123])[^>]*>(.*?)</\1>', chunk, re.DOTALL)
        if m:
            tag = "h3" if m.group(1) == "h3" else ("h4" if m.group(1) == "h2" else "h3")
            text = re.sub(r'<[^>]+>', '', m.group(2)).strip()
            nodes.append({"tag": tag, "children": [text]})
            continue
        m = re.match(r'<p[^>]*>(.*?)</p>', chunk, re.DOTALL)
        if m:
            inner = m.group(1).strip()
            # Handle <a href>
            parts = []
            last = 0
            for am in re.finditer(r'<a\s+href="([^"]+)"[^>]*>(.*?)</a>', inner, re.DOTALL):
                pre = re.sub(r'<[^>]+>', '', inner[last:am.start()])
                if pre:
                    parts.append(pre)
                parts.append({"tag": "a", "attrs": {"href": am.group(1)}, "children": [am.group(2)]})
                last = am.end()
            tail = re.sub(r'<[^>]+>', '', inner[last:])
            if tail:
                parts.append(tail)
            if parts:
                nodes.append({"tag": "p", "children": parts})
            continue
        m = re.match(r'<ul>(.*?)</ul>', chunk, re.DOTALL)
        if m:
            items = re.findall(r'<li>(.*?)</li>', m.group(1), re.DOTALL)
            for item in items:
                text = re.sub(r'<[^>]+>', '', item).strip()
                nodes.append({"tag": "p", "children": [f"• {text}"]})
            continue
        if re.match(r'<hr', chunk):
            nodes.append({"tag": "p", "children": ["—"]})
            continue
        plain = re.sub(r'<[^>]+>', '', chunk).strip()
        if plain:
            nodes.append({"tag": "p", "children": [plain]})
    return nodes


def post_telegraph(art):
    token = _get_telegraph_token()
    if not token:
        return False, "", "Telegra.ph", 93
    nodes = _html_to_telegraph_nodes(art["html"])
    try:
        r = requests.post(
            "https://api.telegra.ph/createPage",
            json={
                "access_token": token,
                "title": art["title"][:256],
                "author_name": "Scotle High School",
                "author_url": TARGET_URL,
                "content": nodes,
                "return_content": False,
            },
            timeout=20,
        )
        d = r.json()
        if d.get("ok"):
            url = d["result"].get("url", "")
            if url:
                return True, url, "Telegra.ph", 93
    except Exception:
        pass
    return False, "", "Telegra.ph", 93


# ─────────────────────────────────────────────────────────────
# PLATFORM 4 — Ubuntu Paste (DA 90)
# ─────────────────────────────────────────────────────────────
def post_ubuntu(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://paste.ubuntu.com", timeout=15)
        csrf = ""
        m = re.search(r"csrfmiddlewaretoken['\"][^'\"]*['\"][^'\"]*['\"]([^'\"]+)", home.text)
        if not m:
            m = re.search(r"name=['\"]csrfmiddlewaretoken['\"] value=['\"]([^'\"]+)", home.text)
        if m:
            csrf = m.group(1)
        if not csrf:
            csrf = s.cookies.get("csrftoken", "")
        r = s.post(
            "https://paste.ubuntu.com/",
            data={
                "csrfmiddlewaretoken": csrf,
                "poster": "Scotle High School",
                "syntax": "text",
                "content": f"# {art['title']}\n\n{art['md']}",
                "expiration": "week",
            },
            headers={"Referer": "https://paste.ubuntu.com/"},
            timeout=20,
        )
        if r.status_code == 200 and "paste.ubuntu.com/p/" in r.url:
            return True, r.url, "Ubuntu Paste", 90
    except Exception:
        pass
    return False, "", "Ubuntu Paste", 90


# ─────────────────────────────────────────────────────────────
# PLATFORM 5 — HackMD (DA 70) — form-based note creation
# ─────────────────────────────────────────────────────────────
def post_hackmd(art):
    # Try API v1 (no auth required for public notes)
    for endpoint in [
        "https://hackmd.io/api/notes",
        "https://hackmd.io/new",
    ]:
        try:
            if endpoint.endswith("/notes"):
                r = requests.post(
                    endpoint,
                    json={
                        "title": art["title"],
                        "content": f"# {art['title']}\n\n{art['md']}",
                        "readPermission": "guest",
                        "writePermission": "owner",
                    },
                    headers={"Content-Type": "application/json", "User-Agent": UA},
                    timeout=20,
                )
                if r.status_code in (200, 201):
                    d = r.json()
                    pub = d.get("publishLink") or d.get("id") or d.get("noteId")
                    if pub:
                        url = pub if pub.startswith("http") else f"https://hackmd.io/{pub}"
                        return True, url, "HackMD", 70
            else:
                r = requests.post(
                    endpoint,
                    data={"markdown": f"# {art['title']}\n\n{art['md']}"},
                    headers={
                        "User-Agent": UA,
                        "Content-Type": "application/x-www-form-urlencoded",
                    },
                    allow_redirects=True,
                    timeout=20,
                )
                if r.status_code == 200 and "hackmd.io" in r.url and "/new" not in r.url:
                    return True, r.url, "HackMD", 70
        except Exception:
            pass
    return False, "", "HackMD", 70


# ─────────────────────────────────────────────────────────────
# PLATFORM 6 — Write.as (DA 69) — real blog
# ─────────────────────────────────────────────────────────────
def post_writeas(art):
    try:
        r = requests.post(
            "https://write.as/api/posts",
            json={"title": art["title"], "body": art["md"], "font": "sans"},
            headers={"Content-Type": "application/json", "User-Agent": UA},
            timeout=20,
        )
        if r.status_code in (200, 201):
            d = r.json().get("data", {})
            pid = d.get("id", "")
            if pid:
                return True, f"https://write.as/{pid}", "Write.as", 69
    except Exception:
        pass
    return False, "", "Write.as", 69


# ─────────────────────────────────────────────────────────────
# PLATFORM 7 — Debian Paste (DA 65)
# ─────────────────────────────────────────────────────────────
def post_debian(art):
    try:
        r = requests.post(
            "https://paste.debian.net/",
            data={
                "poster": "Scotle",
                "paste": f"# {art['title']}\n\n{art['md']}",
                "expire": "604800",
                "submit": "Submit",
            },
            headers={"User-Agent": UA, "Referer": "https://paste.debian.net/"},
            timeout=20,
        )
        if r.status_code == 200:
            m = re.search(r'(https?://paste\.debian\.net/\d+)', r.url + " " + r.text)
            if m:
                return True, m.group(1), "Debian Paste", 65
    except Exception:
        pass
    return False, "", "Debian Paste", 65


# ─────────────────────────────────────────────────────────────
# PLATFORM 8 — Codepad (DA 60)
# ─────────────────────────────────────────────────────────────
def post_codepad(art):
    try:
        r = requests.post(
            "http://codepad.org",
            data={
                "lang": "Plain Text",
                "code": f"# {art['title']}\n\n{art['md']}",
                "run": "False",
                "submit": "Submit",
            },
            headers={"User-Agent": UA, "Referer": "http://codepad.org"},
            timeout=20,
        )
        if r.status_code == 200:
            m = re.search(r'(https?://codepad\.org/\w+)', r.url + " " + r.text)
            if m and "/show/" not in m.group(1):
                return True, m.group(1), "Codepad", 60
            if "codepad.org/" in r.url and r.url != "http://codepad.org":
                return True, r.url, "Codepad", 60
    except Exception:
        pass
    return False, "", "Codepad", 60


# ─────────────────────────────────────────────────────────────
# PLATFORM 9 — Catbox.moe (DA 55) — HTML file hosting
# ─────────────────────────────────────────────────────────────
def post_catbox(art):
    try:
        content = (
            f"<html><head><title>{art['title']}</title>"
            f"<meta charset='utf-8'><style>body{{font-family:Georgia,serif;"
            f"max-width:720px;margin:40px auto;padding:0 20px;line-height:1.7}}"
            f"h1{{color:#1a1a1a}}h2{{color:#333}}a{{color:#1558D6}}</style></head>"
            f"<body>{art['html']}</body></html>"
        )
        fname = f"scotle-education-{random.randint(10000,99999)}.html"
        r = requests.post(
            "https://catbox.moe/user/api.php",
            data={"reqtype": "fileupload", "userhash": ""},
            files={"fileToUpload": (fname, content.encode("utf-8"), "text/html")},
            headers={"User-Agent": UA},
            timeout=20,
        )
        if r.status_code == 200:
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, "Catbox.moe", 55
    except Exception:
        pass
    return False, "", "Catbox.moe", 55


# ─────────────────────────────────────────────────────────────
# PLATFORM 10 — Filebin.net (DA 55)
# ─────────────────────────────────────────────────────────────
def post_filebin(art):
    try:
        bin_id = "".join(random.choices("abcdefghijklmnopqrstuvwxyz0123456789", k=14))
        fname = f"scotle-blog-{random.randint(1000,9999)}.md"
        content = f"# {art['title']}\n\n{art['md']}"
        r = requests.post(
            f"https://filebin.net/{bin_id}/{fname}",
            data=content.encode("utf-8"),
            headers={
                "Content-Type": "text/plain",
                "User-Agent": UA,
                "Accept": "application/json",
            },
            timeout=20,
        )
        if r.status_code in (200, 201):
            return True, f"https://filebin.net/{bin_id}", "Filebin.net", 55
    except Exception:
        pass
    return False, "", "Filebin.net", 55


# ─────────────────────────────────────────────────────────────
# PLATFORM 11 — Rentry.co (DA 55)
# ─────────────────────────────────────────────────────────────
def post_rentry(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        s.get("https://rentry.co", timeout=15)
        csrf = s.cookies.get("csrftoken", "")
        r = s.post(
            "https://rentry.co/api/new",
            data={
                "csrfmiddlewaretoken": csrf,
                "text": f"# {art['title']}\n\n{art['md']}",
                "edit_code": "".join(random.choices("abcdefghijklmnopqrstuvwxyz0123456789", k=12)),
            },
            headers={"Referer": "https://rentry.co"},
            timeout=20,
        )
        if r.status_code == 200:
            d = r.json()
            if d.get("status") == "200":
                url = d.get("url", "")
                if url:
                    return True, url, "Rentry.co", 55
    except Exception:
        pass
    return False, "", "Rentry.co", 55


# ─────────────────────────────────────────────────────────────
# PLATFORM 12 — paste.ee (DA 52)
# ─────────────────────────────────────────────────────────────
def post_paste_ee(art):
    try:
        r = requests.post(
            "https://paste.ee/api",
            data={
                "key": "public",
                "description": art["title"][:100],
                "paste": f"# {art['title']}\n\n{art['md']}",
                "expire": 0,
            },
            headers={"User-Agent": UA},
            timeout=20,
        )
        if r.status_code in (200, 201):
            d = r.json()
            link = d.get("link", d.get("paste", {}).get("link", ""))
            if link and link.startswith("http"):
                return True, link, "paste.ee", 52
    except Exception:
        pass
    return False, "", "paste.ee", 52


# ─────────────────────────────────────────────────────────────
# PLATFORM 13 — cl1p.net (DA 50)
# ─────────────────────────────────────────────────────────────
def post_cl1p(art):
    try:
        slug = "".join(random.choices("abcdefghijklmnopqrstuvwxyz0123456789", k=12))
        r = requests.post(
            f"https://cl1p.net/{slug}",
            data={"cl1pTextArea": f"# {art['title']}\n\n{art['md']}"},
            headers={"User-Agent": UA, "Referer": f"https://cl1p.net/{slug}"},
            timeout=15,
        )
        if r.status_code in (200, 201):
            return True, f"https://cl1p.net/{slug}", "cl1p.net", 50
    except Exception:
        pass
    return False, "", "cl1p.net", 50


# ─────────────────────────────────────────────────────────────
# PLATFORM 14 — Bytebin / pastes.dev (DA 50)
# ─────────────────────────────────────────────────────────────
def post_bytebin(art):
    content = f"# {art['title']}\n\n{art['md']}"
    for base in ["https://pastes.dev", "https://bytebin.lucko.me"]:
        try:
            r = requests.post(
                f"{base}/post",
                data=content.encode("utf-8"),
                headers={"Content-Type": "text/markdown", "User-Agent": UA},
                timeout=15,
            )
            if r.status_code in (200, 201):
                key = r.json().get("key", "")
                if key:
                    return True, f"{base}/{key}", "Bytebin", 50
        except Exception:
            pass
    return False, "", "Bytebin", 50


# ─────────────────────────────────────────────────────────────
# Ordered by DA — best first
# ─────────────────────────────────────────────────────────────
BEST_PLATFORMS = [
    (post_mozilla,    "Mozilla Paste",  97),
    (post_gist,       "GitHub Gist",    96),
    (post_telegraph,  "Telegra.ph",     93),
    (post_ubuntu,     "Ubuntu Paste",   90),
    (post_hackmd,     "HackMD",         70),
    (post_writeas,    "Write.as",       69),
    (post_debian,     "Debian Paste",   65),
    (post_codepad,    "Codepad",        60),
    (post_catbox,     "Catbox.moe",     55),
    (post_filebin,    "Filebin.net",    55),
    (post_rentry,     "Rentry.co",      55),
    (post_paste_ee,   "paste.ee",       52),
    (post_cl1p,       "cl1p.net",       50),
    (post_bytebin,    "Bytebin",        50),
]


# ─────────────────────────────────────────────────────────────
# Excel report
# ─────────────────────────────────────────────────────────────
def save_excel(results):
    pub = [r for r in results if r["success"]]
    fail = [r for r in results if not r["success"]]
    if not pub:
        print("No posts to save.")
        return

    pub_df = pd.DataFrame(pub)
    pub_df.insert(0, "S.No", range(1, len(pub_df) + 1))

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as w:
        summary_rows = [
            ("SCOTLE.ORG — HIGH DA BLOG REPORT", ""),
            ("", ""),
            ("Target", TARGET_URL),
            ("Blog Style", "Scotle High School — Education"),
            ("Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("Total Published", len(pub)),
            ("Total Failed", len(fail)),
            ("Backlinks Created", f"{len(pub)} (1 per post)"),
            ("Avg DA", f"{sum(r['da'] for r in pub)//max(len(pub),1)}"),
            ("", ""),
            ("LIVE BACKLINKS (by DA)", ""),
        ]
        for r in sorted(pub, key=lambda x: -x["da"]):
            summary_rows.append((f"  {r['platform']} (DA {r['da']})", r["url"]))

        pd.DataFrame(summary_rows, columns=["Metric", "Value"]).to_excel(
            w, sheet_name="Dashboard", index=False)
        pub_df.to_excel(w, sheet_name="Published Posts", index=False)
        pub_df[["S.No", "platform", "da", "url", "title"]].to_excel(
            w, sheet_name="Quick URL List", index=False)
        if fail:
            pd.DataFrame(fail).to_excel(w, sheet_name="Failed", index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = load_workbook(OUTPUT_XLSX)
        hfill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        lfont = Font(name="Arial", size=10, color="1565C0", underline="single")
        gfill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        gfont = Font(name="Arial", size=10, bold=True, color="2E7D32")

        for sn in wb.sheetnames:
            ws = wb[sn]
            for c in range(1, ws.max_column + 1):
                ws.cell(row=1, column=c).fill = hfill
                ws.cell(row=1, column=c).font = hfont
                ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    h = ws.cell(row=1, column=col).value
                    cell = ws.cell(row=row, column=col)
                    val = str(cell.value or "")
                    if h in ("url", "Value") and val.startswith("http"):
                        cell.font = lfont
                    if h == "da":
                        try:
                            da = int(cell.value)
                            if da >= 70:
                                cell.fill = gfill
                                cell.font = gfont
                            elif da >= 50:
                                cell.fill = PatternFill(start_color="FFF9C4",
                                    end_color="FFF9C4", fill_type="solid")
                        except Exception:
                            pass
            for c in range(1, ws.max_column + 1):
                ml = max(len(str(ws.cell(row=r, column=c).value or ""))
                         for r in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(c)].width = min(ml + 4, 70)
            ws.freeze_panes = "A2"
        wb.save(OUTPUT_XLSX)
    except Exception as e:
        logging.warning(f"Excel styling skipped: {e}")


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main(dry_run=False):
    print(f"\n{'='*65}")
    print(f"  SCOTLE.ORG — HIGH DA BLOG POSTING")
    print(f"  Target: {TARGET_URL}")
    print(f"  Platforms: {len(BEST_PLATFORMS)} (DA 50+ only)")
    if dry_run:
        print(f"  Mode: DRY RUN")
    print(f"{'='*65}\n")

    if dry_run:
        articles = get_all_scotle_articles()
        print(f"  {len(articles)} Scotle-style articles ready:\n")
        for i, a in enumerate(articles, 1):
            md = article_to_markdown(a, TARGET_URL)
            print(f"  {i:>2}. {a['title']}")
            print(f"      Words: {len(md.split())} | 1 backlink at end")
        print()
        print("  Platforms to post to (by DA):")
        for fn, name, da in BEST_PLATFORMS:
            print(f"    DA {da:>2}  {name}")
        print()
        return

    results = []
    art_index = 0

    print(f"  {'#':<3} {'DA':>4}  {'Platform':<20} {'Result'}")
    print(f"  {'-'*60}")

    for i, (fn, name, da) in enumerate(BEST_PLATFORMS):
        art = gen_article(art_index)
        art_index += 1

        title_preview = art["title"][:35]
        print(f"  [{i+1:>2}] DA{da:>3}  {name:<20} {title_preview}... ", end="", flush=True)

        ok, url, platform, actual_da = fn(art)

        status = "OK" if ok else "FAIL"
        print(f"{status}" + (f" -> {url}" if ok else ""))

        results.append({
            "success": ok,
            "platform": platform,
            "da": actual_da,
            "url": url if ok else "",
            "title": art["title"],
            "word_count": art["word_count"],
            "target": TARGET_URL,
            "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        })

        if i < len(BEST_PLATFORMS) - 1:
            time.sleep(random.uniform(5, 10))

    save_excel(results)

    pub = [r for r in results if r["success"]]
    fail_names = [r["platform"] for r in results if not r["success"]]

    print(f"\n{'='*65}")
    print(f"  DONE — {len(pub)}/{len(BEST_PLATFORMS)} posted successfully")
    avg_da = sum(r["da"] for r in pub) // max(len(pub), 1)
    print(f"  Average DA of live backlinks: {avg_da}")
    print(f"  Excel report: {OUTPUT_XLSX}")
    print(f"{'='*65}\n")

    if pub:
        print("  LIVE BACKLINKS (best first):\n")
        for r in sorted(pub, key=lambda x: -x["da"]):
            print(f"    DA {r['da']:>2}  {r['platform']:<22}  {r['url']}")

    if fail_names:
        print(f"\n  Failed ({len(fail_names)}): {', '.join(fail_names)}")

    print()
    return pub


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Preview without posting")
    args = parser.parse_args()
    main(dry_run=args.dry_run)
