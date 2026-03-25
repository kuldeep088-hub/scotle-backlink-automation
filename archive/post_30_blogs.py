"""
post_30_blogs.py
================
Posts 30 unique blog articles to multiple platforms with scotle.org backlinks.
Saves all links in Google Sheets format Excel.
"""

import os
import sys
import time
import random
import logging
import requests
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from blog_poster import blog_config as config
from blog_poster.blog_utils import (
    generate_article_from_template, get_next_topic, spin_text,
    inject_backlinks, build_author_bio, add_authority_links,
    html_to_markdown, html_to_telegraph_nodes,
)
from blog_poster.templates.article_templates import get_templates, get_all_niches

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger(__name__)

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)

OUTPUT_XLSX = os.path.join(DATA_DIR, "Scotle_30_Blog_Posts.xlsx")
TARGET_URL = config.TARGET_URL
TARGET_BRAND = config.TARGET_BRAND
KEYWORDS = config.TARGET_KEYWORDS
ANCHORS = config.ANCHOR_TEXTS

used_topics = []


def generate_unique_article():
    """Generate a unique article with scotle.org backlinks."""
    global used_topics
    niche = random.choice(get_all_niches())
    templates = get_templates(niche)
    template = random.choice(templates)
    keyword = random.choice(KEYWORDS)
    topic = get_next_topic(KEYWORDS, used_topics=used_topics)
    used_topics.append(topic)

    article = generate_article_from_template(
        template=template, keyword=keyword,
        brand=TARGET_BRAND, city="Jaipur", spin=True,
    )
    article["title"] = topic

    content = article["content_html"]
    anchor = random.choice(ANCHORS)

    content = inject_backlinks(content, TARGET_URL, ANCHORS, max_links=2, nofollow_pct=30)
    content = add_authority_links(content, config.AUTHORITY_LINKS)
    if config.ADD_AUTHOR_BIO:
        content += "\n\n" + build_author_bio(TARGET_URL, TARGET_BRAND, anchor)

    article["content_html"] = content
    article["anchor"] = anchor
    article["keyword"] = keyword
    return article


# ===========================================================
# PLATFORM POSTERS (anonymous, no signup needed)
# ===========================================================

def post_to_writeas(article):
    """Post to Write.as (anonymous, DA 69)."""
    md = html_to_markdown(article["content_html"])
    try:
        resp = requests.post(
            "https://write.as/api/posts",
            json={"title": article["title"], "body": md, "font": "sans"},
            headers={"Content-Type": "application/json"},
            timeout=30,
        )
        if resp.status_code in (200, 201):
            data = resp.json().get("data", {})
            post_id = data.get("id", "")
            return {"success": True, "url": f"https://write.as/{post_id}", "platform": "Write.as", "da": 69}
        return {"success": False, "error": f"HTTP {resp.status_code}", "platform": "Write.as"}
    except Exception as e:
        return {"success": False, "error": str(e), "platform": "Write.as"}


def post_to_telegra_ph(article):
    """Post to Telegraph (anonymous, DA 93)."""
    token_file = os.path.join(os.path.dirname(__file__), "blog_poster", "telegraph_token.txt")
    token = ""

    if os.path.exists(token_file):
        with open(token_file) as f:
            token = f.read().strip()

    if not token:
        try:
            resp = requests.post("https://api.telegra.ph/createAccount", json={
                "short_name": "Scotle", "author_name": TARGET_BRAND,
                "author_url": TARGET_URL,
            }, timeout=15)
            data = resp.json()
            if data.get("ok"):
                token = data["result"]["access_token"]
                with open(token_file, "w") as f:
                    f.write(token)
        except Exception as e:
            return {"success": False, "error": str(e), "platform": "Telegraph"}

    if not token:
        return {"success": False, "error": "No token", "platform": "Telegraph"}

    nodes = html_to_telegraph_nodes(article["content_html"])
    try:
        resp = requests.post("https://api.telegra.ph/createPage", json={
            "access_token": token, "title": article["title"][:256],
            "author_name": TARGET_BRAND, "author_url": TARGET_URL,
            "content": json.dumps(nodes), "return_content": False,
        }, timeout=15)
        data = resp.json()
        if data.get("ok"):
            url = data["result"].get("url", "")
            return {"success": True, "url": url, "platform": "Telegraph", "da": 93}
        return {"success": False, "error": data.get("error", "Unknown"), "platform": "Telegraph"}
    except Exception as e:
        return {"success": False, "error": str(e), "platform": "Telegraph"}


def post_to_rentry(article):
    """Post to Rentry.co (anonymous, DA ~55)."""
    md = html_to_markdown(article["content_html"])
    full_content = f"# {article['title']}\n\n{md}"
    edit_code = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=10))

    try:
        # Get CSRF token first
        session = requests.Session()
        resp = session.get("https://rentry.co", timeout=15)
        csrf = ""
        for line in resp.text.split("\n"):
            if "csrfmiddlewaretoken" in line and "value=" in line:
                csrf = line.split('value="')[1].split('"')[0]
                break

        if not csrf:
            # Try from cookies
            csrf = session.cookies.get("csrftoken", "")

        resp = session.post("https://rentry.co/api/new", data={
            "csrfmiddlewaretoken": csrf,
            "text": full_content,
            "edit_code": edit_code,
        }, headers={
            "Referer": "https://rentry.co",
        }, timeout=15)

        if resp.status_code == 200:
            data = resp.json()
            if data.get("status") == "200":
                url = data.get("url", "")
                return {"success": True, "url": url, "platform": "Rentry.co", "da": 55}
            return {"success": False, "error": data.get("content", "Unknown"), "platform": "Rentry.co"}
        return {"success": False, "error": f"HTTP {resp.status_code}", "platform": "Rentry.co"}
    except Exception as e:
        return {"success": False, "error": str(e), "platform": "Rentry.co"}


def post_to_friendpaste(article):
    """Post to FriendPaste (anonymous)."""
    md = html_to_markdown(article["content_html"])
    full_content = f"# {article['title']}\n\n{md}"
    try:
        resp = requests.post("https://friendpaste.com/", json={
            "title": article["title"],
            "snippet": full_content,
            "language": "markdown",
        }, headers={"Content-Type": "application/json"}, timeout=15)
        if resp.status_code in (200, 201):
            data = resp.json()
            paste_id = data.get("id", data.get("ok", ""))
            if paste_id:
                return {"success": True, "url": f"https://friendpaste.com/{paste_id}", "platform": "FriendPaste", "da": 42}
        return {"success": False, "error": f"HTTP {resp.status_code}", "platform": "FriendPaste"}
    except Exception as e:
        return {"success": False, "error": str(e), "platform": "FriendPaste"}


def post_to_dpaste(article):
    """Post to dpaste.org (anonymous)."""
    md = html_to_markdown(article["content_html"])
    full_content = f"# {article['title']}\n\n{md}"
    try:
        resp = requests.post("https://dpaste.org/api/", data={
            "content": full_content,
            "format": "url",
            "expires": "365",
        }, timeout=15)
        if resp.status_code in (200, 201):
            url = resp.text.strip()
            if url.startswith("http"):
                return {"success": True, "url": url, "platform": "dpaste.org", "da": 52}
        return {"success": False, "error": f"HTTP {resp.status_code}", "platform": "dpaste.org"}
    except Exception as e:
        return {"success": False, "error": str(e), "platform": "dpaste.org"}


def post_to_paste_mozilla(article):
    """Post to paste.mozilla.org."""
    md = html_to_markdown(article["content_html"])
    full_content = f"# {article['title']}\n\n{md}"
    try:
        resp = requests.post("https://paste.mozilla.org/api/", data={
            "content": full_content,
            "format": "url",
            "expires": "3153600000",
            "lexer": "_markdown",
        }, timeout=15)
        if resp.status_code in (200, 201):
            url = resp.text.strip()
            if url.startswith("http"):
                return {"success": True, "url": url, "platform": "Mozilla Paste", "da": 97}
        return {"success": False, "error": f"HTTP {resp.status_code}", "platform": "Mozilla Paste"}
    except Exception as e:
        return {"success": False, "error": str(e), "platform": "Mozilla Paste"}


# All platform functions in round-robin order
PLATFORMS = [
    post_to_writeas,
    post_to_telegra_ph,
    post_to_rentry,
    post_to_dpaste,
    post_to_friendpaste,
    post_to_paste_mozilla,
]


def save_excel(results):
    """Save results to formatted Google Sheets-ready Excel."""
    rows = []
    for i, r in enumerate(results, 1):
        rows.append({
            "S.No": i,
            "Platform": r.get("platform", ""),
            "DA": r.get("da", ""),
            "Title": r.get("title", ""),
            "Article URL": r.get("url", ""),
            "Target Website": TARGET_URL,
            "Anchor Text": r.get("anchor", ""),
            "Word Count": r.get("word_count", ""),
            "Status": "Published" if r.get("success") else "Failed",
            "Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        })

    df = pd.DataFrame(rows)
    published = df[df["Status"] == "Published"]
    failed = df[df["Status"] == "Failed"]

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        # Summary
        summary = pd.DataFrame({
            "Metric": [
                "SCOTLE.ORG BLOG POSTING REPORT", "",
                "Target Website", "Brand",
                "", "RESULTS",
                "Total Attempted", "Successfully Published", "Failed",
                "Success Rate", "Total Backlinks Created",
                "", "PLATFORMS USED",
            ],
            "Value": [
                "", "",
                TARGET_URL, TARGET_BRAND,
                "", "",
                len(df), len(published), len(failed),
                f"{len(published)/max(len(df),1)*100:.0f}%",
                f"{len(published) * 3} (2 in-content + 1 bio per article)",
                "", "",
            ],
        })
        # Add platform counts
        for platform in published["Platform"].unique():
            count = len(published[published["Platform"] == platform])
            summary = pd.concat([summary, pd.DataFrame({
                "Metric": [f"  {platform}"],
                "Value": [f"{count} posts"],
            })], ignore_index=True)

        summary.to_excel(writer, sheet_name="Summary", index=False)
        published.to_excel(writer, sheet_name="Published Posts", index=False)
        if len(failed) > 0:
            failed.to_excel(writer, sheet_name="Failed Posts", index=False)

    # Format
    wb = load_workbook(OUTPUT_XLSX)
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    link_font = Font(name="Arial", size=10, color="1A73E8", underline="single")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                cell = ws.cell(row=row, column=col)
                if header in ("Article URL", "Target Website") and cell.value and str(cell.value).startswith("http"):
                    cell.font = link_font
                if header == "Status" and str(cell.value) == "Published":
                    cell.fill = green_fill
                    cell.font = Font(color="137333", bold=True)

        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 55)
        ws.freeze_panes = "A2"

    wb.save(OUTPUT_XLSX)


def main():
    print("\n" + "=" * 60)
    print("  SCOTLE.ORG - 30 BLOG POSTS")
    print(f"  Target: {TARGET_URL}")
    print("=" * 60 + "\n")

    results = []
    platform_idx = 0
    attempt = 0
    max_attempts = 60  # Max attempts to get 30 posts

    while len([r for r in results if r.get("success")]) < 30 and attempt < max_attempts:
        attempt += 1
        article = generate_unique_article()

        # Pick platform (round-robin)
        poster_fn = PLATFORMS[platform_idx % len(PLATFORMS)]
        platform_idx += 1

        print(f"  [{attempt}] {poster_fn.__name__.replace('post_to_', '').upper()}: {article['title'][:50]}...", end=" ", flush=True)

        result = poster_fn(article)
        result["title"] = article["title"]
        result["anchor"] = article.get("anchor", "")
        result["word_count"] = article.get("word_count", 0)

        if result.get("success"):
            results.append(result)
            success_count = len([r for r in results if r.get("success")])
            print(f"OK [{success_count}/30] -> {result['url']}")
        else:
            results.append(result)
            print(f"FAIL -> {result.get('error', '')[:60]}")

        # Delay between posts (longer for same platform)
        delay = random.uniform(20, 35)
        if attempt < max_attempts:
            time.sleep(delay)

    # Save
    save_excel(results)

    published = [r for r in results if r.get("success")]
    print(f"\n{'=' * 60}")
    print(f"  DONE! {len(published)}/30 published")
    print(f"  Excel: {OUTPUT_XLSX}")
    print(f"{'=' * 60}")
    print(f"\n  ALL LIVE ARTICLE LINKS:\n")
    for i, r in enumerate(published, 1):
        print(f"  {i:>2}. [{r['platform']}] {r['url']}")
    print()

    # Open Excel
    import platform as plat
    if plat.system() == "Windows":
        os.startfile(os.path.abspath(OUTPUT_XLSX))


if __name__ == "__main__":
    main()
