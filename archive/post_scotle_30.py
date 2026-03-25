"""
post_scotle_30.py
=================
Post Scotle-style education blogs to 30+ websites.

Uses scotle_templates.py articles:
- 300-450 words, empathetic/educational tone
- SCOTLE HIGH SCHOOL mentioned naturally in one section
- ONE backlink only — at the very end in CTA

Run:
    python post_scotle_30.py
    python post_scotle_30.py --count 10    # post to first 10 platforms
    python post_scotle_30.py --dry-run     # show articles, don't post
"""

import os, sys, time, random, argparse, logging
import pandas as pd
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from blog_poster import blog_config as config
from blog_poster.blog_utils import html_to_markdown
from blog_poster.templates.scotle_templates import (
    get_all_scotle_articles, article_to_markdown, article_to_html
)

# Import all posting functions from post_30_sites
from post_30_sites import (
    ALL_PLATFORMS, save_excel,
    post_writeas, post_friendpaste, post_catbox, post_dpaste_com,
    post_bytebin, post_centos_paste, post_bpaste, post_controlc,
    post_ix, post_debian_paste, post_filebin, post_pastebin,
    post_gist, post_telegraph, post_hackmd, post_rentry,
    post_pastebin_pl, post_paste2, post_pastelink, post_ghostbin,
    post_pastes_io, post_notepad_pw, post_ideone,
    post_ubuntu_paste, post_mozilla_paste, post_pst_innomi,
    post_pastetext, post_nopaste, post_openstack_paste, post_justpaste,
    post_snippet_host, post_pastiebin, post_codeshare, post_pasteall,
    post_snippi, post_cl1p, post_bitbin,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)
OUTPUT_XLSX = os.path.join(DATA_DIR, "Scotle_Style_30_Sites.xlsx")

TARGET_URL = config.TARGET_URL  # https://scotle.org/


def gen_scotle_article(index: int) -> dict:
    """
    Build an art dict from a Scotle-style template.
    Rotates through all 10 templates. ONE backlink at end only.
    """
    articles = get_all_scotle_articles()
    article = articles[index % len(articles)]

    md = article_to_markdown(article, TARGET_URL)
    html = article_to_html(article, TARGET_URL)

    return {
        "title": article["title"],
        "content_html": html,
        "md": md,
        "word_count": len(md.split()),
        "anchor": TARGET_URL,
        "target": TARGET_URL,
    }


def print_dry_run(articles):
    """Print all article previews without posting."""
    print(f"\n{'='*65}")
    print("  DRY RUN — Scotle-Style Articles Preview")
    print(f"{'='*65}\n")
    for i, art in enumerate(articles, 1):
        print(f"--- Article {i} ---")
        print(f"Title: {art['title']}")
        print(f"Words: {art['word_count']}")
        print(f"Backlink: {art['anchor']}")
        print()
        # Show first 200 chars of markdown
        preview = art["md"][:200].replace("\n", " ")
        print(f"Preview: {preview}...")
        print()


def main(target_count=30, dry_run=False):
    articles = get_all_scotle_articles()

    print(f"\n{'='*65}")
    print(f"  SCOTLE.ORG — EDUCATION BLOG POSTING (SCOTLE STYLE)")
    print(f"  Target: {TARGET_URL}")
    print(f"  Templates available: {len(articles)}")
    print(f"  Platforms to try: {len(ALL_PLATFORMS)}")
    print(f"  Target posts: {target_count}")
    print(f"  Style: 1 backlink at end, SCOTLE HIGH SCHOOL in body")
    print(f"{'='*65}\n")

    if dry_run:
        sample_arts = [gen_scotle_article(i) for i in range(min(3, len(articles)))]
        print_dry_run(sample_arts)
        return

    results = []
    used_platforms = set()
    success_count = 0
    art_index = 0

    for i, platform_fn in enumerate(ALL_PLATFORMS):
        if success_count >= target_count:
            break

        art = gen_scotle_article(art_index)
        art_index += 1

        pname = platform_fn.__name__.replace("post_", "")
        title_preview = art["title"][:38]
        print(f"  [{i+1:>2}/{len(ALL_PLATFORMS)}] {pname:<22} {title_preview}... ", end="", flush=True)

        try:
            ok, url, platform, da = platform_fn(art)
        except Exception as e:
            ok, url, platform, da = False, "", pname, 0
            logging.debug(f"Exception on {pname}: {e}")

        if ok and platform not in used_platforms:
            used_platforms.add(platform)
            success_count += 1
            results.append({
                "success": True, "url": url, "platform": platform, "da": da,
                "title": art["title"], "anchor": TARGET_URL,
                "word_count": art["word_count"], "target": TARGET_URL,
                "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
            print(f"OK [{success_count}/{target_count}] -> {url}")
        elif ok and platform in used_platforms:
            print(f"SKIP (dup: {platform})")
        else:
            results.append({
                "success": False, "url": "", "platform": platform, "da": da,
                "title": art["title"], "anchor": "", "word_count": 0,
                "target": TARGET_URL, "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
            print("FAIL")

        if i < len(ALL_PLATFORMS) - 1 and success_count < target_count:
            time.sleep(random.uniform(4, 8))

    # Save Excel
    _save_scotle_excel(results)

    pub = [r for r in results if r["success"]]
    fail = [r for r in results if not r["success"]]

    print(f"\n{'='*65}")
    print(f"  DONE! {len(pub)} Scotle-style posts published")
    print(f"  Platforms: {len(set(r['platform'] for r in pub))} different sites")
    print(f"  Excel report: {OUTPUT_XLSX}")
    print(f"{'='*65}\n")

    if pub:
        print("  ALL LIVE LINKS:\n")
        for j, r in enumerate(pub, 1):
            print(f"  {j:>2}. [{r['da']:>2} DA] {r['platform']:<22} {r['url']}")

    print()
    return pub


def _save_scotle_excel(results):
    """Save results to Excel with Scotle-style report."""
    pub = [r for r in results if r["success"]]
    fail = [r for r in results if not r["success"]]

    if not pub:
        print("No posts published — nothing to save.")
        return

    pub_df = pd.DataFrame(pub)
    pub_df.insert(0, "S.No", range(1, len(pub_df) + 1))
    fail_df = pd.DataFrame(fail) if fail else pd.DataFrame()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as w:
        summary_rows = [
            ("SCOTLE.ORG - SCOTLE STYLE EDUCATION BLOG REPORT", ""),
            ("", ""),
            ("Target Website", TARGET_URL),
            ("Blog Style", "Scotle High School — Education / Parenting"),
            ("Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("Total Published", len(pub)),
            ("Total Failed", len(fail)),
            ("Sites Used", len(set(r["platform"] for r in pub))),
            ("Success Rate", f"{len(pub)/max(len(results), 1)*100:.0f}%"),
            ("Backlinks Created", f"{len(pub)} (1 per post — end CTA only)"),
            ("", ""),
            ("PLATFORMS USED", ""),
        ]
        for platform in sorted(set(r["platform"] for r in pub)):
            da = next((r["da"] for r in pub if r["platform"] == platform), "")
            summary_rows.append((f"  {platform} (DA {da})", "1 post"))

        pd.DataFrame(summary_rows, columns=["Metric", "Value"]).to_excel(
            w, sheet_name="Dashboard", index=False)
        pub_df.to_excel(w, sheet_name="Published Posts", index=False)
        pub_df[["S.No", "platform", "da", "url", "title"]].to_excel(
            w, sheet_name="Quick URL List", index=False)
        if len(fail_df) > 0:
            fail_df.to_excel(w, sheet_name="Failed Platforms", index=False)

    # Style with openpyxl
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = load_workbook(OUTPUT_XLSX)
        hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        lfont = Font(name="Arial", size=10, color="1558D6", underline="single")
        gfill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        gfont = Font(name="Arial", size=10, bold=True, color="137333")

        for sn in wb.sheetnames:
            ws = wb[sn]
            for c in range(1, ws.max_column + 1):
                ws.cell(row=1, column=c).fill = hfill
                ws.cell(row=1, column=c).font = hfont
                ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    h = ws.cell(row=1, column=col).value
                    cell = ws.cell(row=row, column=col)
                    val = str(cell.value or "")
                    if h == "url" and val.startswith("http"):
                        cell.font = lfont
                    if h == "da":
                        try:
                            da = int(cell.value)
                            if da >= 65:
                                cell.fill = gfill
                                cell.font = gfont
                            elif da >= 50:
                                cell.fill = PatternFill(start_color="FFF9C4",
                                    end_color="FFF9C4", fill_type="solid")
                        except Exception:
                            pass
            for c in range(1, ws.max_column + 1):
                ml = max(len(str(ws.cell(row=r, column=c).value or ""))
                         for r in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(c)].width = min(ml + 3, 65)
            ws.freeze_panes = "A2"
        wb.save(OUTPUT_XLSX)
    except Exception as e:
        logging.warning(f"Excel formatting skipped: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Post Scotle-style education blogs to 30 sites")
    parser.add_argument("--count", type=int, default=30,
                        help="Number of successful posts to target (default: 30)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview articles without posting")
    args = parser.parse_args()

    main(target_count=args.count, dry_run=args.dry_run)
