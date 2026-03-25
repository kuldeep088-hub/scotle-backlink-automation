"""
post_30_sites.py
================
Post blogs to 30 DIFFERENT websites — 1 post per site.
Tries 40+ platforms, saves all live links to Excel.
Target: https://scotle.org/
"""

import os, sys, time, random, json, logging, requests, re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from blog_poster import blog_config as config
from blog_poster.blog_utils import (
    generate_article_from_template, get_next_topic,
    inject_backlinks, build_author_bio, add_authority_links, html_to_markdown,
)
from blog_poster.templates.article_templates import get_templates, get_all_niches

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)
OUTPUT_XLSX = os.path.join(DATA_DIR, "Scotle_30_Different_Sites.xlsx")

used_topics = []

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"


def gen_article():
    global used_topics
    niche = random.choice(get_all_niches())
    tpl = random.choice(get_templates(niche))
    kw = random.choice(config.TARGET_KEYWORDS)
    topic = get_next_topic(config.TARGET_KEYWORDS, used_topics=used_topics)
    used_topics.append(topic)
    art = generate_article_from_template(tpl, kw, config.TARGET_BRAND, "Jaipur", True)
    art["title"] = topic
    c = art["content_html"]
    anchor = random.choice(config.ANCHOR_TEXTS)
    c = inject_backlinks(c, config.TARGET_URL, config.ANCHOR_TEXTS, 2, 30)
    c = add_authority_links(c, config.AUTHORITY_LINKS)
    c += "\n\n" + build_author_bio(config.TARGET_URL, config.TARGET_BRAND, anchor)
    art["content_html"] = c
    art["anchor"] = anchor
    art["md"] = html_to_markdown(c)
    return art


# ============================================================
# PLATFORM 1 — Write.as (DA 69) — anonymous blog
# ============================================================
def post_writeas(art):
    try:
        r = requests.post("https://write.as/api/posts",
            json={"title": art["title"], "body": art["md"], "font": "sans"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json().get("data", {})
            pid = d.get("id", "")
            if pid:
                return True, f"https://write.as/{pid}", "Write.as", 69
    except: pass
    return False, "", "Write.as", 69


# ============================================================
# PLATFORM 2 — FriendPaste (DA 42)
# ============================================================
def post_friendpaste(art):
    try:
        r = requests.post("https://friendpaste.com/",
            json={"title": art["title"],
                  "snippet": f"# {art['title']}\n\n{art['md']}",
                  "language": "markdown"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            pid = d.get("id", d.get("ok", ""))
            if pid:
                return True, f"https://friendpaste.com/{pid}", "FriendPaste", 42
    except: pass
    return False, "", "FriendPaste", 42


# ============================================================
# PLATFORM 3 — Catbox.moe (DA 55) — file hosting HTML
# ============================================================
def post_catbox(art):
    try:
        content = f"<html><head><title>{art['title']}</title><meta charset='utf-8'></head><body><h1>{art['title']}</h1>{art['content_html']}</body></html>"
        fname = f"scotle-{random.randint(10000,99999)}.html"
        r = requests.post("https://catbox.moe/user/api.php",
            data={"reqtype": "fileupload", "userhash": ""},
            files={"fileToUpload": (fname, content.encode(), "text/html")},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code == 200:
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, "Catbox.moe", 55
    except: pass
    return False, "", "Catbox.moe", 55


# ============================================================
# PLATFORM 4 — dpaste.com (DA 55) — text paste
# ============================================================
def post_dpaste_com(art):
    try:
        r = requests.post("https://dpaste.com/api/v2/",
            data={"content": f"# {art['title']}\n\n{art['md']}",
                  "syntax": "markdown", "expiry_days": "365"},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, "dpaste.com", 55
    except: pass
    return False, "", "dpaste.com", 55


# ============================================================
# PLATFORM 5 — Bytebin (lucko) (DA ~40)
# ============================================================
def post_bytebin(art):
    for base in ["https://bytebin.lucko.me", "https://pastes.dev"]:
        try:
            content = f"# {art['title']}\n\n{art['md']}"
            r = requests.post(f"{base}/post",
                data=content.encode(),
                headers={"Content-Type": "text/markdown", "User-Agent": UA}, timeout=15)
            if r.status_code in (200, 201):
                try:
                    key = r.json().get("key", "")
                    if key:
                        return True, f"{base}/{key}", "Bytebin", 40
                except:
                    loc = r.headers.get("Location", "")
                    if loc:
                        url = loc if loc.startswith("http") else f"{base}{loc}"
                        return True, url, "Bytebin", 40
        except: pass
    return False, "", "Bytebin", 40


# ============================================================
# PLATFORM 6 — paste.centos.org (DA ~60)
# ============================================================
def post_centos_paste(art):
    try:
        r = requests.post("https://paste.centos.org/api/create",
            data={"title": art["title"][:60],
                  "text": f"# {art['title']}\n\n{art['md']}",
                  "name": "Scotle Blog", "lang": "text", "private": "0",
                  "expire": "0"},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            try:
                d = r.json()
                pid = d.get("result", {}).get("id", "") or d.get("id", "")
                if pid:
                    return True, f"https://paste.centos.org/view/{pid}", "CentOS Paste", 60
            except:
                if "view" in r.text.lower() or "paste" in r.text.lower():
                    m = re.search(r'https?://paste\.centos\.org/\S+', r.text)
                    if m:
                        return True, m.group(), "CentOS Paste", 60
    except: pass
    return False, "", "CentOS Paste", 60


# ============================================================
# PLATFORM 7 — bpaste.net (DA ~50)
# ============================================================
def post_bpaste(art):
    try:
        r = requests.post("https://bpaste.net/json/new",
            data={"code": f"# {art['title']}\n\n{art['md']}",
                  "lexer": "rst", "expiry": "never"},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            paste_id = d.get("pastes", [{}])[0].get("paste_id", "")
            if paste_id:
                return True, f"https://bpaste.net/show/{paste_id}", "Bpaste.net", 50
    except: pass
    return False, "", "Bpaste.net", 50


# ============================================================
# PLATFORM 8 — controlc.com (DA ~45)
# ============================================================
def post_controlc(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://controlc.com", timeout=10)
        csrf = ""
        m = re.search(r'name=["\']_token["\'] value=["\']([^"\']+)', home.text)
        if m:
            csrf = m.group(1)
        r = s.post("https://controlc.com/index.php",
            data={"_token": csrf,
                  "paste_title": art["title"][:50],
                  "paste_data": f"# {art['title']}\n\n{art['md']}",
                  "paste_code": "0", "paste_expire": "0"},
            headers={"Referer": "https://controlc.com"}, timeout=20)
        if r.status_code == 200 and "controlc.com/" in r.url:
            return True, r.url, "ControlC", 45
    except: pass
    return False, "", "ControlC", 45


# ============================================================
# PLATFORM 9 — ix.io (DA ~55) — simple pipe paste
# ============================================================
def post_ix(art):
    try:
        content = f"# {art['title']}\n\n{art['md']}"
        r = requests.post("http://ix.io",
            data={"f:1": content.encode()},
            headers={"User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, "ix.io", 55
    except: pass
    return False, "", "ix.io", 55


# ============================================================
# PLATFORM 10 — paste.debian.net (DA ~65)
# ============================================================
def post_debian_paste(art):
    try:
        r = requests.post("https://paste.debian.net/",
            data={"poster": "Scotle",
                  "paste": f"# {art['title']}\n\n{art['md']}",
                  "expire": "604800",
                  "submit": "Submit"},
            headers={"User-Agent": UA, "Referer": "https://paste.debian.net/"}, timeout=20)
        if r.status_code == 200:
            m = re.search(r'(https?://paste\.debian\.net/\d+)', r.url + " " + r.text)
            if m:
                return True, m.group(1), "paste.debian.net", 65
    except: pass
    return False, "", "paste.debian.net", 65


# ============================================================
# PLATFORM 11 — filebin.net (DA ~55) — file hosting
# ============================================================
def post_filebin(art):
    try:
        bin_id = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=12))
        fname = f"scotle-blog-{random.randint(1000, 9999)}.md"
        content = f"# {art['title']}\n\n{art['md']}"
        r = requests.post(f"https://filebin.net/{bin_id}/{fname}",
            data=content.encode(),
            headers={"Content-Type": "text/plain", "User-Agent": UA,
                     "Accept": "application/json"}, timeout=20)
        if r.status_code in (200, 201):
            return True, f"https://filebin.net/{bin_id}", "Filebin.net", 55
    except: pass
    return False, "", "Filebin.net", 55


# ============================================================
# PLATFORM 12 — pastebin.com (DA 92) — guest paste
# ============================================================
def post_pastebin(art):
    try:
        # Guest paste with free API key from pastebin.com (no login needed for dev key)
        api_dev_key = "d2f7aa1e7a2a39c12b99d7dc6aa74c03"  # public test key
        r = requests.post("https://pastebin.com/api/api_post.php",
            data={"api_dev_key": api_dev_key,
                  "api_option": "paste",
                  "api_paste_code": f"# {art['title']}\n\n{art['md']}",
                  "api_paste_name": art["title"][:100],
                  "api_paste_expire_date": "N",
                  "api_paste_private": "0",
                  "api_paste_format": "markdown"},
            headers={"User-Agent": UA}, timeout=20)
        if r.status_code == 200 and r.text.startswith("http"):
            return True, r.text.strip(), "Pastebin.com", 92
    except: pass
    return False, "", "Pastebin.com", 92


# ============================================================
# PLATFORM 13 — GitHub Gist anonymous (DA 96)
# ============================================================
def post_gist(art):
    try:
        r = requests.post("https://api.github.com/gists",
            json={"description": art["title"],
                  "public": True,
                  "files": {
                      f"scotle-blog-{random.randint(1000,9999)}.md": {
                          "content": f"# {art['title']}\n\n{art['md']}"
                      }
                  }},
            headers={"Accept": "application/vnd.github.v3+json", "User-Agent": UA},
            timeout=20)
        if r.status_code == 201:
            url = r.json().get("html_url", "")
            if url:
                return True, url, "GitHub Gist", 96
    except: pass
    return False, "", "GitHub Gist", 96


# ============================================================
# PLATFORM 14 — Telegra.ph (DA 93)
# ============================================================
def post_telegraph(art):
    from blog_poster.blog_utils import html_to_telegraph_nodes
    tf = os.path.join(os.path.dirname(__file__), "blog_poster", "telegraph_token.txt")
    token = ""
    if os.path.exists(tf):
        with open(tf) as f: token = f.read().strip()
    if not token:
        try:
            r = requests.post("https://api.telegra.ph/createAccount",
                json={"short_name": "Scotle", "author_name": config.TARGET_BRAND,
                      "author_url": config.TARGET_URL}, timeout=10)
            d = r.json()
            if d.get("ok"):
                token = d["result"]["access_token"]
                with open(tf, "w") as f: f.write(token)
        except: return False, "", "Telegraph", 93
    if not token: return False, "", "Telegraph", 93
    nodes = html_to_telegraph_nodes(art["content_html"])
    try:
        r = requests.post("https://api.telegra.ph/createPage",
            json={"access_token": token, "title": art["title"][:256],
                  "author_name": config.TARGET_BRAND, "author_url": config.TARGET_URL,
                  "content": json.dumps(nodes), "return_content": False}, timeout=15)
        d = r.json()
        if d.get("ok"):
            return True, d["result"].get("url", ""), "Telegraph", 93
    except: pass
    return False, "", "Telegraph", 93


# ============================================================
# PLATFORM 15 — HackMD.io (DA ~70) — anonymous note
# ============================================================
def post_hackmd(art):
    try:
        r = requests.post("https://hackmd.io/new",
            data={"markdown": f"# {art['title']}\n\n{art['md']}"},
            headers={"User-Agent": UA, "Content-Type": "application/x-www-form-urlencoded"},
            allow_redirects=True, timeout=20)
        if r.status_code == 200 and "hackmd.io" in r.url and "/new" not in r.url:
            return True, r.url, "HackMD.io", 70
    except: pass
    return False, "", "HackMD.io", 70


# ============================================================
# PLATFORM 16 — Rentry.co (DA 55) — markdown paste
# ============================================================
def post_rentry(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        s.get("https://rentry.co", timeout=10)
        csrf = s.cookies.get("csrftoken", "")
        r = s.post("https://rentry.co/api/new",
            data={"csrfmiddlewaretoken": csrf,
                  "text": f"# {art['title']}\n\n{art['md']}",
                  "edit_code": ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=10))},
            headers={"Referer": "https://rentry.co"}, timeout=15)
        if r.status_code == 200:
            d = r.json()
            if d.get("status") == "200":
                return True, d.get("url", ""), "Rentry.co", 55
    except: pass
    return False, "", "Rentry.co", 55


# ============================================================
# PLATFORM 17 — Pastebin.pl (DA ~40)
# ============================================================
def post_pastebin_pl(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://pastebin.pl", timeout=10)
        csrf = ""
        m = re.search(r'name="csrf_token"\s+value="([^"]+)"', home.text)
        if m: csrf = m.group(1)
        r = s.post("https://pastebin.pl/submit",
            data={"csrf_token": csrf,
                  "title": art["title"][:80],
                  "content": f"# {art['title']}\n\n{art['md']}",
                  "format": "text", "expire": "0"},
            headers={"Referer": "https://pastebin.pl"}, timeout=20)
        if r.status_code == 200 and "/view/" in r.url:
            return True, r.url, "Pastebin.pl", 40
    except: pass
    return False, "", "Pastebin.pl", 40


# ============================================================
# PLATFORM 18 — paste2.org (DA ~45)
# ============================================================
def post_paste2(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://paste2.org", timeout=10)
        r = s.post("https://paste2.org",
            data={"lang": "Plain Text",
                  "code": f"# {art['title']}\n\n{art['md']}",
                  "parent": "0"},
            headers={"Referer": "https://paste2.org",
                     "Content-Type": "application/x-www-form-urlencoded"}, timeout=20)
        if r.status_code == 200 and "paste2.org/" in r.url and r.url != "https://paste2.org/":
            return True, r.url, "paste2.org", 45
    except: pass
    return False, "", "paste2.org", 45


# ============================================================
# PLATFORM 19 — pastelink.net (DA ~42)
# ============================================================
def post_pastelink(art):
    try:
        r = requests.post("https://pastelink.net/api/paste",
            json={"content": f"# {art['title']}\n\n{art['md']}",
                  "title": art["title"]},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            link = d.get("link", d.get("url", d.get("id", "")))
            if link:
                if not link.startswith("http"):
                    link = f"https://pastelink.net/{link}"
                return True, link, "pastelink.net", 42
    except: pass
    return False, "", "pastelink.net", 42


# ============================================================
# PLATFORM 20 — ghostbin.org (DA ~45)
# ============================================================
def post_ghostbin(art):
    for base in ["https://ghostbin.org", "https://ghostbin.co", "https://ghostbin.com"]:
        try:
            r = requests.post(f"{base}/paste/new",
                data={"text": f"# {art['title']}\n\n{art['md']}",
                      "expire": "2d",
                      "lang": "markdown"},
                headers={"User-Agent": UA, "Referer": base}, timeout=15)
            if r.status_code in (200, 201) and base.replace("https://", "") in r.url:
                return True, r.url, "Ghostbin", 45
        except: pass
    return False, "", "Ghostbin", 45


# ============================================================
# PLATFORM 21 — pastes.io (DA ~45)
# ============================================================
def post_pastes_io(art):
    try:
        r = requests.post("https://pastes.io/api/paste",
            json={"title": art["title"], "content": f"# {art['title']}\n\n{art['md']}"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            slug = d.get("slug", d.get("key", d.get("id", "")))
            if slug:
                return True, f"https://pastes.io/{slug}", "pastes.io", 45
    except: pass
    return False, "", "pastes.io", 45


# ============================================================
# PLATFORM 22 — notepad.pw (DA ~40)
# ============================================================
def post_notepad_pw(art):
    try:
        slug = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=8))
        r = requests.post(f"https://notepad.pw/save/{slug}",
            data={"text": f"# {art['title']}\n\n{art['md']}"},
            headers={"User-Agent": UA, "Referer": "https://notepad.pw"}, timeout=20)
        if r.status_code in (200, 201, 302):
            return True, f"https://notepad.pw/{slug}", "notepad.pw", 40
    except: pass
    return False, "", "notepad.pw", 40


# ============================================================
# PLATFORM 23 — write.as (2nd post — different account token, skip if already used)
# PLATFORM 23 — Telegra.ph already in #14, use ide.one instead
# ============================================================
def post_ideone(art):
    try:
        r = requests.post("https://ideone.com/api/1/submit",
            json={"clientId": "guest", "clientSecret": "guest",
                  "sourceCode": f"# {art['title']}\n\n{art['md']}",
                  "language": 62, "input": ""},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            link = d.get("result", {}).get("link", "") or d.get("link", "")
            if link:
                if not link.startswith("http"):
                    link = f"https://ideone.com/{link}"
                return True, link, "Ideone.com", 50
    except: pass
    return False, "", "Ideone.com", 50


# ============================================================
# PLATFORM 24 — quickforget.com (DA ~38)
# ============================================================
def post_quickforget(art):
    try:
        r = requests.post("https://quickforget.com/api",
            json={"data": f"# {art['title']}\n\n{art['md']}", "expire": 0},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            d = r.json()
            key = d.get("key", d.get("id", ""))
            if key:
                return True, f"https://quickforget.com/{key}", "quickforget.com", 38
    except: pass
    return False, "", "quickforget.com", 38


# ============================================================
# PLATFORM 25 — stashpaste.com / pasteofcode.org (DA ~40)
# ============================================================
def post_paste_ofcode(art):
    try:
        r = requests.post("https://paste.ofcode.org/",
            data={"code": f"# {art['title']}\n\n{art['md']}",
                  "lang": "text", "notabot": "moo"},
            headers={"User-Agent": UA, "Referer": "https://paste.ofcode.org/"}, timeout=15)
        if r.status_code == 200 and "paste.ofcode.org/" in r.url:
            return True, r.url, "paste.ofcode.org", 40
    except: pass
    return False, "", "paste.ofcode.org", 40


# ============================================================
# PLATFORM 26 — tutpaste.com (DA ~38)
# ============================================================
def post_tutpaste(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("http://tutpaste.com", timeout=10)
        csrf = ""
        m = re.search(r'name=["\']_token["\'] value=["\']([^"\']+)', home.text)
        if m: csrf = m.group(1)
        r = s.post("http://tutpaste.com",
            data={"_token": csrf, "title": art["title"][:80],
                  "body": f"# {art['title']}\n\n{art['md']}",
                  "syntax": "text"},
            headers={"Referer": "http://tutpaste.com"}, timeout=20)
        if r.status_code == 200 and "tutpaste.com/" in r.url:
            return True, r.url, "tutpaste.com", 38
    except: pass
    return False, "", "tutpaste.com", 38


# ============================================================
# PLATFORM 27 — codepad.org (DA ~60)
# ============================================================
def post_codepad(art):
    try:
        r = requests.post("http://codepad.org",
            data={"lang": "Plain Text",
                  "code": f"# {art['title']}\n\n{art['md']}",
                  "run": "False",
                  "submit": "Submit"},
            headers={"User-Agent": UA, "Referer": "http://codepad.org"}, timeout=20)
        if r.status_code == 200 and "codepad.org/" in r.url and "/show/" in r.url:
            return True, r.url, "codepad.org", 60
        if r.status_code == 200:
            m = re.search(r'(https?://codepad\.org/\w+)', r.text)
            if m:
                return True, m.group(1), "codepad.org", 60
    except: pass
    return False, "", "codepad.org", 60


# ============================================================
# PLATFORM 28 — dumpz.org (DA ~40)
# ============================================================
def post_dumpz(art):
    try:
        r = requests.post("http://dumpz.org/api/dump",
            json={"content": f"# {art['title']}\n\n{art['md']}", "lexer": "text"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            d = r.json()
            key = d.get("key", "")
            if key:
                return True, f"http://dumpz.org/{key}/", "dumpz.org", 40
    except: pass
    return False, "", "dumpz.org", 40


# ============================================================
# PLATFORM 29 — bitbin.it (DA ~42)
# ============================================================
def post_bitbin(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://bitbin.it", timeout=10)
        csrf = s.cookies.get("XSRF-TOKEN", "")
        r = s.post("https://bitbin.it/post/",
            data={"content": f"# {art['title']}\n\n{art['md']}",
                  "_token": csrf, "lang": "text", "expire": "never"},
            headers={"Referer": "https://bitbin.it", "X-CSRF-TOKEN": csrf}, timeout=20)
        if r.status_code == 200 and "bitbin.it/" in r.url and r.url != "https://bitbin.it/":
            return True, r.url, "bitbin.it", 42
    except: pass
    return False, "", "bitbin.it", 42


# ============================================================
# PLATFORM 30 — snippet.host (DA ~42)
# ============================================================
def post_snippet_host(art):
    try:
        r = requests.post("https://snippet.host/api/create",
            json={"content": f"# {art['title']}\n\n{art['md']}", "listed": True},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            d = r.json()
            key = d.get("key", d.get("id", ""))
            if key:
                return True, f"https://snippet.host/{key}", "snippet.host", 42
    except: pass
    return False, "", "snippet.host", 42


# ============================================================
# PLATFORM 31 — pastiebin.com (DA ~38)
# ============================================================
def post_pastiebin(art):
    try:
        r = requests.post("https://www.pastiebin.com/paste",
            data={"paste_content": f"# {art['title']}\n\n{art['md']}",
                  "paste_title": art["title"][:60],
                  "paste_expire": "never"},
            headers={"User-Agent": UA, "Referer": "https://www.pastiebin.com"}, timeout=15)
        if r.status_code == 200 and "pastiebin.com/" in r.url:
            return True, r.url, "pastiebin.com", 38
    except: pass
    return False, "", "pastiebin.com", 38


# ============================================================
# PLATFORM 32 — codeshare.io (DA ~55)
# ============================================================
def post_codeshare(art):
    try:
        r = requests.post("https://codeshare.io/",
            json={"value": f"# {art['title']}\n\n{art['md']}", "mode": "markdown"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            d = r.json()
            code = d.get("code", d.get("id", ""))
            if code:
                return True, f"https://codeshare.io/{code}", "codeshare.io", 55
    except: pass
    return False, "", "codeshare.io", 55


# ============================================================
# PLATFORM 33 — pasteall.org (DA ~50)
# ============================================================
def post_pasteall(art):
    try:
        r = requests.post("https://www.pasteall.org/paste/create",
            data={"language": "text",
                  "paste_content": f"# {art['title']}\n\n{art['md']}"},
            headers={"User-Agent": UA, "Referer": "https://www.pasteall.org/paste/"},
            allow_redirects=True, timeout=20)
        if r.status_code == 200 and "pasteall.org/paste/" in r.url:
            return True, r.url, "pasteall.org", 50
    except: pass
    return False, "", "pasteall.org", 50


# ============================================================
# PLATFORM 34 — Snippi.com (DA ~42)
# ============================================================
def post_snippi(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://snippi.com", timeout=10)
        csrf = ""
        m = re.search(r'_token.*?value="([^"]+)"', home.text)
        if m: csrf = m.group(1)
        r = s.post("https://snippi.com/save",
            data={"_token": csrf,
                  "title": art["title"][:60],
                  "code": f"# {art['title']}\n\n{art['md']}",
                  "language": "text"},
            headers={"Referer": "https://snippi.com"}, timeout=20)
        if r.status_code == 200 and "snippi.com/" in r.url and r.url != "https://snippi.com/":
            return True, r.url, "Snippi.com", 42
    except: pass
    return False, "", "Snippi.com", 42


# ============================================================
# PLATFORM 35 — cl1p.net (DA ~50)
# ============================================================
def post_cl1p(art):
    try:
        slug = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=10))
        r = requests.post(f"https://cl1p.net/{slug}",
            data={"cl1pTextArea": f"# {art['title']}\n\n{art['md']}"},
            headers={"User-Agent": UA, "Referer": f"https://cl1p.net/{slug}"}, timeout=15)
        if r.status_code in (200, 201):
            return True, f"https://cl1p.net/{slug}", "cl1p.net", 50
    except: pass
    return False, "", "cl1p.net", 50


# ============================================================
# PLATFORM 36 — paste.ubuntu.com (DA ~90)
# ============================================================
def post_ubuntu_paste(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://paste.ubuntu.com", timeout=10)
        csrf = ""
        m = re.search(r"csrfmiddlewaretoken.*?value='([^']+)'", home.text)
        if not m:
            m = re.search(r'csrfmiddlewaretoken.*?value="([^"]+)"', home.text)
        if m: csrf = m.group(1)
        if not csrf:
            csrf = s.cookies.get("csrftoken", "")
        r = s.post("https://paste.ubuntu.com/",
            data={"csrfmiddlewaretoken": csrf,
                  "poster": "Scotle",
                  "syntax": "text",
                  "content": f"# {art['title']}\n\n{art['md']}",
                  "expiration": "week"},
            headers={"Referer": "https://paste.ubuntu.com/"}, timeout=20)
        if r.status_code == 200 and "paste.ubuntu.com/p/" in r.url:
            return True, r.url, "paste.ubuntu.com", 90
    except: pass
    return False, "", "paste.ubuntu.com", 90


# ============================================================
# PLATFORM 37 — Mozilla Paste (DA 97)
# ============================================================
def post_mozilla_paste(art):
    try:
        r = requests.post("https://paste.mozilla.org/api/",
            data={"content": f"# {art['title']}\n\n{art['md']}",
                  "format": "url", "expires": "3153600000", "lexer": "_markdown"},
            headers={"User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            url = r.text.strip()
            if url.startswith("http"):
                return True, url, "Mozilla Paste", 97
    except: pass
    return False, "", "Mozilla Paste", 97


# ============================================================
# PLATFORM 38 — pst.innomi.net (DA ~38)
# ============================================================
def post_pst_innomi(art):
    try:
        r = requests.post("https://pst.innomi.net/paste/new",
            data={"text": f"# {art['title']}\n\n{art['md']}", "ttl": 0, "lang": "text"},
            headers={"User-Agent": UA, "Referer": "https://pst.innomi.net"}, timeout=15)
        if r.status_code == 200 and "pst.innomi.net/paste/" in r.url:
            return True, r.url, "pst.innomi.net", 38
    except: pass
    return False, "", "pst.innomi.net", 38


# ============================================================
# PLATFORM 39 — pastetext.net (DA ~40)
# ============================================================
def post_pastetext(art):
    try:
        r = requests.post("https://pastetext.net/",
            data={"text": f"# {art['title']}\n\n{art['md']}", "name": art["title"][:60]},
            headers={"User-Agent": UA, "Referer": "https://pastetext.net/"}, timeout=15)
        if r.status_code == 200 and "pastetext.net/" in r.url and r.url != "https://pastetext.net/":
            return True, r.url, "pastetext.net", 40
    except: pass
    return False, "", "pastetext.net", 40


# ============================================================
# PLATFORM 40 — nopaste.net (DA ~38)
# ============================================================
def post_nopaste(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://nopaste.net", timeout=10)
        csrf = ""
        m = re.search(r'name=["\']_token["\'] value=["\']([^"\']+)', home.text)
        if m: csrf = m.group(1)
        r = s.post("https://nopaste.net",
            data={"_token": csrf, "title": art["title"][:60],
                  "content": f"# {art['title']}\n\n{art['md']}",
                  "syntax": "text", "expire": "0"},
            headers={"Referer": "https://nopaste.net"}, timeout=20)
        if r.status_code == 200 and "nopaste.net/" in r.url and r.url != "https://nopaste.net/":
            return True, r.url, "nopaste.net", 38
    except: pass
    return False, "", "nopaste.net", 38


# ============================================================
# PLATFORM 41 — paste.openstack.org (DA ~85)
# ============================================================
def post_openstack_paste(art):
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://paste.openstack.org", timeout=10)
        csrf = ""
        m = re.search(r'name=["\']_csrf_token["\'] value=["\']([^"\']+)', home.text)
        if m: csrf = m.group(1)
        r = s.post("https://paste.openstack.org/api/pastes",
            json={"code": f"# {art['title']}\n\n{art['md']}", "language": "text", "expire": "1week"},
            headers={"Content-Type": "application/json", "Referer": "https://paste.openstack.org"}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            pid = d.get("paste_id", d.get("id", ""))
            if pid:
                return True, f"https://paste.openstack.org/show/{pid}", "paste.openstack.org", 85
    except: pass
    return False, "", "paste.openstack.org", 85


# ============================================================
# PLATFORM 42 — paste.centos.org backup path
# PLATFORM 42 — just-paste.it (DA ~45)
# ============================================================
def post_justpaste(art):
    try:
        r = requests.post("https://just-paste.it/documents",
            data=f"# {art['title']}\n\n{art['md']}".encode(),
            headers={"Content-Type": "text/plain", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            d = r.json()
            key = d.get("key", "")
            if key:
                return True, f"https://just-paste.it/{key}", "just-paste.it", 45
    except: pass
    return False, "", "just-paste.it", 45


# ============================================================
# All platforms list — 42 platforms, 1 post each
# ============================================================
ALL_PLATFORMS = [
    post_writeas,         # 1  DA 69
    post_friendpaste,     # 2  DA 42
    post_catbox,          # 3  DA 55
    post_dpaste_com,      # 4  DA 55
    post_bytebin,         # 5  DA 40
    post_centos_paste,    # 6  DA 60
    post_bpaste,          # 7  DA 50
    post_controlc,        # 8  DA 45
    post_ix,              # 9  DA 55
    post_debian_paste,    # 10 DA 65
    post_filebin,         # 11 DA 55
    post_pastebin,        # 12 DA 92
    post_gist,            # 13 DA 96
    post_telegraph,       # 14 DA 93
    post_hackmd,          # 15 DA 70
    post_rentry,          # 16 DA 55
    post_pastebin_pl,     # 17 DA 40
    post_paste2,          # 18 DA 45
    post_pastelink,       # 19 DA 42
    post_ghostbin,        # 20 DA 45
    post_pastes_io,       # 21 DA 45
    post_notepad_pw,      # 22 DA 40
    post_ideone,          # 23 DA 50
    post_quickforget,     # 24 DA 38
    post_paste_ofcode,    # 25 DA 40
    post_tutpaste,        # 26 DA 38
    post_codepad,         # 27 DA 60
    post_dumpz,           # 28 DA 40
    post_bitbin,          # 29 DA 42
    post_snippet_host,    # 30 DA 42
    post_pastiebin,       # 31 DA 38
    post_codeshare,       # 32 DA 55
    post_pasteall,        # 33 DA 50
    post_snippi,          # 34 DA 42
    post_cl1p,            # 35 DA 50
    post_ubuntu_paste,    # 36 DA 90
    post_mozilla_paste,   # 37 DA 97
    post_pst_innomi,      # 38 DA 38
    post_pastetext,       # 39 DA 40
    post_nopaste,         # 40 DA 38
    post_openstack_paste, # 41 DA 85
    post_justpaste,       # 42 DA 45
]


def save_excel(results):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    pub = [r for r in results if r["success"]]
    fail = [r for r in results if not r["success"]]

    if not pub:
        print("No published posts to save.")
        return

    pub_df = pd.DataFrame(pub)
    pub_df.insert(0, "S.No", range(1, len(pub_df) + 1))

    fail_df = pd.DataFrame(fail) if fail else pd.DataFrame()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as w:
        # Summary
        summary_rows = [
            ("SCOTLE.ORG - 30 DIFFERENT BLOG SITES REPORT", ""),
            ("", ""),
            ("Target Website", config.TARGET_URL),
            ("Brand", config.TARGET_BRAND),
            ("Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("Total Published", len(pub)),
            ("Total Failed", len(fail)),
            ("Different Sites Used", len(set(r["platform"] for r in pub))),
            ("Success Rate", f"{len(pub)/max(len(results),1)*100:.0f}%"),
            ("Total Backlinks Created", f"{len(pub)*3} (3 per post)"),
            ("", ""),
            ("PLATFORMS USED", ""),
        ]
        for platform in sorted(set(r["platform"] for r in pub)):
            da = next((r["da"] for r in pub if r["platform"] == platform), "")
            summary_rows.append((f"  {platform} (DA {da})", "1 post"))

        summary = pd.DataFrame(summary_rows, columns=["Metric", "Value"])
        summary.to_excel(w, sheet_name="Dashboard", index=False)
        pub_df.to_excel(w, sheet_name="All 30 Published Posts", index=False)
        pub_df[["S.No", "platform", "da", "url", "title", "target"]].to_excel(
            w, sheet_name="Quick URL List", index=False)
        if len(fail_df) > 0:
            fail_df.to_excel(w, sheet_name="Failed Platforms", index=False)

    # Format
    wb = load_workbook(OUTPUT_XLSX)
    hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    gfill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    lfont = Font(name="Arial", size=10, color="1558D6", underline="single")
    gfont = Font(name="Arial", size=10, bold=True, color="137333")

    for sn in wb.sheetnames:
        ws = wb[sn]
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=col).value
                cell = ws.cell(row=row, column=col)
                val = str(cell.value or "")
                if h in ("url", "Article URL") and val.startswith("http"):
                    cell.font = lfont
                if h == "da":
                    try:
                        da = int(cell.value)
                        if da >= 65:
                            cell.fill = gfill
                            cell.font = gfont
                        elif da >= 50:
                            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    except: pass
        for c in range(1, ws.max_column + 1):
            ml = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(c)].width = min(ml + 3, 65)
        ws.freeze_panes = "A2"
    wb.save(OUTPUT_XLSX)


def main():
    print(f"\n{'='*62}")
    print(f"  SCOTLE.ORG — POSTING TO 30 DIFFERENT WEBSITES")
    print(f"  Target: {config.TARGET_URL}")
    print(f"  Total platforms to try: {len(ALL_PLATFORMS)}")
    print(f"{'='*62}\n")

    results = []
    used_platforms = set()
    success_count = 0

    for i, platform_fn in enumerate(ALL_PLATFORMS):
        if success_count >= 30:
            break

        art = gen_article()
        pname = platform_fn.__name__.replace("post_", "")
        print(f"  [{i+1:>2}/{len(ALL_PLATFORMS)}] {pname:<22} {art['title'][:40]}... ", end="", flush=True)

        ok, url, platform, da = platform_fn(art)

        if ok and platform not in used_platforms:
            used_platforms.add(platform)
            success_count += 1
            results.append({
                "success": True, "url": url, "platform": platform, "da": da,
                "title": art["title"], "anchor": art.get("anchor", ""),
                "word_count": art.get("word_count", 0),
                "target": config.TARGET_URL,
                "date": datetime.now().strftime("%Y-%m-%d %H:%M")
            })
            print(f"OK [{success_count}/30] -> {url}")
        elif ok and platform in used_platforms:
            print(f"SKIP (duplicate platform: {platform})")
        else:
            results.append({
                "success": False, "url": "", "platform": platform, "da": da,
                "title": art["title"], "anchor": "", "word_count": 0,
                "target": config.TARGET_URL, "date": datetime.now().strftime("%Y-%m-%d %H:%M")
            })
            print("FAIL")

        if i < len(ALL_PLATFORMS) - 1 and success_count < 30:
            time.sleep(random.uniform(5, 10))

    # Save
    save_excel(results)

    pub = [r for r in results if r["success"]]
    print(f"\n{'='*62}")
    print(f"  DONE! {len(pub)} posts on {len(set(r['platform'] for r in pub))} different websites")
    print(f"  Excel: {OUTPUT_XLSX}")
    print(f"{'='*62}\n")
    print(f"  ALL LIVE ARTICLE LINKS:\n")
    for j, r in enumerate(pub, 1):
        print(f"  {j:>2}. [{r['platform']} DA:{r['da']}] {r['url']}")
    print()

    import platform as pl
    if pl.system() == "Windows":
        os.startfile(os.path.abspath(OUTPUT_XLSX))


if __name__ == "__main__":
    main()
