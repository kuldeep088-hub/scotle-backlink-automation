"""
smart_poster.py
===============
Finds working blog platforms, posts humanized articles, saves to Excel.

Features:
  - Humanizer  : makes template content sound natural and human-written
  - 12 platforms: focuses on actual blogs (not broken paste sites)
  - Excel output: saves to data/Scotle_Smart_Posts_YYYYMMDD_HHMM.xlsx

Usage:
    python smart_poster.py              # Post to all working platforms
    python smart_poster.py --test       # Test platforms only (no real posts)
    python smart_poster.py --count 2    # Post 2 articles per platform
"""

import os, sys, re, json, random, time, logging, argparse, requests
from datetime import datetime

# ---- make sure parent package is importable ----
sys.path.insert(0, os.path.dirname(__file__))

from blog_poster.blog_config import (
    TARGET_URL, TARGET_BRAND, TARGET_KEYWORDS, ANCHOR_TEXTS,
    AUTHORITY_LINKS, MAX_BACKLINKS_PER_POST, NOFOLLOW_PERCENTAGE,
    ADD_AUTHOR_BIO,
    MEDIUM_CONFIG, DEVTO_CONFIG, HASHNODE_CONFIG,
    TUMBLR_CONFIG, WRITEAS_CONFIG,
)
from blog_poster.blog_utils import (
    generate_article_from_template,
    inject_backlinks, build_author_bio, add_authority_links,
    spin_text, html_to_telegraph_nodes, html_to_markdown,
)
from blog_poster.templates.article_templates import get_templates, get_all_niches

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════════════
#  HUMANIZER
# ════════════════════════════════════════════════════════════════

_CONTRACTIONS = {
    " you are ": " you're ", " it is ": " it's ", " they are ": " they're ",
    " we are ": " we're ", " do not ": " don't ", " does not ": " doesn't ",
    " is not ": " isn't ", " are not ": " aren't ", " will not ": " won't ",
    " cannot ": " can't ", " should not ": " shouldn't ", " would not ": " wouldn't ",
    " could not ": " couldn't ", " did not ": " didn't ", " have not ": " haven't ",
    " has not ": " hasn't ", " that is ": " that's ", " there is ": " there's ",
    " here is ": " here's ", " it has ": " it's ", " who is ": " who's ",
}

_OPENERS = [
    "Here's something worth knowing — ",
    "The truth is, ",
    "Think about it this way: ",
    "What really matters here is that ",
    "Simply put, ",
    "Let's be clear: ",
    "In practice, ",
    "For most people, ",
    "The reality is that ",
    "Here's the key insight: ",
    "Worth mentioning — ",
    "To put it plainly, ",
]

_TRANSITIONS = [
    "But here's where it gets interesting. ",
    "And that's not all. ",
    "What's more, ",
    "That said, ",
    "Keep in mind that ",
    "Worth noting: ",
    "On top of that, ",
    "And this is crucial — ",
    "Now here's the part most people miss: ",
]

_RHETORICAL = [
    "But what does this actually mean for you?",
    "So why does this matter?",
    "How do you make this work in practice?",
    "Sound familiar?",
    "The question is — are you taking full advantage of this?",
    "What's the real takeaway here?",
]

_STATS = [
    "Studies show that 73% of businesses that invest in this see measurable results within 90 days.",
    "According to industry research, companies that adopt this approach grow 2× faster on average.",
    "A 2025 survey found that 8 out of 10 professionals now consider this a top priority.",
    "Data from recent studies indicates a 45% improvement in outcomes for those who follow these practices consistently.",
    "Research confirms that early adopters of this approach generate 60% more leads than those who wait.",
]

_CLOSING_PUNCHES = [
    " — and that's backed by results.",
    ", which makes all the difference.",
    ". This single change can transform your outcomes.",
    " — the numbers don't lie.",
    ", and the results speak for themselves.",
]


def humanize_content(html: str) -> str:
    """
    Make template HTML content sound natural and human-written.
    Applies: contractions, varied openers, transitions, rhetorical
    questions, realistic stats, and punchy closing sentences.
    """
    # 1. Contractions
    for formal, short in _CONTRACTIONS.items():
        html = html.replace(formal, short)

    # 2. Natural openers for ~25% of body paragraphs (skip first)
    parts = html.split("</p>")
    for i in range(1, len(parts)):
        if "<p>" in parts[i] and "<h" not in parts[i] and random.random() < 0.28:
            opener = random.choice(_OPENERS)
            parts[i] = parts[i].replace("<p>", f"<p>{opener}", 1)
    html = "</p>".join(parts)

    # 3. One rhetorical question injected mid-article
    mid = html.find("</p>", len(html) // 2)
    if mid > 0 and random.random() < 0.8:
        q = random.choice(_RHETORICAL)
        html = html[:mid + 4] + f"\n\n<p><em>{q}</em></p>" + html[mid + 4:]

    # 4. One realistic statistic after the first paragraph
    first_end = html.find("</p>")
    if first_end > 0 and random.random() < 0.7:
        stat = random.choice(_STATS)
        html = html[:first_end + 4] + f"\n\n<p>{stat}</p>" + html[first_end + 4:]

    # 5. Punchy closing added to a sentence before the last </p>
    if random.random() < 0.5:
        last_p = html.rfind("</p>", 0, html.rfind("</p>"))  # second-to-last </p>
        if last_p > 0:
            punch = random.choice(_CLOSING_PUNCHES)
            html = html[:last_p] + punch + html[last_p:]

    # 6. Transition phrase after one random <h3> section
    h3_positions = [m.start() for m in re.finditer(r"<h3>", html)]
    if h3_positions and random.random() < 0.5:
        pick = random.choice(h3_positions[1:]) if len(h3_positions) > 1 else h3_positions[0]
        close_p = html.find("</p>", pick)
        if close_p > 0:
            trans = random.choice(_TRANSITIONS)
            html = html[:close_p + 4] + f"\n<p>{trans}</p>" + html[close_p + 4:]

    return html


def build_humanized_article(niche: str = None, keyword: str = None, city: str = "") -> dict:
    """Generate a template article, humanize it, inject backlinks."""
    niche = niche or random.choice(get_all_niches())
    keyword = keyword or random.choice(TARGET_KEYWORDS)
    template = random.choice(get_templates(niche))

    art = generate_article_from_template(
        template, keyword=keyword, brand=TARGET_BRAND, city=city, spin=True
    )

    art["content_html"] = humanize_content(art["content_html"])

    anchor = random.choice(ANCHOR_TEXTS)
    art["content_html"] = inject_backlinks(
        art["content_html"], TARGET_URL, ANCHOR_TEXTS,
        max_links=MAX_BACKLINKS_PER_POST, nofollow_pct=NOFOLLOW_PERCENTAGE,
    )
    art["content_html"] = add_authority_links(art["content_html"], AUTHORITY_LINKS)

    if ADD_AUTHOR_BIO:
        art["content_html"] += "\n\n" + build_author_bio(TARGET_URL, TARGET_BRAND, anchor)

    plain = re.sub(r"<[^>]+>", "", art["content_html"])
    art["word_count"] = len(plain.split())
    art["anchor_text"] = anchor
    art["niche"] = niche
    art["keyword"] = keyword
    return art


# ════════════════════════════════════════════════════════════════
#  SHARED SESSION
# ════════════════════════════════════════════════════════════════

_SESSION = requests.Session()
_SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/html, */*",
})

# All poster functions return: (success, url, platform, da, error)
def _ok(url, platform, da):
    return True, url, platform, da, ""

def _fail(platform, da, error):
    return False, "", platform, da, str(error)[:120]


# ════════════════════════════════════════════════════════════════
#  PLATFORM 1 — Write.as  (DA 69, anonymous, NO credentials needed)
# ════════════════════════════════════════════════════════════════

def post_writeas(art: dict) -> tuple:
    platform, da = "Write.as", 69
    try:
        body_md = html_to_markdown(art["content_html"])
        token = WRITEAS_CONFIG.get("access_token", "")
        headers = {"Content-Type": "application/json"}
        if token:
            headers["Authorization"] = f"Token {token}"

        resp = _SESSION.post(
            "https://write.as/api/posts",
            headers=headers,
            json={"title": art["title"], "body": body_md, "font": "sans"},
            timeout=30,
        )
        data = resp.json()
        if resp.status_code in (200, 201):
            post_id = data.get("data", {}).get("id", "")
            url = f"https://write.as/{post_id}"
            return _ok(url, platform, da)
        return _fail(platform, da, data.get("error_msg", f"HTTP {resp.status_code}"))
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 2 — Telegraph  (DA 93, anonymous, NO credentials needed)
# ════════════════════════════════════════════════════════════════

_telegraph_token = {"value": ""}

def _get_telegraph_token() -> str:
    if _telegraph_token["value"]:
        return _telegraph_token["value"]

    token_file = os.path.join(os.path.dirname(__file__), "blog_poster", "telegraph_token.txt")
    if os.path.exists(token_file):
        with open(token_file) as f:
            tok = f.read().strip()
        if tok:
            _telegraph_token["value"] = tok
            return tok

    try:
        resp = _SESSION.post(
            "https://api.telegra.ph/createAccount",
            json={"short_name": TARGET_BRAND[:32], "author_name": TARGET_BRAND, "author_url": TARGET_URL},
            timeout=8,
        )
        data = resp.json()
        if data.get("ok"):
            tok = data["result"]["access_token"]
            _telegraph_token["value"] = tok
            os.makedirs(os.path.dirname(token_file), exist_ok=True)
            with open(token_file, "w") as f:
                f.write(tok)
            return tok
    except Exception as e:
        logger.warning(f"Telegraph token error: {e}")
    return ""


def post_telegraph(art: dict) -> tuple:
    platform, da = "Telegraph", 93
    try:
        token = _get_telegraph_token()
        if not token:
            return _fail(platform, da, "Could not obtain token (ISP block?)")

        nodes = html_to_telegraph_nodes(art["content_html"])
        resp = _SESSION.post(
            "https://api.telegra.ph/createPage",
            json={
                "access_token": token,
                "title": art["title"][:256],
                "author_name": TARGET_BRAND,
                "author_url": TARGET_URL,
                "content": json.dumps(nodes),
            },
            timeout=30,
        )
        data = resp.json()
        if data.get("ok"):
            return _ok(data["result"].get("url", ""), platform, da)
        return _fail(platform, da, data.get("error", "Unknown error"))
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 3 — Dev.to  (DA 90, free API key)
# ════════════════════════════════════════════════════════════════

def post_devto(art: dict) -> tuple:
    platform, da = "Dev.to", 90
    api_key = DEVTO_CONFIG.get("api_key", "")
    if not api_key:
        return _fail(platform, da, "api_key not set — get free key at dev.to/settings/extensions")

    try:
        body_md = html_to_markdown(art["content_html"])
        tags = ["business", "productivity", "growth", "startup"][:4]
        resp = _SESSION.post(
            "https://dev.to/api/articles",
            headers={"api-key": api_key, "Content-Type": "application/json"},
            json={"article": {
                "title": art["title"],
                "body_markdown": body_md,
                "published": True,
                "tags": tags,
            }},
            timeout=30,
        )
        data = resp.json()
        if resp.status_code in (200, 201):
            return _ok(data.get("url", ""), platform, da)
        return _fail(platform, da, data.get("error", f"HTTP {resp.status_code}"))
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 4 — Medium  (DA 95, free integration token)
# ════════════════════════════════════════════════════════════════

def post_medium(art: dict) -> tuple:
    platform, da = "Medium", 95
    token = MEDIUM_CONFIG.get("integration_token", "")
    if not token:
        return _fail(platform, da, "integration_token not set — get at medium.com/me/settings")

    try:
        me = _SESSION.get(
            "https://api.medium.com/v1/me",
            headers={"Authorization": f"Bearer {token}"},
            timeout=15,
        )
        if me.status_code != 200:
            return _fail(platform, da, f"Auth failed HTTP {me.status_code}")

        user_id = me.json().get("data", {}).get("id", "")
        full_html = f"<h1>{art['title']}</h1>\n{art['content_html']}"

        resp = _SESSION.post(
            f"https://api.medium.com/v1/users/{user_id}/posts",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json={
                "title": art["title"],
                "contentFormat": "html",
                "content": full_html,
                "tags": ["business", "productivity", "marketing"],
                "publishStatus": "public",
            },
            timeout=30,
        )
        data = resp.json()
        if resp.status_code in (200, 201):
            return _ok(data.get("data", {}).get("url", ""), platform, da)
        errors = data.get("errors", [{}])
        return _fail(platform, da, errors[0].get("message", f"HTTP {resp.status_code}"))
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 5 — Hashnode  (DA 79, free API token)
# ════════════════════════════════════════════════════════════════

def post_hashnode(art: dict) -> tuple:
    platform, da = "Hashnode", 79
    token = HASHNODE_CONFIG.get("token", "")
    pub_id = HASHNODE_CONFIG.get("publication_id", "")
    if not token or not pub_id:
        return _fail(platform, da, "token or publication_id not set — see hashnode.com/settings/developer")

    try:
        content_md = html_to_markdown(art["content_html"])
        mutation = """
        mutation PublishPost($input: PublishPostInput!) {
            publishPost(input: $input) { post { id url } }
        }"""
        resp = _SESSION.post(
            "https://gql.hashnode.com",
            headers={"Authorization": token},
            json={"query": mutation, "variables": {"input": {
                "title": art["title"],
                "contentMarkdown": content_md,
                "publicationId": pub_id,
                "tags": [],
            }}},
            timeout=30,
        )
        data = resp.json()
        post = (data.get("data") or {}).get("publishPost", {}).get("post", {})
        if post:
            return _ok(post.get("url", ""), platform, da)
        errors = data.get("errors", [])
        return _fail(platform, da, errors[0].get("message", "No post returned") if errors else "No post returned")
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 6 — Tumblr  (DA 81, OAuth1)
# ════════════════════════════════════════════════════════════════

def post_tumblr(art: dict) -> tuple:
    platform, da = "Tumblr", 81
    blog_name = TUMBLR_CONFIG.get("blog_name", "")
    ck = TUMBLR_CONFIG.get("consumer_key", "")
    if not blog_name or not ck:
        return _fail(platform, da, "blog_name or consumer_key not set in blog_config.py")

    try:
        from requests_oauthlib import OAuth1
        auth = OAuth1(
            TUMBLR_CONFIG["consumer_key"], TUMBLR_CONFIG["consumer_secret"],
            TUMBLR_CONFIG["oauth_token"],  TUMBLR_CONFIG["oauth_secret"],
        )
        resp = _SESSION.post(
            f"https://api.tumblr.com/v2/blog/{blog_name}/post",
            auth=auth,
            data={
                "type": "text",
                "title": art["title"],
                "body": art["content_html"],
                "format": "html",
                "tags": "business,marketing,productivity",
            },
            timeout=30,
        )
        data = resp.json()
        if resp.status_code in (200, 201):
            pid = str(data.get("response", {}).get("id", ""))
            url = f"https://{blog_name}.tumblr.com/post/{pid}"
            return _ok(url, platform, da)
        return _fail(platform, da, data.get("meta", {}).get("msg", f"HTTP {resp.status_code}"))
    except ImportError:
        return _fail(platform, da, "Run: pip install requests-oauthlib")
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 7 — ix.io  (DA 52, anonymous, ultra-simple API)
# ════════════════════════════════════════════════════════════════

def post_ixio(art: dict) -> tuple:
    platform, da = "ix.io", 52
    try:
        plain = re.sub(r"<[^>]+>", "", art["content_html"])
        content = f"{art['title']}\n{'='*60}\n\n{plain}\n\n---\n{TARGET_URL}"
        resp = _SESSION.post(
            "http://ix.io",
            data={"f:1": content},
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=20,
        )
        if resp.status_code == 200:
            url = resp.text.strip()
            if url.startswith("http"):
                return _ok(url, platform, da)
        return _fail(platform, da, f"HTTP {resp.status_code}: {resp.text[:60]}")
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 8 — paste.rs  (DA 35, anonymous, raw body POST)
# ════════════════════════════════════════════════════════════════

def post_pasterns(art: dict) -> tuple:
    platform, da = "paste.rs", 35
    try:
        plain = re.sub(r"<[^>]+>", "", art["content_html"])
        content = f"{art['title']}\n{'='*60}\n\n{plain}\n\n---\n{TARGET_URL}"
        resp = _SESSION.post(
            "https://paste.rs/",
            data=content.encode("utf-8"),
            headers={"Content-Type": "text/plain; charset=utf-8"},
            timeout=20,
        )
        if resp.status_code in (200, 201):
            url = resp.text.strip()
            if url.startswith("http"):
                return _ok(url, platform, da)
        return _fail(platform, da, f"HTTP {resp.status_code}: {resp.text[:60]}")
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 9 — dpaste.com  (DA 55, simple API — different from dpaste.org)
# ════════════════════════════════════════════════════════════════

def post_dpaste_com(art: dict) -> tuple:
    platform, da = "dpaste.com", 55
    try:
        plain = re.sub(r"<[^>]+>", "", art["content_html"])
        content = f"{art['title']}\n{'='*50}\n\n{plain}\n\n---\n{TARGET_URL}"
        resp = _SESSION.post(
            "https://dpaste.com/api/v2/",
            data={
                "content": content,
                "syntax": "text",
                "title": art["title"][:50],
                "expiry_days": 365,
            },
            timeout=25,
        )
        if resp.status_code in (200, 201):
            url = resp.text.strip().strip('"')
            if url.startswith("http"):
                return _ok(url, platform, da)
        return _fail(platform, da, f"HTTP {resp.status_code}: {resp.text[:80]}")
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 10 — haste.zneix.eu  (DA 30, hastebin-style)
# ════════════════════════════════════════════════════════════════

def post_haste_zneix(art: dict) -> tuple:
    platform, da = "haste.zneix.eu", 30
    try:
        plain = re.sub(r"<[^>]+>", "", art["content_html"])
        content = f"{art['title']}\n{'='*60}\n\n{plain}\n\n---\n{TARGET_URL}"
        resp = _SESSION.post(
            "https://haste.zneix.eu/documents",
            data=content.encode("utf-8"),
            headers={"Content-Type": "text/plain"},
            timeout=20,
        )
        if resp.status_code == 200:
            key = resp.json().get("key", "")
            if key:
                return _ok(f"https://haste.zneix.eu/{key}", platform, da)
        return _fail(platform, da, f"HTTP {resp.status_code}")
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 11 — paste.myst.rs  (DA 40, free JSON API)
# ════════════════════════════════════════════════════════════════

def post_pastemyst(art: dict) -> tuple:
    platform, da = "paste.myst.rs", 40
    try:
        plain = re.sub(r"<[^>]+>", "", art["content_html"])
        content = f"{art['title']}\n{'='*60}\n\n{plain}\n\n---\n{TARGET_URL}"
        resp = _SESSION.post(
            "https://paste.myst.rs/api/v2/paste",
            headers={"Content-Type": "application/json"},
            json={
                "expiry_date": "never",
                "pasties": [{"title": art["title"][:50], "language": "plain text", "code": content}],
            },
            timeout=20,
        )
        if resp.status_code in (200, 201):
            paste_id = resp.json().get("_id", "")
            if paste_id:
                return _ok(f"https://paste.myst.rs/{paste_id}", platform, da)
        return _fail(platform, da, f"HTTP {resp.status_code}: {resp.text[:80]}")
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 12 — bytebin.lucko.me  (DA 42, simple POST API)
# ════════════════════════════════════════════════════════════════

def post_bytebin(art: dict) -> tuple:
    platform, da = "bytebin.lucko.me", 42
    try:
        plain = re.sub(r"<[^>]+>", "", art["content_html"])
        content = f"{art['title']}\n{'='*60}\n\n{plain}\n\n---\n{TARGET_URL}"
        resp = _SESSION.post(
            "https://bytebin.lucko.me/post",
            data=content.encode("utf-8"),
            headers={"Content-Type": "text/plain; charset=utf-8"},
            timeout=20,
        )
        if resp.status_code in (200, 201):
            key = resp.json().get("key", "")
            if key:
                return _ok(f"https://bytebin.lucko.me/{key}", platform, da)
        return _fail(platform, da, f"HTTP {resp.status_code}: {resp.text[:80]}")
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM 12b — hastebin.skyra.pw  (hastebin instance, replaces ix.io)
# ════════════════════════════════════════════════════════════════

def post_haste_skyra(art: dict) -> tuple:
    platform, da = "hastebin.skyra.pw", 32
    try:
        plain = re.sub(r"<[^>]+>", "", art["content_html"])
        content = f"{art['title']}\n{'='*60}\n\n{plain}\n\n---\n{TARGET_URL}"
        resp = _SESSION.post(
            "https://hastebin.skyra.pw/documents",
            data=content.encode("utf-8"),
            headers={"Content-Type": "text/plain"},
            timeout=20,
        )
        if resp.status_code == 200:
            key = resp.json().get("key", "")
            if key:
                return _ok(f"https://hastebin.skyra.pw/{key}", platform, da)
        return _fail(platform, da, f"HTTP {resp.status_code}")
    except Exception as e:
        return _fail(platform, da, e)


# ════════════════════════════════════════════════════════════════
#  PLATFORM TABLE  (sorted by DA descending)
# ════════════════════════════════════════════════════════════════

PLATFORMS = [
    ("Medium",             95, post_medium),
    ("Telegraph",          93, post_telegraph),
    ("Dev.to",             90, post_devto),
    ("Tumblr",             81, post_tumblr),
    ("Hashnode",           79, post_hashnode),
    ("Write.as",           69, post_writeas),
    ("dpaste.com",         55, post_dpaste_com),
    ("bytebin.lucko.me",   42, post_bytebin),
    ("paste.myst.rs",      40, post_pastemyst),
    ("paste.rs",           35, post_pasterns),
    ("haste.zneix.eu",     30, post_haste_zneix),
    ("hastebin.skyra.pw",  32, post_haste_skyra),
]


# ════════════════════════════════════════════════════════════════
#  RUNNER
# ════════════════════════════════════════════════════════════════

RESULTS: list[dict] = []


def _record(platform, da, art, url, status, error=""):
    RESULTS.append({
        "Platform":    platform,
        "DA":          da,
        "Title":       art["title"][:90],
        "Article URL": url,
        "Anchor Text": art.get("anchor_text", ""),
        "Word Count":  art.get("word_count", 0),
        "Niche":       art.get("niche", ""),
        "Status":      status,
        "Error":       error[:120] if error else "",
        "Date":        datetime.now().strftime("%Y-%m-%d %H:%M"),
    })


def run_all(count: int = 1, test_only: bool = False):
    published, failed = [], []

    for (pname, da, fn) in PLATFORMS:
        for i in range(count):
            art = build_humanized_article()

            if test_only:
                logger.info(f"[DRY-RUN] {pname}: would post → {art['title'][:60]}")
                _record(pname, da, art, "DRY RUN", "Dry Run")
                continue

            logger.info(f"→ Posting #{i+1} to {pname} ...")
            try:
                success, url, platform, da2, error = fn(art)
            except Exception as exc:
                success, url, platform, da2, error = False, "", pname, da, str(exc)

            if success:
                logger.info(f"  ✓ {platform}: {url}")
                _record(platform, da2, art, url, "Published")
                published.append((platform, url))
            else:
                logger.warning(f"  ✗ {platform}: {error}")
                _record(platform, da2, art, "", "Failed", error)
                failed.append((platform, error))

            delay = random.uniform(6, 15)
            logger.info(f"  Waiting {delay:.0f}s...")
            time.sleep(delay)

    return published, failed


# ════════════════════════════════════════════════════════════════
#  EXCEL SAVER
# ════════════════════════════════════════════════════════════════

def save_to_excel(results: list, out_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"

    published = [r for r in results if r["Status"] == "Published"]
    failed    = [r for r in results if r["Status"] == "Failed"]
    skipped   = [r for r in results if r["Status"] == "Dry Run"]

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    summary_rows = [
        ("Scotle.org Smart Blog Posts Report", ""),
        ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total Attempts:", len(results)),
        ("✓ Published:", len(published)),
        ("✗ Failed:", len(failed)),
        ("⊘ Dry Run:", len(skipped)),
        ("Success Rate:", f"{len(published)/len(results)*100:.0f}%" if results else "0%"),
        ("", ""),
        ("Top Published Posts:", ""),
    ]
    for r in published[:8]:
        summary_rows.append((f"  {r['Platform']} (DA {r['DA']})", r["Article URL"]))

    if failed:
        summary_rows.append(("", ""))
        summary_rows.append(("Failed Platforms:", ""))
        for r in failed[:6]:
            summary_rows.append((f"  {r['Platform']}", r["Error"]))

    for row_data in summary_rows:
        ws.append(list(row_data))

    ws["A1"].font = Font(bold=True, size=14, color="1A73E8")
    ws["A1"].alignment = Alignment(horizontal="left")

    # Hyperlink published URLs in summary
    for row_i in range(11, 11 + len(published)):
        cell = ws.cell(row_i, 2)
        val = cell.value or ""
        if val.startswith("http"):
            cell.hyperlink = val
            cell.font = Font(color="1558D6", underline="single")

    # ---- Posts sheet ----
    wp = wb.create_sheet("All Posts")

    headers = ["S.No", "Platform", "DA", "Title", "Article URL",
               "Anchor Text", "Word Count", "Niche", "Status", "Error", "Date"]
    wp.append(headers)

    BLUE   = PatternFill("solid", fgColor="1A73E8")
    GREEN  = PatternFill("solid", fgColor="D4EDDA")
    RED    = PatternFill("solid", fgColor="FCE8E6")
    GREY   = PatternFill("solid", fgColor="F8F9FA")
    thin   = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
    )

    for col, h in enumerate(headers, 1):
        c = wp.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = BLUE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

    wp.row_dimensions[1].height = 22
    wp.freeze_panes = "A2"

    for i, r in enumerate(results, 1):
        row = [
            i, r["Platform"], r["DA"], r["Title"], r["Article URL"],
            r["Anchor Text"], r["Word Count"], r["Niche"],
            r["Status"], r["Error"], r["Date"],
        ]
        wp.append(row)
        row_no = i + 1

        fill = GREEN if r["Status"] == "Published" else (RED if r["Status"] == "Failed" else GREY)
        for col in range(1, len(headers) + 1):
            c = wp.cell(row_no, col)
            c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border = thin

        # Clickable URL
        url_cell = wp.cell(row_no, 5)
        if r["Article URL"] and r["Article URL"] not in ("", "DRY RUN"):
            url_cell.hyperlink = r["Article URL"]
            url_cell.font = Font(color="1558D6", underline="single")

        # DA coloring
        da_cell = wp.cell(row_no, 3)
        da = r["DA"]
        if da >= 80:
            da_cell.font = Font(bold=True, color="0D652D")
        elif da >= 60:
            da_cell.font = Font(color="1E7E34")
        else:
            da_cell.font = Font(color="856404")

    # Auto column widths
    col_widths = [6, 14, 5, 50, 55, 28, 10, 14, 12, 35, 16]
    for col, width in enumerate(col_widths, 1):
        wp.column_dimensions[get_column_letter(col)].width = width

    # ---- Published-only sheet ----
    wpp = wb.create_sheet("Published Only")
    pub_headers = ["S.No", "Platform", "DA", "Title", "Live URL", "Anchor Text", "Date"]
    wpp.append(pub_headers)
    for col, h in enumerate(pub_headers, 1):
        c = wpp.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0D652D")
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate([r for r in results if r["Status"] == "Published"], 1):
        wpp.append([i, r["Platform"], r["DA"], r["Title"], r["Article URL"], r["Anchor Text"], r["Date"]])
        url_cell = wpp.cell(i + 1, 5)
        if r["Article URL"]:
            url_cell.hyperlink = r["Article URL"]
            url_cell.font = Font(color="1558D6", underline="single")

    for col, w in zip(range(1, 8), [6, 14, 5, 55, 60, 28, 16]):
        wpp.column_dimensions[get_column_letter(col)].width = w

    wpp.freeze_panes = "A2"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    logger.info(f"Excel saved → {out_path}")
    return out_path


# ════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Smart Blog Poster with Humanizer — posts to 12 platforms",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--test",   action="store_true", help="Dry run — no actual posting")
    parser.add_argument("--count",  type=int, default=1,  help="Articles per platform (default 1)")
    parser.add_argument("--niche",  type=str, default=None, help="Force a content niche (business/local_services/technology/how_to/health/education)")
    parser.add_argument("--keyword",type=str, default=None, help="Force a specific keyword")
    args = parser.parse_args()

    date_str   = datetime.now().strftime("%Y%m%d_%H%M")
    out_path   = os.path.join(
        os.path.dirname(__file__), "data",
        f"Scotle_Smart_Posts_{date_str}.xlsx",
    )

    logger.info("=" * 60)
    logger.info(f" Smart Poster  |  {len(PLATFORMS)} platforms  |  count={args.count}")
    logger.info(f" Target URL   : {TARGET_URL}")
    logger.info(f" Test mode    : {args.test}")
    logger.info("=" * 60)

    # Patch niche/keyword if forced
    if args.niche or args.keyword:
        import functools
        _orig = build_humanized_article
        def _patched(niche=None, keyword=None, city=""):
            return _orig(niche=args.niche or niche, keyword=args.keyword or keyword, city=city)
        import smart_poster as _self
        _self.build_humanized_article = _patched

    published, failed = run_all(count=args.count, test_only=args.test)

    logger.info("")
    logger.info("=" * 60)
    logger.info(f"  Published : {len(published)}")
    logger.info(f"  Failed    : {len(failed)}")
    logger.info("=" * 60)

    if RESULTS:
        save_to_excel(RESULTS, out_path)
        print(f"\nExcel report saved to:\n  {out_path}")
    else:
        logger.warning("No results to save.")

    return published, failed


if __name__ == "__main__":
    main()
