"""
mega_post.py - Post 30 blogs to multiple high-DA platforms
"""
import os, sys, time, random, json, logging, requests, re
import pandas as pd
from datetime import datetime
from urllib.parse import quote

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from blog_poster import blog_config as config
from blog_poster.blog_utils import (
    generate_article_from_template, get_next_topic,
    inject_backlinks, build_author_bio, add_authority_links, html_to_markdown,
)
from blog_poster.templates.article_templates import get_templates, get_all_niches

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)
OUTPUT_XLSX = os.path.join(DATA_DIR, "Scotle_30_Blogs_Report.xlsx")

used_topics = []

def gen_article():
    global used_topics
    niche = random.choice(get_all_niches())
    tpl = random.choice(get_templates(niche))
    kw = random.choice(config.TARGET_KEYWORDS)
    topic = get_next_topic(config.TARGET_KEYWORDS, used_topics=used_topics)
    used_topics.append(topic)
    art = generate_article_from_template(tpl, kw, config.TARGET_BRAND, "Jaipur", True)
    art["title"] = topic
    c = art["content_html"]
    anchor = random.choice(config.ANCHOR_TEXTS)
    c = inject_backlinks(c, config.TARGET_URL, config.ANCHOR_TEXTS, 2, 30)
    c = add_authority_links(c, config.AUTHORITY_LINKS)
    c += "\n\n" + build_author_bio(config.TARGET_URL, config.TARGET_BRAND, anchor)
    art["content_html"] = c
    art["anchor"] = anchor
    art["md"] = html_to_markdown(c)
    return art

# ========== PLATFORM POSTERS ==========

def post_writeas(art):
    """Write.as DA 69"""
    try:
        r = requests.post("https://write.as/api/posts",
            json={"title": art["title"], "body": art["md"], "font": "sans"},
            headers={"Content-Type": "application/json"}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json().get("data", {})
            return True, f"https://write.as/{d.get('id','')}", "Write.as", 69
    except: pass
    return False, "", "Write.as", 69

def post_telegraph(art):
    """Telegraph DA 93"""
    from blog_poster.blog_utils import html_to_telegraph_nodes
    tf = os.path.join(os.path.dirname(__file__), "blog_poster", "telegraph_token.txt")
    token = ""
    if os.path.exists(tf):
        with open(tf) as f: token = f.read().strip()
    if not token:
        try:
            r = requests.post("https://api.telegra.ph/createAccount",
                json={"short_name":"Scotle","author_name":config.TARGET_BRAND,"author_url":config.TARGET_URL}, timeout=10)
            d = r.json()
            if d.get("ok"):
                token = d["result"]["access_token"]
                with open(tf,"w") as f: f.write(token)
        except: return False, "", "Telegraph", 93
    if not token: return False, "", "Telegraph", 93
    nodes = html_to_telegraph_nodes(art["content_html"])
    try:
        r = requests.post("https://api.telegra.ph/createPage",
            json={"access_token":token,"title":art["title"][:256],"author_name":config.TARGET_BRAND,
                  "author_url":config.TARGET_URL,"content":json.dumps(nodes),"return_content":False}, timeout=10)
        d = r.json()
        if d.get("ok"): return True, d["result"].get("url",""), "Telegraph", 93
    except: pass
    return False, "", "Telegraph", 93

def post_friendpaste(art):
    """FriendPaste DA 42"""
    try:
        r = requests.post("https://friendpaste.com/",
            json={"title":art["title"],"snippet":f"# {art['title']}\n\n{art['md']}","language":"markdown"},
            headers={"Content-Type":"application/json"}, timeout=15)
        if r.status_code in (200,201):
            d = r.json(); pid = d.get("id",d.get("ok",""))
            if pid: return True, f"https://friendpaste.com/{pid}", "FriendPaste", 42
    except: pass
    return False, "", "FriendPaste", 42

def post_rentry(art):
    """Rentry.co DA 55"""
    try:
        s = requests.Session()
        s.get("https://rentry.co", timeout=10)
        csrf = s.cookies.get("csrftoken","")
        r = s.post("https://rentry.co/api/new",
            data={"csrfmiddlewaretoken":csrf,"text":f"# {art['title']}\n\n{art['md']}",
                  "edit_code":''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789',k=10))},
            headers={"Referer":"https://rentry.co"}, timeout=15)
        if r.status_code == 200:
            d = r.json()
            if d.get("status")=="200": return True, d.get("url",""), "Rentry.co", 55
    except: pass
    return False, "", "Rentry.co", 55

def post_notepin(art):
    """Notepin DA ~45"""
    try:
        slug = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz',k=8))
        r = requests.post("https://notepin.co/api/posts",
            json={"title":art["title"],"body":art["md"],"blog":slug},
            headers={"Content-Type":"application/json"}, timeout=15)
        if r.status_code in (200,201):
            d = r.json()
            url = d.get("url", f"https://notepin.co/{slug}")
            return True, url, "Notepin", 45
    except: pass
    return False, "", "Notepin", 45

def post_txtfyi(art):
    """txt.fyi DA ~40"""
    try:
        r = requests.post("https://txt.fyi/api/post",
            data={"content": f"# {art['title']}\n\n{art['md']}"},
            timeout=15)
        if r.status_code in (200,201):
            url = r.text.strip() if r.text.startswith("http") else r.json().get("url","")
            if url: return True, url, "txt.fyi", 40
    except: pass
    return False, "", "txt.fyi", 40

def post_paste_ee(art):
    """paste.ee DA ~48"""
    try:
        r = requests.post("https://api.paste.ee/v1/pastes",
            json={"sections":[{"name":art["title"],"contents":f"# {art['title']}\n\n{art['md']}"}]},
            headers={"Content-Type":"application/json","X-Auth-Token":"public"},
            timeout=15)
        if r.status_code in (200,201):
            d = r.json()
            if d.get("success"): return True, d.get("link",""), "Paste.ee", 48
    except: pass
    return False, "", "Paste.ee", 48

def post_dpaste_com(art):
    """dpaste.com DA ~55"""
    try:
        r = requests.post("https://dpaste.com/api/v2/",
            data={"content":f"# {art['title']}\n\n{art['md']}","syntax":"markdown","expiry_days":"365"},
            timeout=15)
        if r.status_code in (200,201):
            url = r.text.strip()
            if url.startswith("http"): return True, url, "dpaste.com", 55
    except: pass
    return False, "", "dpaste.com", 55

def post_hastebin(art):
    """Hastebin DA ~52"""
    try:
        r = requests.post("https://hastebin.com/documents",
            data=f"# {art['title']}\n\n{art['md']}".encode(),
            headers={"Content-Type":"text/plain"}, timeout=15)
        if r.status_code in (200,201):
            d = r.json(); key = d.get("key","")
            if key: return True, f"https://hastebin.com/{key}", "Hastebin", 52
    except: pass
    return False, "", "Hastebin", 52

def post_paste_rs(art):
    """paste.rs DA ~38"""
    try:
        r = requests.post("https://paste.rs/",
            data=f"# {art['title']}\n\n{art['md']}".encode(),
            headers={"Content-Type":"text/plain"}, timeout=15)
        if r.status_code in (200,201):
            url = r.text.strip()
            if url.startswith("http"): return True, url, "paste.rs", 38
    except: pass
    return False, "", "paste.rs", 38

def post_sprunge(art):
    """sprunge.us DA ~42"""
    try:
        r = requests.post("http://sprunge.us",
            data={"sprunge":f"# {art['title']}\n\n{art['md']}"},timeout=15)
        if r.status_code in (200,201):
            url = r.text.strip()
            if url.startswith("http"): return True, url, "sprunge.us", 42
    except: pass
    return False, "", "sprunge.us", 42

def post_0x0(art):
    """0x0.st file hosting DA ~45"""
    try:
        content = f"# {art['title']}\n\n{art['md']}"
        r = requests.post("https://0x0.st",
            files={"file": (f"scotle-{random.randint(1000,9999)}.md", content.encode(), "text/markdown")},
            timeout=15)
        if r.status_code in (200,201):
            url = r.text.strip()
            if url.startswith("http"): return True, url, "0x0.st", 45
    except: pass
    return False, "", "0x0.st", 45

def post_catbox(art):
    """catbox.moe file hosting DA ~55"""
    try:
        content = f"<html><head><title>{art['title']}</title></head><body>{art['content_html']}</body></html>"
        r = requests.post("https://catbox.moe/user/api.php",
            data={"reqtype":"fileupload","userhash":""},
            files={"fileToUpload":(f"scotle-blog-{random.randint(1000,9999)}.html", content.encode(),"text/html")},
            timeout=15)
        if r.status_code == 200:
            url = r.text.strip()
            if url.startswith("http"): return True, url, "Catbox.moe", 55
    except: pass
    return False, "", "Catbox.moe", 55

def post_telegra_ph_direct(art):
    """Direct Telegraph page via form post"""
    try:
        s = requests.Session()
        s.headers.update({"User-Agent":"Mozilla/5.0 Chrome/120.0.0.0"})
        r = s.get("https://edit.telegra.ph/", timeout=10)
        if r.status_code == 200:
            # Extract save hash
            token_match = re.search(r'tph_token\s*=\s*["\']([^"\']+)', r.text)
            if token_match:
                token = token_match.group(1)
                return post_telegraph(art)  # Use API with discovered token
    except: pass
    return False, "", "Telegraph", 93


ALL_POSTERS = [
    post_writeas, post_telegraph, post_friendpaste, post_rentry,
    post_dpaste_com, post_hastebin, post_paste_ee,
    post_paste_rs, post_0x0, post_catbox, post_sprunge,
    post_notepin, post_txtfyi,
]


def save_excel(results):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    pub = [r for r in results if r["success"]]
    fail = [r for r in results if not r["success"]]

    pub_df = pd.DataFrame(pub)
    if len(pub_df) > 0:
        pub_df.insert(0, "S.No", range(1, len(pub_df)+1))

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as w:
        # Summary
        summary = pd.DataFrame({
            "Metric": ["SCOTLE.ORG - 30 BLOG POSTS REPORT","","Target Website","Brand","Date","",
                       "RESULTS","Total Published","Total Failed","Success Rate","Total Backlinks (3 per post)","",
                       "PLATFORMS"],
            "Value": ["","",config.TARGET_URL,config.TARGET_BRAND,datetime.now().strftime("%Y-%m-%d %H:%M"),"",
                      "",len(pub),len(fail),f"{len(pub)/max(len(results),1)*100:.0f}%",f"{len(pub)*3}","",""]
        })
        for platform in set(r["platform"] for r in pub):
            cnt = sum(1 for r in pub if r["platform"]==platform)
            da = next((r["da"] for r in pub if r["platform"]==platform), "")
            summary = pd.concat([summary, pd.DataFrame({"Metric":[f"  {platform} (DA {da})"],"Value":[f"{cnt} posts"]})], ignore_index=True)
        summary.to_excel(w, sheet_name="Summary", index=False)

        # Published
        if len(pub_df) > 0:
            pub_df.to_excel(w, sheet_name="All Published Posts", index=False)

    # Format
    wb = load_workbook(OUTPUT_XLSX)
    hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    gfill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    lfont = Font(name="Arial", size=10, color="1A73E8", underline="single")

    for sn in wb.sheetnames:
        ws = wb[sn]
        for c in range(1, ws.max_column+1):
            ws.cell(row=1,column=c).fill = hfill
            ws.cell(row=1,column=c).font = hfont
            ws.cell(row=1,column=c).alignment = Alignment(horizontal="center")
        for r in range(2, ws.max_row+1):
            for c in range(1, ws.max_column+1):
                h = ws.cell(row=1,column=c).value
                cell = ws.cell(row=r,column=c)
                if h == "url" and cell.value and str(cell.value).startswith("http"):
                    cell.font = lfont
                if h == "success" and cell.value == True:
                    cell.fill = gfill
        for c in range(1, ws.max_column+1):
            ml = max(len(str(ws.cell(row=r,column=c).value or "")) for r in range(1,ws.max_row+1))
            ws.column_dimensions[get_column_letter(c)].width = min(ml+3, 60)
        ws.freeze_panes = "A2"
    wb.save(OUTPUT_XLSX)


def main():
    print(f"\n{'='*60}\n  SCOTLE.ORG - POSTING 30 BLOGS\n  Target: {config.TARGET_URL}\n{'='*60}\n")

    results = []
    pidx = 0

    while len([r for r in results if r["success"]]) < 30 and pidx < 90:
        art = gen_article()
        fn = ALL_POSTERS[pidx % len(ALL_POSTERS)]
        pname = fn.__name__.replace("post_","")
        pidx += 1

        print(f"  [{pidx:>2}] {pname:<15} {art['title'][:45]}...", end=" ", flush=True)
        ok, url, platform, da = fn(art)

        results.append({"success":ok,"url":url,"platform":platform,"da":da,
                        "title":art["title"],"anchor":art.get("anchor",""),
                        "word_count":art.get("word_count",0),
                        "target":config.TARGET_URL,
                        "date":datetime.now().strftime("%Y-%m-%d %H:%M")})

        pub_count = len([r for r in results if r["success"]])
        if ok:
            print(f"OK [{pub_count}/30] -> {url}")
        else:
            print(f"FAIL")

        time.sleep(random.uniform(8, 15))

    save_excel(results)
    pub = [r for r in results if r["success"]]

    print(f"\n{'='*60}")
    print(f"  DONE! {len(pub)} blogs published")
    print(f"  Excel saved: {OUTPUT_XLSX}")
    print(f"{'='*60}\n")
    for i,r in enumerate(pub,1):
        print(f"  {i:>2}. [{r['platform']} DA:{r['da']}] {r['url']}")
    print()

    import platform as pl
    if pl.system()=="Windows": os.startfile(os.path.abspath(OUTPUT_XLSX))

if __name__=="__main__":
    main()
