"""
big_poster.py
=============
Posts humanized Scotle.org backlink articles to 40+ platforms.
Finds every working site and saves all live URLs to Excel.

Usage:
    python big_poster.py
    python big_poster.py --count 2   # post 2 articles per platform
"""

import os, sys, re, json, random, time, logging, argparse, requests
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))

from blog_poster.blog_config import (
    TARGET_URL, TARGET_BRAND, TARGET_KEYWORDS, ANCHOR_TEXTS,
    AUTHORITY_LINKS, MAX_BACKLINKS_PER_POST, NOFOLLOW_PERCENTAGE, ADD_AUTHOR_BIO,
    MEDIUM_CONFIG, DEVTO_CONFIG, HASHNODE_CONFIG, TUMBLR_CONFIG, WRITEAS_CONFIG,
    GITHUB_CONFIG, GITLAB_CONFIG, PASTEBIN_CONFIG,
)
from blog_poster.blog_utils import (
    generate_article_from_template, inject_backlinks, build_author_bio,
    add_authority_links, html_to_telegraph_nodes, html_to_markdown,
)
from blog_poster.templates.article_templates import get_templates, get_all_niches

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

# ── Humanizer ──────────────────────────────────────────────────

_CONTRACTIONS = {
    " you are ": " you're ", " it is ": " it's ", " they are ": " they're ",
    " we are ": " we're ", " do not ": " don't ", " does not ": " doesn't ",
    " is not ": " isn't ", " are not ": " aren't ", " will not ": " won't ",
    " cannot ": " can't ", " should not ": " shouldn't ", " would not ": " wouldn't ",
    " could not ": " couldn't ", " did not ": " didn't ", " have not ": " haven't ",
    " that is ": " that's ", " there is ": " there's ", " here is ": " here's ",
}
_OPENERS = [
    "Here's something worth knowing — ", "The truth is, ", "Think about it this way: ",
    "Simply put, ", "Let's be clear: ", "In practice, ", "For most people, ",
    "The reality is that ", "Here's the key insight: ", "Worth mentioning — ",
]
_RHETORICAL = [
    "But what does this actually mean for you?", "So why does this matter?",
    "How do you make this work in practice?", "Sound familiar?",
    "The question is — are you taking full advantage of this?",
]
_STATS = [
    "Studies show that 73% of businesses that invest in this see measurable results within 90 days.",
    "According to industry research, companies that adopt this approach grow 2× faster on average.",
    "A 2025 survey found that 8 out of 10 professionals now consider this a top priority.",
    "Data confirms that early adopters generate 60% more leads than those who wait.",
]

def humanize(html: str) -> str:
    for f, s in _CONTRACTIONS.items():
        html = html.replace(f, s)
    parts = html.split("</p>")
    for i in range(1, len(parts)):
        if "<p>" in parts[i] and random.random() < 0.25:
            parts[i] = parts[i].replace("<p>", f"<p>{random.choice(_OPENERS)}", 1)
    html = "</p>".join(parts)
    mid = html.find("</p>", len(html) // 2)
    if mid > 0 and random.random() < 0.7:
        html = html[:mid+4] + f"\n\n<p><em>{random.choice(_RHETORICAL)}</em></p>" + html[mid+4:]
    first_end = html.find("</p>")
    if first_end > 0 and random.random() < 0.6:
        html = html[:first_end+4] + f"\n\n<p>{random.choice(_STATS)}</p>" + html[first_end+4:]
    return html

# Local business niches — rotated for topical diversity
_LOCAL_NICHES = [
    ("local business directory Jaipur", "business"),
    ("best restaurants in Jaipur", "local_services"),
    ("top hospitals in Jaipur", "health"),
    ("best schools in Jaipur", "education"),
    ("top gyms in Jaipur", "health"),
    ("best hotels in Jaipur", "business"),
    ("Jaipur business listings", "business"),
    ("top services Jaipur", "local_services"),
    ("best shopping in Jaipur", "business"),
    ("trusted businesses Jaipur", "business"),
]

def build_article():
    keyword, niche = random.choice(_LOCAL_NICHES)
    tpl     = random.choice(get_templates(niche))
    art     = generate_article_from_template(tpl, keyword=keyword, brand=TARGET_BRAND, spin=True)
    art["content_html"] = humanize(art["content_html"])
    anchor  = random.choice(ANCHOR_TEXTS)
    art["content_html"] = inject_backlinks(art["content_html"], TARGET_URL, ANCHOR_TEXTS,
                                            max_links=MAX_BACKLINKS_PER_POST, nofollow_pct=NOFOLLOW_PERCENTAGE)
    art["content_html"] = add_authority_links(art["content_html"], AUTHORITY_LINKS)
    if ADD_AUTHOR_BIO:
        art["content_html"] += "\n\n" + build_author_bio(TARGET_URL, TARGET_BRAND, anchor)
    plain = re.sub(r"<[^>]+>", "", art["content_html"])
    art["word_count"]   = len(plain.split())
    art["anchor_text"]  = anchor
    art["plain_text"]   = plain
    return art

# ── Session ────────────────────────────────────────────────────

S = requests.Session()
S.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/123.0 Safari/537.36"})

def ok(url, name, da):   return True,  url, name, da, ""
def fail(name, da, err): return False, "",  name, da, str(err)[:150]

def txt(art):
    return f"{art['title']}\n{'='*60}\n\n{art['plain_text']}\n\n---\nVisit: {TARGET_URL}"

# ══════════════════════════════════════════════════════════════
#  PLATFORM FUNCTIONS
# ══════════════════════════════════════════════════════════════

# ── 1. Write.as (DA 69) ────────────────────────────────────────
def p_writeas(art):
    name, da = "Write.as", 69
    try:
        md = html_to_markdown(art["content_html"])
        r  = S.post("https://write.as/api/posts",
                    headers={"Content-Type":"application/json"},
                    json={"title": art["title"], "body": md, "font": "sans"}, timeout=30)
        d  = r.json()
        if r.status_code in (200,201):
            pid = d.get("data",{}).get("id","")
            return ok(f"https://write.as/{pid}", name, da)
        return fail(name, da, d.get("error_msg", f"HTTP {r.status_code}"))
    except Exception as e: return fail(name, da, e)

# ── 2. Telegraph (DA 93) ───────────────────────────────────────
_tph_token = {"v": ""}
def _tph():
    if _tph_token["v"]: return _tph_token["v"]
    tf = os.path.join(os.path.dirname(__file__), "blog_poster", "telegraph_token.txt")
    if os.path.exists(tf):
        _tph_token["v"] = open(tf).read().strip()
        if _tph_token["v"]: return _tph_token["v"]
    try:
        r = S.post("https://api.telegra.ph/createAccount",
                   json={"short_name": TARGET_BRAND[:32], "author_name": TARGET_BRAND, "author_url": TARGET_URL},
                   timeout=8)
        if r.json().get("ok"):
            _tph_token["v"] = r.json()["result"]["access_token"]
            os.makedirs(os.path.dirname(tf), exist_ok=True)
            open(tf,"w").write(_tph_token["v"])
            return _tph_token["v"]
    except: pass
    return ""

def p_telegraph(art):
    name, da = "Telegraph", 93
    try:
        tok = _tph()
        if not tok: return fail(name, da, "ISP blocked")
        nodes = html_to_telegraph_nodes(art["content_html"])
        r = S.post("https://api.telegra.ph/createPage",
                   json={"access_token": tok, "title": art["title"][:256],
                         "author_name": TARGET_BRAND, "author_url": TARGET_URL,
                         "content": json.dumps(nodes)}, timeout=25)
        d = r.json()
        if d.get("ok"): return ok(d["result"]["url"], name, da)
        return fail(name, da, d.get("error","unknown"))
    except Exception as e: return fail(name, da, e)

# ── 3. paste.ubuntu.com (DA 91) ────────────────────────────────
def p_ubuntu_paste(art):
    name, da = "paste.ubuntu.com", 91
    try:
        r = S.post("https://paste.ubuntu.com/",
                   data={"poster": TARGET_BRAND, "syntax": "text",
                         "content": txt(art), "expiration": "year"},
                   allow_redirects=True, timeout=25)
        if r.status_code == 200 and "paste.ubuntu.com/" in r.url and r.url != "https://paste.ubuntu.com/":
            return ok(r.url, name, da)
        m = re.search(r'paste\.ubuntu\.com/(\d+)', r.text)
        if m: return ok(f"https://paste.ubuntu.com/{m.group(1)}/", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 4. dpaste.com (DA 55) ──────────────────────────────────────
def p_dpaste(art):
    name, da = "dpaste.com", 55
    try:
        r = S.post("https://dpaste.com/api/v2/",
                   data={"content": txt(art), "syntax": "text",
                         "title": art["title"][:50], "expiry_days": 365}, timeout=20)
        if r.status_code in (200,201):
            u = r.text.strip().strip('"')
            if u.startswith("http"): return ok(u, name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 5. paste.rs (DA 35) ────────────────────────────────────────
def p_pasterns(art):
    name, da = "paste.rs", 35
    try:
        r = S.post("https://paste.rs/", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain; charset=utf-8"}, timeout=20)
        if r.status_code in (200,201):
            u = r.text.strip()
            if u.startswith("http"): return ok(u, name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 6. bytebin.lucko.me (DA 42) ────────────────────────────────
def p_bytebin(art):
    name, da = "bytebin.lucko.me", 42
    try:
        r = S.post("https://bytebin.lucko.me/post", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain; charset=utf-8"}, timeout=20)
        if r.status_code in (200,201):
            key = r.json().get("key","")
            if key: return ok(f"https://bytebin.lucko.me/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 7. paste.myst.rs (DA 40) ───────────────────────────────────
def p_pastemyst(art):
    name, da = "paste.myst.rs", 40
    try:
        r = S.post("https://paste.myst.rs/api/v2/paste",
                   headers={"Content-Type":"application/json"},
                   json={"expiry_date":"never",
                         "pasties":[{"title":art["title"][:50],"language":"plain text","code":txt(art)}]},
                   timeout=20)
        if r.status_code in (200,201):
            pid = r.json().get("_id","")
            if pid: return ok(f"https://paste.myst.rs/{pid}", name, da)
        return fail(name, da, f"HTTP {r.status_code}: {r.text[:80]}")
    except Exception as e: return fail(name, da, e)

# ── 8. haste.zneix.eu (DA 30) ──────────────────────────────────
def p_haste_zneix(art):
    name, da = "haste.zneix.eu", 30
    try:
        r = S.post("https://haste.zneix.eu/documents", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain"}, timeout=20)
        if r.status_code == 200:
            key = r.json().get("key","")
            if key: return ok(f"https://haste.zneix.eu/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 9. mclo.gs (DA 45) ─────────────────────────────────────────
def p_mclogs(art):
    name, da = "mclo.gs", 45
    try:
        r = S.post("https://api.mclo.gs/1/log",
                   data={"content": txt(art)},
                   headers={"Content-Type":"application/x-www-form-urlencoded"}, timeout=20)
        if r.status_code == 200:
            d = r.json()
            if d.get("success"): return ok(d.get("url",""), name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 10. pastefy.app (DA 40) ────────────────────────────────────
def p_pastefy(art):
    name, da = "pastefy.app", 40
    try:
        r = S.post("https://pastefy.app/api/v2/paste",
                   headers={"Content-Type":"application/json"},
                   json={"content": txt(art), "title": art["title"][:60], "type": "PASTE"},
                   timeout=20)
        if r.status_code in (200,201):
            d   = r.json()
            pid = d.get("paste",{}).get("id","") or d.get("id","")
            if pid: return ok(f"https://pastefy.app/{pid}", name, da)
        return fail(name, da, f"HTTP {r.status_code}: {r.text[:80]}")
    except Exception as e: return fail(name, da, e)

# ── 11. hastebin.com (DA 65) ───────────────────────────────────
def p_hastebin(art):
    name, da = "hastebin.com", 65
    try:
        r = S.post("https://hastebin.com/documents", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain"}, timeout=20)
        if r.status_code == 200:
            key = r.json().get("key","")
            if key: return ok(f"https://hastebin.com/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 12. pastes.dev (DA 30) ─────────────────────────────────────
def p_pastesdev(art):
    name, da = "pastes.dev", 30
    try:
        r = S.post("https://pastes.dev/documents", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain"}, timeout=20)
        if r.status_code == 200:
            key = r.json().get("key","")
            if key: return ok(f"https://pastes.dev/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 13. haste.unethical.engineer (DA 25) ───────────────────────
def p_haste_unethical(art):
    name, da = "haste.unethical.engineer", 25
    try:
        r = S.post("https://haste.unethical.engineer/documents", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain"}, timeout=15)
        if r.status_code == 200:
            key = r.json().get("key","")
            if key: return ok(f"https://haste.unethical.engineer/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 14. haste.schel.de (DA 25) ─────────────────────────────────
def p_haste_schel(art):
    name, da = "haste.schel.de", 25
    try:
        r = S.post("https://haste.schel.de/documents", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain"}, timeout=15)
        if r.status_code == 200:
            key = r.json().get("key","")
            if key: return ok(f"https://haste.schel.de/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 15. haste.pluralkit.me (DA 28) ─────────────────────────────
def p_haste_plural(art):
    name, da = "haste.pluralkit.me", 28
    try:
        r = S.post("https://haste.pluralkit.me/documents", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain"}, timeout=15)
        if r.status_code == 200:
            key = r.json().get("key","")
            if key: return ok(f"https://haste.pluralkit.me/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 16. haste.spg.gg (DA 25) ───────────────────────────────────
def p_haste_spg(art):
    name, da = "haste.spg.gg", 25
    try:
        r = S.post("https://haste.spg.gg/documents", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain"}, timeout=15)
        if r.status_code == 200:
            key = r.json().get("key","")
            if key: return ok(f"https://haste.spg.gg/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 17. toptal hastebin (DA 72) ────────────────────────────────
def p_toptal(art):
    name, da = "toptal-hastebin", 72
    try:
        r = S.post("https://hastebin.com/documents", data=txt(art).encode(),
                   headers={"Content-Type":"text/plain",
                             "Origin":"https://www.toptal.com",
                             "Referer":"https://www.toptal.com/developers/hastebin"},
                   timeout=20)
        if r.status_code == 200:
            key = r.json().get("key","")
            if key: return ok(f"https://www.toptal.com/developers/hastebin/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 18. rentry.co (DA 55) ──────────────────────────────────────
def p_rentry(art):
    name, da = "rentry.co", 55
    try:
        init = S.get("https://rentry.co", timeout=12)
        csrf = init.cookies.get("csrftoken","")
        if not csrf:
            m = re.search(r'csrfmiddlewaretoken.*?value=["\']([^"\']+)', init.text)
            if m: csrf = m.group(1)
        if not csrf: return fail(name, da, "no CSRF token")
        plain = re.sub(r"<[^>]+>", "", art["content_html"])
        content = f"# {art['title']}\n\n{plain}\n\n---\n[{TARGET_BRAND}]({TARGET_URL})"
        slug = "sc" + str(random.randint(10000, 99999))
        r = S.post("https://rentry.co/api/new",
                   data={"csrfmiddlewaretoken": csrf, "url": slug,
                         "edit_code": str(random.randint(1000,9999)), "text": content},
                   headers={"Referer":"https://rentry.co","X-CSRFToken":csrf,
                             "Content-Type":"application/x-www-form-urlencoded"},
                   timeout=20)
        if r.status_code == 200:
            try:
                u = r.json().get("url","")
                if u:
                    if not u.startswith("http"): u = "https://rentry.co/" + u.lstrip("/")
                    return ok(u, name, da)
            except: pass
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 19. HedgeDoc demo (DA 45) ──────────────────────────────────
def p_hedgedoc(art):
    name, da = "demo.hedgedoc.org", 45
    try:
        md = html_to_markdown(art["content_html"])
        full_md = f"# {art['title']}\n\n{md}\n\n---\n[{TARGET_BRAND}]({TARGET_URL})"
        r = S.post("https://demo.hedgedoc.org/new",
                   data=full_md.encode("utf-8"),
                   headers={"Content-Type":"text/markdown"},
                   allow_redirects=True, timeout=25)
        if r.status_code == 200 and "demo.hedgedoc.org" in r.url and "/new" not in r.url:
            return ok(r.url, name, da)
        return fail(name, da, f"HTTP {r.status_code} url={r.url}")
    except Exception as e: return fail(name, da, e)

# ── 20. md.vern.cc (HedgeDoc, DA 40) ──────────────────────────
def p_vern_md(art):
    name, da = "md.vern.cc", 40
    try:
        md = html_to_markdown(art["content_html"])
        full_md = f"# {art['title']}\n\n{md}\n\n---\n[{TARGET_BRAND}]({TARGET_URL})"
        r = S.post("https://md.vern.cc/new",
                   data=full_md.encode("utf-8"),
                   headers={"Content-Type":"text/markdown"},
                   allow_redirects=True, timeout=20)
        if r.status_code == 200 and "md.vern.cc" in r.url and "/new" not in r.url:
            return ok(r.url, name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 21. ghostbin.co (DA 40) ────────────────────────────────────
def p_ghostbin(art):
    name, da = "ghostbin.co", 40
    try:
        r = S.post("https://ghostbin.co/paste/new",
                   data={"text": txt(art), "lang": "text", "expire": "2d",
                         "password": "", "title": art["title"][:60]},
                   headers={"Content-Type":"application/x-www-form-urlencoded"},
                   allow_redirects=True, timeout=20)
        if r.status_code == 200 and "ghostbin.co/paste/" in r.url:
            return ok(r.url, name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 22. bpa.st (DA 35) ─────────────────────────────────────────
def p_bpaste(art):
    name, da = "bpa.st", 35
    try:
        r = S.post("https://bpa.st/api/v1/paste",
                   json={"expiry": "1day",
                         "files": [{"name": art["title"][:50] + ".txt",
                                    "lexer": "text", "content": txt(art)}]},
                   timeout=20)
        data = r.json()
        if r.status_code == 200 and "link" in data:
            return ok(data["link"], name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 23. pasteio.com (DA 35) ────────────────────────────────────
def p_pasteio(art):
    name, da = "pasteio.com", 35
    try:
        r = S.post("https://pasteio.com/api/v1/pastes",
                   headers={"Content-Type":"application/json"},
                   json={"content": txt(art), "visibility": "public"},
                   timeout=20)
        if r.status_code in (200,201):
            d   = r.json()
            pid = d.get("pasteId","") or d.get("id","") or d.get("key","")
            if pid: return ok(f"https://pasteio.com/{pid}", name, da)
        return fail(name, da, f"HTTP {r.status_code}: {r.text[:80]}")
    except Exception as e: return fail(name, da, e)

# ── 24. pastebin.pl (DA 35) ────────────────────────────────────
def p_pastebin_pl(art):
    name, da = "pastebin.pl", 35
    try:
        init = S.get("https://pastebin.pl", timeout=12)
        csrf = re.search(r'csrf.*?value=["\']([^"\']{20,})', init.text)
        token = csrf.group(1) if csrf else ""
        data = {"paste": txt(art), "submit": "paste"}
        if token: data["csrf_token"] = token
        r = S.post("https://pastebin.pl/", data=data,
                   headers={"Referer":"https://pastebin.pl/"},
                   allow_redirects=True, timeout=20)
        if r.status_code == 200 and "pastebin.pl" in r.url and r.url != "https://pastebin.pl/":
            return ok(r.url, name, da)
        m = re.search(r'href=["\']https?://pastebin\.pl/([a-zA-Z0-9]+)', r.text)
        if m: return ok(f"https://pastebin.pl/{m.group(1)}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 25. controlc.com (DA 45) ───────────────────────────────────
def p_controlc(art):
    name, da = "controlc.com", 45
    try:
        init = S.get("https://controlc.com/", timeout=12)
        ts = re.search(r"name=timestamp\s+value=['\"]([^'\"]+)['\"]", init.text)
        token = ts.group(1) if ts else ""
        r = S.post("https://controlc.com/index.php?act=submit",
                   data={"input_text": txt(art), "paste_title": art["title"][:60],
                         "antispam": "1", "website": "", "subdomain": "",
                         "timestamp": token},
                   headers={"Content-Type":"application/x-www-form-urlencoded",
                             "Referer":"https://controlc.com/"},
                   allow_redirects=True, timeout=20)
        m = re.search(r'href=["\']?(https://controlc\.com/[a-fA-F0-9]{6,})', r.text)
        if m: return ok(m.group(1), name, da)
        m2 = re.search(r'controlc\.com/([a-fA-F0-9]{6,})', r.text)
        if m2: return ok(f"https://controlc.com/{m2.group(1)}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 26. just-paste.it (DA 40) ──────────────────────────────────
def p_justpaste(art):
    name, da = "just-paste.it", 40
    try:
        init = S.get("https://just-paste.it", timeout=12)
        csrf = init.cookies.get("XSRF-TOKEN","")
        r = S.post("https://just-paste.it/api/article/save",
                   headers={"Content-Type":"application/json",
                             "X-XSRF-TOKEN": csrf},
                   json={"content": f"<h1>{art['title']}</h1>\n{art['content_html']}",
                         "visibility": "public"},
                   timeout=20)
        if r.status_code in (200,201):
            d   = r.json()
            pid = d.get("id","") or d.get("articleId","")
            if pid: return ok(f"https://just-paste.it/{pid}", name, da)
        return fail(name, da, f"HTTP {r.status_code}: {r.text[:80]}")
    except Exception as e: return fail(name, da, e)

# ── 27a. paste.ee / pastee.dev (DA 40) ────────────────────────
def p_pasteee(art):
    name, da = "pastee.dev", 40
    try:
        r = S.post("https://paste.ee/api",
                   data={"key": "public", "paste": txt(art), "format": "json"},
                   timeout=20)
        d = r.json()
        if r.status_code == 200 and d.get("status") == "success":
            return ok(d["paste"]["link"], name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 27b. nopaste.net (DA 35) ───────────────────────────────────
def p_nopaste(art):
    name, da = "nopaste.net", 35
    try:
        r = S.post("https://nopaste.net", data=txt(art).encode(),
                   headers={"Content-Type": "text/plain"}, timeout=20)
        m = re.search(r'(https://nopaste\.net/\S+)', r.text)
        if m: return ok(m.group(1).rstrip("\"'<>"), name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 27c. paste.armbian.com (DA 35) ─────────────────────────────
def p_armbian(art):
    name, da = "paste.armbian.com", 35
    try:
        r = S.post("https://paste.armbian.com/documents",
                   data=txt(art).encode(),
                   headers={"Content-Type": "text/plain"}, timeout=20)
        key = r.json().get("key", "")
        if key: return ok(f"https://paste.armbian.com/{key}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 27d. hedgedoc.envs.net (DA 40) ─────────────────────────────
def p_hedgedoc_envs(art):
    name, da = "hedgedoc.envs.net", 40
    try:
        r = S.post("https://hedgedoc.envs.net/new", allow_redirects=False, timeout=20)
        note_url = r.headers.get("Location", "")
        if not note_url:
            return fail(name, da, "No redirect URL")
        # Upload content to the note via API
        note_id = note_url.rstrip("/").split("/")[-1]
        content = f"# {art['title']}\n\n{txt(art)}"
        S.put(f"https://hedgedoc.envs.net/{note_id}/content",
              data=content.encode(), headers={"Content-Type": "text/markdown"}, timeout=20)
        return ok(note_url, name, da)
    except Exception as e: return fail(name, da, e)

# ── 27e. paste.centos.org (DA 50) ──────────────────────────────
def p_centos_paste(art):
    name, da = "paste.centos.org", 50
    try:
        r = S.post("https://paste.centos.org/",
                   data={"code": txt(art), "name": "anon",
                         "title": art["title"][:60], "lang": "text",
                         "snipurl": "", "private": "no",
                         "expire": "2678400", "submit": "Submit"},
                   allow_redirects=False, timeout=20)
        loc = r.headers.get("Location", "")
        if loc and "paste.centos.org/view/" in loc:
            return ok(loc if loc.startswith("http") else f"https://paste.centos.org{loc}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 27f. paste.opensuse.org (DA 55) ────────────────────────────
def p_opensuse_paste(art):
    name, da = "paste.opensuse.org", 55
    try:
        init = S.get("https://paste.opensuse.org/", timeout=12)
        csrf_m = re.search(r'<meta name="csrf-token" content="([^"]+)"', init.text)
        if not csrf_m:
            return fail(name, da, "no CSRF token")
        csrf = csrf_m.group(1)
        r = S.post("https://paste.opensuse.org/pastes",
                   data={"authenticity_token": csrf, "paste[author]": "anon",
                         "paste[title]": art["title"][:60],
                         "paste[code]": txt(art), "paste[private]": "0"},
                   allow_redirects=False, timeout=20)
        loc = r.headers.get("Location", "")
        if loc and "paste.opensuse.org/pastes/" in loc:
            return ok(loc if loc.startswith("http") else f"https://paste.opensuse.org{loc}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 27. heypasteit.com (DA 35) ─────────────────────────────────
def p_heypaste(art):
    name, da = "heypasteit.com", 35
    try:
        r = S.post("https://heypasteit.com/clip",
                   data={"YVCONTENTS": txt(art), "YVLANGUAGE": "TEXT",
                         "YVPRIVATE": "0"},
                   headers={"Content-Type":"application/x-www-form-urlencoded",
                             "Referer":"https://heypasteit.com/"},
                   allow_redirects=True, timeout=20)
        if r.status_code == 200 and "heypasteit.com/clip/" in r.url:
            return ok(r.url, name, da)
        m = re.search(r'heypasteit\.com/clip/([A-Z0-9]+)', r.text)
        if m: return ok(f"https://heypasteit.com/clip/{m.group(1)}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 28. paste.mozilla.org (DA 92) ──────────────────────────────
def p_mozilla_paste(art):
    name, da = "paste.mozilla.org", 92
    try:
        r = S.post("https://paste.mozilla.org/api/",
                   data={"content": txt(art), "lexer": "_text",
                         "expires": str(60*60*24*365), "format": "url"},
                   timeout=20)
        if r.status_code in (200,201):
            u = r.text.strip()
            if u.startswith("http"): return ok(u, name, da)
        return fail(name, da, f"HTTP {r.status_code}: {r.text[:80]}")
    except Exception as e: return fail(name, da, e)

# ── 29. pastebin.fi (DA 30) ────────────────────────────────────
def p_pastebin_fi(art):
    name, da = "pastebin.fi", 30
    try:
        r = S.post("https://pastebin.fi/",
                   data={"paste": txt(art), "submit": "Paste"},
                   headers={"Content-Type":"application/x-www-form-urlencoded"},
                   allow_redirects=True, timeout=20)
        if r.status_code == 200 and "pastebin.fi/" in r.url and r.url != "https://pastebin.fi/":
            return ok(r.url, name, da)
        m = re.search(r'pastebin\.fi/([a-zA-Z0-9]+)', r.text)
        if m: return ok(f"https://pastebin.fi/{m.group(1)}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 30. cl1p.net (DA 40) ───────────────────────────────────────
def p_cl1p(art):
    name, da = "cl1p.net", 40
    try:
        slug = "scotle-" + str(random.randint(100000, 999999))
        r = S.post(f"https://cl1p.net/{slug}",
                   data={"cl1pContent": txt(art), "cl1pURL": slug},
                   headers={"Content-Type":"application/x-www-form-urlencoded",
                             "Referer":f"https://cl1p.net/{slug}"},
                   allow_redirects=False, timeout=20)
        if r.status_code in (200, 201, 302):
            return ok(f"https://cl1p.net/{slug}", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── 31. Dev.to (DA 90, needs API key) ──────────────────────────
def p_devto(art):
    name, da = "Dev.to", 90
    api_key = DEVTO_CONFIG.get("api_key","")
    if not api_key: return fail(name, da, "api_key not set — get free at dev.to/settings/extensions")
    try:
        md = html_to_markdown(art["content_html"])
        r  = S.post("https://dev.to/api/articles",
                    headers={"api-key": api_key, "Content-Type":"application/json"},
                    json={"article":{"title":art["title"],"body_markdown":md,"published":True,
                                     "tags":["business","productivity","growth","tips"]}},
                    timeout=30)
        d  = r.json()
        if r.status_code in (200,201): return ok(d.get("url",""), name, da)
        return fail(name, da, d.get("error", f"HTTP {r.status_code}"))
    except Exception as e: return fail(name, da, e)

# ── 32. Medium (DA 95, needs token) ────────────────────────────
def p_medium(art):
    name, da = "Medium", 95
    token = MEDIUM_CONFIG.get("integration_token","")
    if not token: return fail(name, da, "integration_token not set — get at medium.com/me/settings")
    try:
        me = S.get("https://api.medium.com/v1/me",
                   headers={"Authorization":f"Bearer {token}"}, timeout=15)
        if me.status_code != 200: return fail(name, da, f"Auth failed HTTP {me.status_code}")
        uid = me.json().get("data",{}).get("id","")
        r = S.post(f"https://api.medium.com/v1/users/{uid}/posts",
                   headers={"Authorization":f"Bearer {token}","Content-Type":"application/json"},
                   json={"title":art["title"],"contentFormat":"html",
                         "content":f"<h1>{art['title']}</h1>\n{art['content_html']}",
                         "tags":["business","marketing"],"publishStatus":"public"},
                   timeout=30)
        d = r.json()
        if r.status_code in (200,201): return ok(d.get("data",{}).get("url",""), name, da)
        errors = d.get("errors",[{}])
        return fail(name, da, errors[0].get("message", f"HTTP {r.status_code}"))
    except Exception as e: return fail(name, da, e)

# ── 33. Hashnode (DA 79, needs token) ──────────────────────────
def p_hashnode(art):
    name, da = "Hashnode", 79
    token = HASHNODE_CONFIG.get("token","")
    pub_id = HASHNODE_CONFIG.get("publication_id","")
    if not token or not pub_id: return fail(name, da, "token or publication_id not set — hashnode.com/settings/developer")
    try:
        md = html_to_markdown(art["content_html"])
        mutation = "mutation PublishPost($input: PublishPostInput!) { publishPost(input: $input) { post { id url } } }"
        r = S.post("https://gql.hashnode.com",
                   headers={"Authorization": token},
                   json={"query": mutation, "variables":{"input":{"title":art["title"],"contentMarkdown":md,"publicationId":pub_id,"tags":[]}}},
                   timeout=30)
        d    = r.json()
        post = (d.get("data") or {}).get("publishPost",{}).get("post",{})
        if post: return ok(post.get("url",""), name, da)
        errs = d.get("errors",[])
        return fail(name, da, errs[0].get("message","No post") if errs else "No post")
    except Exception as e: return fail(name, da, e)

# ── 34a. GitHub Gist (DA 95, needs free token) ─────────────────
def p_github_gist(art):
    name, da = "GitHub Gist", 95
    token = GITHUB_CONFIG.get("token", "")
    if not token: return fail(name, da, "token not set — github.com/settings/tokens -> gist scope")
    try:
        r = S.post("https://api.github.com/gists",
                   headers={"Authorization": f"token {token}",
                             "Accept": "application/vnd.github+json"},
                   json={"description": art["title"],
                         "public": True,
                         "files": {art["title"][:50].replace(" ", "-") + ".md":
                                   {"content": f"# {art['title']}\n\n{txt(art)}"}}},
                   timeout=20)
        d = r.json()
        if r.status_code == 201:
            return ok(d.get("html_url", ""), name, da)
        return fail(name, da, d.get("message", f"HTTP {r.status_code}"))
    except Exception as e: return fail(name, da, e)

# ── 34b. GitLab Snippet (DA 95, needs free token) ───────────────
def p_gitlab_snippet(art):
    name, da = "GitLab Snippet", 95
    token = GITLAB_CONFIG.get("token", "")
    if not token: return fail(name, da, "token not set — gitlab.com/-/user_settings/personal_access_tokens")
    try:
        r = S.post("https://gitlab.com/api/v4/snippets",
                   headers={"PRIVATE-TOKEN": token, "Content-Type": "application/json"},
                   json={"title": art["title"][:80],
                         "file_name": "scotle-article.md",
                         "content": f"# {art['title']}\n\n{txt(art)}",
                         "visibility": "public"},
                   timeout=20)
        d = r.json()
        if r.status_code == 201:
            return ok(d.get("web_url", ""), name, da)
        return fail(name, da, str(d.get("message", f"HTTP {r.status_code}")))
    except Exception as e: return fail(name, da, e)

# ── 34c. Pastebin.com (DA 92, needs free API key) ───────────────
def p_pastebin(art):
    name, da = "Pastebin.com", 92
    api_key = PASTEBIN_CONFIG.get("api_dev_key", "")
    if not api_key: return fail(name, da, "api_dev_key not set — pastebin.com/api")
    try:
        r = S.post("https://pastebin.com/api/api_post.php",
                   data={"api_dev_key": api_key,
                         "api_option": "paste",
                         "api_paste_code": txt(art),
                         "api_paste_name": art["title"][:100],
                         "api_paste_private": "0",
                         "api_paste_expire_date": "N"},
                   timeout=20)
        if r.status_code == 200 and r.text.startswith("https://pastebin.com/"):
            return ok(r.text.strip(), name, da)
        return fail(name, da, r.text[:100])
    except Exception as e: return fail(name, da, e)

# ── 34. Tumblr (DA 81, needs OAuth) ────────────────────────────
def p_tumblr(art):
    name, da = "Tumblr", 81
    bn = TUMBLR_CONFIG.get("blog_name","")
    ck = TUMBLR_CONFIG.get("consumer_key","")
    if not bn or not ck: return fail(name, da, "blog_name or consumer_key not set in blog_config.py")
    try:
        from requests_oauthlib import OAuth1
        auth = OAuth1(TUMBLR_CONFIG["consumer_key"], TUMBLR_CONFIG["consumer_secret"],
                      TUMBLR_CONFIG["oauth_token"], TUMBLR_CONFIG["oauth_secret"])
        r = S.post(f"https://api.tumblr.com/v2/blog/{bn}/post", auth=auth,
                   data={"type":"text","title":art["title"],"body":art["content_html"],"format":"html"},
                   timeout=30)
        d = r.json()
        if r.status_code in (200,201):
            pid = str(d.get("response",{}).get("id",""))
            return ok(f"https://{bn}.tumblr.com/post/{pid}", name, da)
        return fail(name, da, d.get("meta",{}).get("msg", f"HTTP {r.status_code}"))
    except ImportError: return fail(name, da, "pip install requests-oauthlib")
    except Exception as e: return fail(name, da, e)

# ── NEW: HackMD (DA 72) — no credentials needed ────────────────
def p_hackmd(art):
    name, da = "HackMD", 72
    try:
        md = html_to_markdown(art["content_html"])
        full_md = f"# {art['title']}\n\n{md}\n\n---\n[{TARGET_BRAND}]({TARGET_URL})"
        r = S.post("https://hackmd.io/api/notes",
                   headers={"Content-Type": "application/json"},
                   json={"title": art["title"], "content": full_md,
                         "readPermission": "guest", "writePermission": "owner",
                         "commentPermission": "disabled"},
                   timeout=25)
        if r.status_code in (200, 201):
            d = r.json()
            nid = d.get("id", "")
            if nid:
                return ok(f"https://hackmd.io/{nid}", name, da)
        return fail(name, da, f"HTTP {r.status_code}: {r.text[:80]}")
    except Exception as e: return fail(name, da, e)

# ── NEW: snippet.host (DA 45) ───────────────────────────────────
def p_snippet_host(art):
    name, da = "snippet.host", 45
    try:
        r = S.post("https://snippet.host/api/create",
                   headers={"Content-Type": "application/json"},
                   json={"content": txt(art), "listed": True},
                   timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            slug = d.get("slug","") or d.get("id","")
            if slug:
                return ok(f"https://snippet.host/{slug}", name, da)
        return fail(name, da, f"HTTP {r.status_code}: {r.text[:80]}")
    except Exception as e: return fail(name, da, e)

# ── NEW: paste.debian.net (DA 60) ──────────────────────────────
def p_debian_paste(art):
    name, da = "paste.debian.net", 60
    try:
        r = S.post("https://paste.debian.net/",
                   data={"code": txt(art), "lang": "text",
                         "expire": "60", "submit": "Submit"},
                   headers={"Content-Type": "application/x-www-form-urlencoded",
                             "Referer": "https://paste.debian.net/"},
                   allow_redirects=True, timeout=20)
        if r.status_code == 200 and "paste.debian.net/" in r.url and r.url != "https://paste.debian.net/":
            return ok(r.url, name, da)
        m = re.search(r'paste\.debian\.net/(\d+)', r.text)
        if m:
            return ok(f"https://paste.debian.net/{m.group(1)}/", name, da)
        return fail(name, da, f"HTTP {r.status_code}")
    except Exception as e: return fail(name, da, e)

# ── NEW: pasty.ee (DA 38) ───────────────────────────────────────
def p_pasty(art):
    name, da = "pasty.ee", 38
    try:
        r = S.post("https://pasty.ee/api/v1/pastes",
                   headers={"Content-Type": "application/json"},
                   json={"content": txt(art)},
                   timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            pid = d.get("id","")
            if pid:
                return ok(f"https://pasty.ee/{pid}", name, da)
        return fail(name, da, f"HTTP {r.status_code}: {r.text[:80]}")
    except Exception as e: return fail(name, da, e)

# ── NEW: Write.as (authenticated blog — DA 69) ──────────────────
def p_writeas_auth(art):
    """Write.as with optional token for named blog posts."""
    name, da = "Write.as (auth)", 69
    token = WRITEAS_CONFIG.get("access_token", "")
    collection = WRITEAS_CONFIG.get("collection_alias", "")
    if not token or not collection:
        return fail(name, da, "access_token or collection_alias not set — optional, skip if not configured")
    try:
        md = html_to_markdown(art["content_html"])
        r = S.post(f"https://write.as/api/collections/{collection}/posts",
                   headers={"Authorization": f"Token {token}",
                             "Content-Type": "application/json"},
                   json={"title": art["title"], "body": md, "font": "sans"},
                   timeout=30)
        d = r.json()
        if r.status_code in (200, 201):
            pid = d.get("data", {}).get("id", "")
            if pid:
                return ok(f"https://write.as/{collection}/{pid}", name, da)
        return fail(name, da, d.get("error_msg", f"HTTP {r.status_code}"))
    except Exception as e: return fail(name, da, e)

# ── NEW: Riseup Pad (DA 72) — anonymous collaborative pad ───────
def p_riseup_pad(art):
    name, da = "Riseup Pad", 72
    try:
        pad_id = "scotle-" + str(random.randint(100000, 999999))
        content = txt(art)
        r = S.post(f"https://pad.riseup.net/p/{pad_id}/export/txt",
                   data=content.encode("utf-8"),
                   headers={"Content-Type": "text/plain"},
                   timeout=20)
        # Riseup creates pad on GET, then we set content
        S.get(f"https://pad.riseup.net/p/{pad_id}", timeout=15)
        r2 = S.post(f"https://pad.riseup.net/p/{pad_id}/import",
                    files={"file": (pad_id + ".txt", content.encode("utf-8"), "text/plain")},
                    timeout=20)
        if r2.status_code in (200, 201, 302):
            return ok(f"https://pad.riseup.net/p/{pad_id}", name, da)
        return fail(name, da, f"HTTP {r2.status_code}")
    except Exception as e: return fail(name, da, e)

# ══════════════════════════════════════════════════════════════
#  PLATFORM TABLE
# ══════════════════════════════════════════════════════════════

PLATFORMS = [
    # High DA — credentials needed (unlock with 5-min setup)
    ("GitHub Gist",             95, p_github_gist),
    ("GitLab Snippet",          95, p_gitlab_snippet),
    ("Medium",                  95, p_medium),
    ("Pastebin.com",            92, p_pastebin),
    ("paste.mozilla.org",       92, p_mozilla_paste),
    ("Telegraph",               93, p_telegraph),
    ("paste.ubuntu.com",        91, p_ubuntu_paste),
    ("Dev.to",                  90, p_devto),
    ("Tumblr",                  81, p_tumblr),
    ("Hashnode",                79, p_hashnode),
    ("toptal-hastebin",         72, p_toptal),
    ("hastebin.com",            65, p_hastebin),
    # Mid DA — no credentials needed
    ("Write.as",                69, p_writeas),
    ("HackMD",               72, p_hackmd),
    ("Riseup Pad",           72, p_riseup_pad),
    ("paste.debian.net",     60, p_debian_paste),
    ("snippet.host",         45, p_snippet_host),
    ("pasty.ee",             38, p_pasty),
    ("Write.as (auth)",      69, p_writeas_auth),
    ("rentry.co",               55, p_rentry),
    ("dpaste.com",              55, p_dpaste),
    ("controlc.com",            45, p_controlc),
    ("mclo.gs",                 45, p_mclogs),
    ("demo.hedgedoc.org",       45, p_hedgedoc),
    ("bytebin.lucko.me",        42, p_bytebin),
    ("pastefy.app",             40, p_pastefy),
    ("paste.myst.rs",           40, p_pastemyst),
    ("ghostbin.co",             40, p_ghostbin),
    ("md.vern.cc",              40, p_vern_md),
    ("just-paste.it",           40, p_justpaste),
    ("cl1p.net",                40, p_cl1p),
    # Lower DA — no credentials needed
    ("paste.opensuse.org",      55, p_opensuse_paste),
    ("paste.centos.org",        50, p_centos_paste),
    ("paste.rs",                35, p_pasterns),
    ("bpa.st",                  35, p_bpaste),
    ("pastee.dev",              40, p_pasteee),
    ("nopaste.net",             35, p_nopaste),
    ("paste.armbian.com",       35, p_armbian),
    ("hedgedoc.envs.net",       40, p_hedgedoc_envs),
    ("heypasteit.com",          35, p_heypaste),
    ("pasteio.com",             35, p_pasteio),
    ("pastebin.pl",             35, p_pastebin_pl),
    ("haste.zneix.eu",          30, p_haste_zneix),
    ("haste.pluralkit.me",      28, p_haste_plural),
    ("pastes.dev",              30, p_pastesdev),
    ("pastebin.fi",             30, p_pastebin_fi),
    ("haste.schel.de",          25, p_haste_schel),
    ("haste.unethical.engineer",25, p_haste_unethical),
    ("haste.spg.gg",            25, p_haste_spg),
]

# ══════════════════════════════════════════════════════════════
#  RUNNER
# ══════════════════════════════════════════════════════════════

RESULTS = []

def _rec(name, da, art, url, status, error=""):
    RESULTS.append({"Platform": name, "DA": da, "Title": art["title"][:80],
                    "Article URL": url, "Anchor Text": art.get("anchor_text",""),
                    "Word Count": art.get("word_count",0), "Status": status,
                    "Error": error[:120], "Date": datetime.now().strftime("%Y-%m-%d %H:%M")})

def run(count=1):
    published, failed = [], []
    for (name, da, fn) in PLATFORMS:
        for i in range(count):
            art = build_article()
            logger.info(f"→ {name} (DA {da}) ...")
            try:
                ok_f, url, pname, pda, err = fn(art)
            except Exception as exc:
                ok_f, url, pname, pda, err = False, "", name, da, str(exc)
            if ok_f:
                logger.info(f"  ✓ {url}")
                _rec(name, da, art, url, "Published")
                published.append((name, da, url))
            else:
                logger.warning(f"  ✗ {err}")
                _rec(name, da, art, "", "Failed", err)
                failed.append(name)
            time.sleep(random.uniform(3, 8))
    return published, failed

# ══════════════════════════════════════════════════════════════
#  EXCEL
# ══════════════════════════════════════════════════════════════

def save_excel(results, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    import datetime as _dt

    wb  = Workbook()
    pub = [r for r in results if r["Status"] == "Published"]
    fai = [r for r in results if r["Status"] == "Failed"]

    DA_HIGH = "C8E6C9"; DA_MED = "DCEDC8"; DA_LOW = "FFF9C4"; DA_VLOW = "F5F5F5"
    HDR_GREEN = "0D652D"; HDR_BLUE = "1A73E8"; LINK = "1558D6"
    ROW_PUB = "E8F5E9"; ROW_FAIL = "FFEBEE"

    def da_color(da):
        if da >= 70: return DA_HIGH
        if da >= 50: return DA_MED
        if da >= 30: return DA_LOW
        return DA_VLOW

    def wc_color(wc):
        if wc >= 500: return "C8E6C9"
        if wc >= 300: return "FFF9C4"
        return "FFCCBC"

    thin = Side(style="thin", color="DDDDDD")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_hdr(ws, headers, fill_hex, txt="FFFFFF", sz=11):
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h)
            c.font = Font(bold=True, color=txt, size=sz)
            c.fill = PatternFill("solid", fgColor=fill_hex)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    def hyperlink(ws, row, col, url, label=None):
        c = ws.cell(row, col, label or url)
        if url and url != "N/A":
            c.hyperlink = url
            c.font = Font(color=LINK, underline="single")

    # ── Sheet 1: Published Links ──────────────────────────────────
    ws = wb.active
    ws.title = "Published Links"
    cols1 = ["#", "Platform", "DA", "DA Band", "Title", "Live URL",
             "Anchor Text", "Words", "Date"]
    set_hdr(ws, cols1, HDR_GREEN)
    for i, r in enumerate(pub, 1):
        da  = r["DA"]
        wc  = r.get("Word Count") or 0
        band = ("DA 70+" if da >= 70 else "DA 50-69" if da >= 50
                else "DA 30-49" if da >= 30 else "DA <30")
        ws.append([i, r["Platform"], da, band, r["Title"],
                   r["Article URL"], r["Anchor Text"], wc, r["Date"]])
        row = i + 1
        rf  = PatternFill("solid", fgColor=da_color(da))
        for col in range(1, len(cols1) + 1):
            ws.cell(row, col).fill = rf
            ws.cell(row, col).border = brd
            ws.cell(row, col).alignment = Alignment(vertical="center")
        hyperlink(ws, row, 6, r["Article URL"])
        ws.cell(row, 8).fill = PatternFill("solid", fgColor=wc_color(wc))
        ws.cell(row, 3).alignment = Alignment(horizontal="center")
        ws.cell(row, 4).alignment = Alignment(horizontal="center")
    for col, w in zip(range(1, 10), [5, 22, 6, 10, 54, 58, 22, 8, 16]):
        ws.column_dimensions[get_column_letter(col)].width = w
    # Legend row
    lr = ws.max_row + 2
    ws.cell(lr, 1, "Legend:").font = Font(bold=True)
    for off, (lbl, clr) in enumerate([
        ("DA 70+", DA_HIGH), ("DA 50-69", DA_MED), ("DA 30-49", DA_LOW),
        ("Words 500+", "C8E6C9"), ("Words <300", "FFCCBC")
    ], 2):
        c = ws.cell(lr, off, lbl)
        c.fill = PatternFill("solid", fgColor=clr)
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=9)

    # ── Sheet 2: All Results ─────────────────────────────────────
    wa = wb.create_sheet("All Results")
    cols2 = ["#", "Platform", "DA", "Status", "Title", "Article URL",
             "Words", "Error Message", "Date"]
    set_hdr(wa, cols2, HDR_BLUE)
    for i, r in enumerate(results, 1):
        is_pub = r["Status"] == "Published"
        wc  = r.get("Word Count") or 0
        url = r["Article URL"] or "N/A"
        wa.append([i, r["Platform"], r["DA"], r["Status"],
                   r["Title"], url, wc, r.get("Error", "") or "", r["Date"]])
        row = i + 1
        rf  = PatternFill("solid", fgColor=ROW_PUB if is_pub else ROW_FAIL)
        for col in range(1, len(cols2) + 1):
            wa.cell(row, col).fill = rf
            wa.cell(row, col).border = brd
            wa.cell(row, col).alignment = Alignment(vertical="center")
        wa.cell(row, 4).font = Font(bold=True,
                                    color="2E7D32" if is_pub else "C62828")
        wa.cell(row, 4).alignment = Alignment(horizontal="center")
        wa.cell(row, 3).alignment = Alignment(horizontal="center")
        if is_pub and r["Article URL"]:
            hyperlink(wa, row, 6, r["Article URL"])
    for col, w in zip(range(1, 10), [5, 22, 6, 11, 52, 58, 8, 42, 16]):
        wa.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: DA Breakdown + Bar Chart ───────────────────────
    wd = wb.create_sheet("DA Breakdown")
    wd.merge_cells("A1:D1")
    tc = wd.cell(1, 1, "Published Backlinks by DA Range")
    tc.font = Font(bold=True, size=13, color=HDR_BLUE)
    tc.alignment = Alignment(horizontal="center")
    wd.row_dimensions[1].height = 24
    wd.append([])
    wd.append(["DA Range", "Platforms", "Links", "Avg Words"])
    for col in range(1, 5):
        c = wd.cell(3, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR_BLUE)
        c.alignment = Alignment(horizontal="center")

    clr_map = {"DA 70+": DA_HIGH, "DA 50-69": DA_MED,
               "DA 30-49": DA_LOW, "DA <30": DA_VLOW}
    bands = [("DA 70+", 70, 999), ("DA 50-69", 50, 69),
             ("DA 30-49", 30, 49), ("DA <30", 0, 29)]
    for label, lo, hi in bands:
        grp = [r for r in pub if lo <= r["DA"] <= hi]
        wcs = [r.get("Word Count") or 0 for r in grp]
        avg_wc = int(sum(wcs) / len(wcs)) if wcs else 0
        wd.append([label, len(set(r["Platform"] for r in grp)), len(grp), avg_wc])
        row = wd.max_row
        for col in range(1, 5):
            wd.cell(row, col).fill = PatternFill("solid", fgColor=clr_map[label])
            wd.cell(row, col).border = brd
            wd.cell(row, col).alignment = Alignment(horizontal="center")

    chart = BarChart()
    chart.type = "col"
    chart.title = "Links Published by DA Range"
    chart.y_axis.title = "Links"
    chart.x_axis.title = "DA Range"
    chart.style = 10; chart.width = 18; chart.height = 12
    chart.add_data(Reference(wd, min_col=3, max_col=3, min_row=3, max_row=7),
                   titles_from_data=True)
    chart.set_categories(Reference(wd, min_col=1, min_row=4, max_row=7))
    wd.add_chart(chart, "F3")
    for col, w in zip(range(1, 5), [14, 14, 10, 13]):
        wd.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 4: Summary ────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 55

    ws2.merge_cells("A1:B1")
    hc = ws2.cell(1, 1, "Scotle.org Backlink Campaign Report")
    hc.font = Font(bold=True, size=16, color=HDR_BLUE)
    hc.alignment = Alignment(horizontal="center")
    hc.fill = PatternFill("solid", fgColor="E3F2FD")
    ws2.row_dimensions[1].height = 30

    row = 2
    ws2.cell(row, 1, "Generated:").font = Font(bold=True)
    ws2.cell(row, 2, _dt.datetime.now().strftime("%Y-%m-%d %H:%M"))

    row += 2
    ws2.cell(row, 1, "CAMPAIGN STATS").font = Font(bold=True, size=12, color=HDR_BLUE)
    plats_tried = len(set(r["Platform"] for r in results))
    plats_ok    = len(set(r["Platform"] for r in pub))
    total_wc    = sum(r.get("Word Count") or 0 for r in pub)
    stats = [
        ("Platforms Tried",         plats_tried),
        ("Platforms That Worked",   plats_ok),
        ("Total Links Published",   len(pub)),
        ("Total Attempts",          len(results)),
        ("Platform Success Rate",   f"{plats_ok / max(plats_tried, 1) * 100:.0f}%"),
        ("Avg DA of Published Links", f"{sum(r['DA'] for r in pub) // max(len(pub), 1)}"),
        ("Total Words Published",   f"{total_wc:,}"),
        ("Avg Words per Post",      f"{total_wc // max(len(pub), 1)}"),
    ]
    for label, value in stats:
        row += 1
        ws2.cell(row, 1, label).font = Font(bold=True)
        ws2.cell(row, 2, value)

    row += 2
    ws2.cell(row, 1, "UNLOCK MORE PLATFORMS").font = Font(bold=True, size=12, color="B71C1C")
    unlock = [
        ("Medium (DA 95)",   "medium.com/me/settings > Integration tokens"),
        ("Dev.to (DA 90)",   "dev.to/settings/extensions > API Keys"),
        ("Tumblr (DA 81)",   "api.tumblr.com/console > OAuth keys"),
        ("Hashnode (DA 79)", "hashnode.com/settings/developer > Token + Publication ID"),
    ]
    for label, where in unlock:
        row += 1
        ws2.cell(row, 1, label).font = Font(bold=True, color="B71C1C")
        ws2.cell(row, 2, where)

    row += 2
    ws2.cell(row, 1, "ALL PUBLISHED LINKS").font = Font(bold=True, size=12, color=HDR_BLUE)
    row += 1
    ws2.cell(row, 1, "Platform (DA)").font = Font(bold=True, underline="single")
    ws2.cell(row, 2, "Live URL").font    = Font(bold=True, underline="single")
    last_plat = None
    for entry in sorted(pub, key=lambda x: (-x["DA"], x["Platform"])):
        row += 1
        if entry["Platform"] != last_plat:
            ws2.cell(row, 1, f"{entry['Platform']}  (DA {entry['DA']})").font = Font(bold=True)
            last_plat = entry["Platform"]
        else:
            ws2.cell(row, 1, "")
        hyperlink(ws2, row, 2, entry["Article URL"])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    logger.info(f"Excel saved: {path}")


def main():
    p = argparse.ArgumentParser(description="Big Blog Poster — 40 platforms")
    p.add_argument("--count", type=int, default=1, help="Posts per platform")
    args = p.parse_args()

    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    out = os.path.join(os.path.dirname(__file__), "data", f"Scotle_BigPost_{date_str}.xlsx")

    logger.info("="*65)
    logger.info(f"  Big Poster  |  {len(PLATFORMS)} platforms  |  count={args.count}")
    logger.info(f"  Target: {TARGET_URL}")
    logger.info("="*65)

    published, failed = run(count=args.count)

    logger.info("")
    logger.info("="*65)
    logger.info(f"  ✓ PUBLISHED : {len(published)}")
    logger.info(f"  ✗ FAILED    : {len(failed)}")
    logger.info("="*65)

    if published:
        logger.info("\n  LIVE LINKS:")
        for name, da, url in sorted(published, key=lambda x: -x[1]):
            logger.info(f"    DA{da:3d}  {name:30s}  {url}")

    if RESULTS:
        save_excel(RESULTS, out)
        print(f"\n  Excel report: {out}")

if __name__ == "__main__":
    main()
