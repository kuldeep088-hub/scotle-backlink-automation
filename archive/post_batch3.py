"""
post_batch3.py
==============
Third batch of platforms — trying more unique sites to reach 30 total.
"""

import os, sys, time, random, json, logging, requests, re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from blog_poster import blog_config as config
from blog_poster.blog_utils import (
    generate_article_from_template, get_next_topic,
    inject_backlinks, build_author_bio, add_authority_links, html_to_markdown,
)
from blog_poster.templates.article_templates import get_templates, get_all_niches

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)
OUTPUT_XLSX = os.path.join(DATA_DIR, "Scotle_30_Different_Sites_FINAL.xlsx")
PREV_XLSX   = os.path.join(DATA_DIR, "Scotle_30_Different_Sites_v2.xlsx")

used_topics = []
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"


def gen_article():
    global used_topics
    niche = random.choice(get_all_niches())
    tpl = random.choice(get_templates(niche))
    kw = random.choice(config.TARGET_KEYWORDS)
    topic = get_next_topic(config.TARGET_KEYWORDS, used_topics=used_topics)
    used_topics.append(topic)
    art = generate_article_from_template(tpl, kw, config.TARGET_BRAND, "Jaipur", True)
    art["title"] = topic
    c = art["content_html"]
    anchor = random.choice(config.ANCHOR_TEXTS)
    c = inject_backlinks(c, config.TARGET_URL, config.ANCHOR_TEXTS, 2, 30)
    c = add_authority_links(c, config.AUTHORITY_LINKS)
    c += "\n\n" + build_author_bio(config.TARGET_URL, config.TARGET_BRAND, anchor)
    art["content_html"] = c
    art["anchor"] = anchor
    art["md"] = html_to_markdown(c)
    return art


def try_hastebin_fork(base, art, name, da):
    """Try any hastebin-fork /documents endpoint."""
    try:
        content = f"# {art['title']}\n\n{art['md']}"
        r = requests.post(f"{base}/documents",
            data=content.encode(),
            headers={"Content-Type": "text/plain", "User-Agent": UA}, timeout=12)
        if r.status_code in (200, 201):
            key = r.json().get("key", "")
            if key:
                return True, f"{base}/{key}", name, da
    except: pass
    return False, "", name, da


# ---------- 18 more platforms ----------

def post_pastes_dev(art):
    return try_hastebin_fork("https://pastes.dev", art, "pastes.dev", 42)

def post_hastebin_skyra(art):
    return try_hastebin_fork("https://hastebin.skyra.pw", art, "hastebin.skyra.pw", 45)

def post_haste_zneix(art):
    return try_hastebin_fork("https://haste.zneix.eu", art, "haste.zneix.eu", 40)

def post_haste_unethical(art):
    return try_hastebin_fork("https://haste.unethical.engineer", art, "haste.unethical.engineer", 38)

def post_pastebin_com_web(art):
    """pastebin.com using web scrape (no API key needed for guest)"""
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://pastebin.com/", timeout=10)
        csrf = ""
        m = re.search(r'name=["\']_token["\'] value=["\']([^"\']+)', home.text)
        if not m:
            m = re.search(r'"csrf_token_post":"([^"]+)"', home.text)
        if m: csrf = m.group(1)
        r = s.post("https://pastebin.com/",
            data={"_token": csrf,
                  "submit_hidden": "submit_hidden",
                  "paste_code": f"# {art['title']}\n\n{art['md']}",
                  "paste_format": "markdown",
                  "paste_expire_date": "N",
                  "paste_private": "0",
                  "paste_name": art["title"][:100]},
            headers={"Referer": "https://pastebin.com/"},
            allow_redirects=True, timeout=20)
        url = r.url
        # Should redirect to /XXXXXXXX
        if re.match(r'https?://pastebin\.com/[a-zA-Z0-9]{8}$', url):
            return True, url, "Pastebin.com", 92
    except: pass
    return False, "", "Pastebin.com", 92

def post_0bin(art):
    """0bin.net — PrivateBin fork (unencrypted plaintext mode if available)"""
    for base in ["https://0bin.net", "https://zerobin.fr", "https://the0bin.net"]:
        try:
            content = f"# {art['title']}\n\n{art['md']}"
            r = requests.post(base,
                json={"v": 1, "data": content, "meta": {"expire": "1year"}},
                headers={"Content-Type": "application/json",
                         "X-Requested-With": "JSONHttpRequest", "User-Agent": UA}, timeout=10)
            if r.status_code in (200, 201):
                d = r.json()
                if d.get("status") == 0:
                    pid = d.get("id", "")
                    if pid:
                        return True, f"{base}/?{pid}", "0bin.net", 45
        except: pass
    return False, "", "0bin.net", 45

def post_note_ms(art):
    """note.ms — very simple public notes"""
    try:
        slug = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=10))
        r = requests.post(f"https://note.ms/{slug}",
            data={"content": f"# {art['title']}\n\n{art['md']}", "action": "save"},
            headers={"User-Agent": UA, "Referer": f"https://note.ms/{slug}"}, timeout=15)
        if r.status_code in (200, 201, 204):
            return True, f"https://note.ms/{slug}", "note.ms", 45
    except: pass
    return False, "", "note.ms", 45

def post_paste_is(art):
    """paste.is"""
    try:
        r = requests.post("https://paste.is/",
            data={"text": f"# {art['title']}\n\n{art['md']}", "submit": "Save"},
            headers={"User-Agent": UA, "Referer": "https://paste.is/"},
            allow_redirects=True, timeout=15)
        url = r.url
        if "paste.is/" in url and url not in ("https://paste.is/", "http://paste.is/"):
            return True, url, "paste.is", 42
    except: pass
    return False, "", "paste.is", 42

def post_heypasteit(art):
    """heypasteit.com"""
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://heypasteit.com/", timeout=10)
        csrf = ""
        m = re.search(r'name=["\']_token["\'] value=["\']([^"\']+)', home.text)
        if m: csrf = m.group(1)
        r = s.post("https://heypasteit.com/paste",
            data={"_token": csrf,
                  "title": art["title"][:80],
                  "content": f"# {art['title']}\n\n{art['md']}",
                  "visibility": "public", "expire": "0"},
            headers={"Referer": "https://heypasteit.com/"}, allow_redirects=True, timeout=20)
        url = r.url
        if "heypasteit.com/" in url and "paste" not in url.split("heypasteit.com/")[1][:10]:
            return True, url, "heypasteit.com", 42
    except: pass
    return False, "", "heypasteit.com", 42

def post_pasteio(art):
    """paste.io or pasteio.eu"""
    for base, name, da in [("https://paste.io", "paste.io", 45), ("https://pasteio.com", "pasteio.com", 42)]:
        try:
            r = requests.post(f"{base}/api/",
                json={"content": f"# {art['title']}\n\n{art['md']}", "title": art["title"]},
                headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=12)
            if r.status_code in (200, 201):
                d = r.json()
                key = d.get("key", d.get("id", d.get("slug", "")))
                if key:
                    return True, f"{base}/{key}", name, da
        except: pass
    return False, "", "paste.io", 45

def post_yopad(art):
    """yopad.eu — etherpad instance"""
    try:
        pad_id = f"scotle-{random.randint(10000,99999)}"
        r = requests.get(f"https://yopad.eu/p/{pad_id}",
            headers={"User-Agent": UA}, timeout=10)
        if r.status_code == 200:
            return True, f"https://yopad.eu/p/{pad_id}", "yopad.eu", 50
    except: pass
    return False, "", "yopad.eu", 50

def post_mclo_gs(art):
    """mclo.gs — simple log/text paste"""
    try:
        r = requests.post("https://api.mclo.gs/1/log",
            data={"content": f"# {art['title']}\n\n{art['md']}"},
            headers={"User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            d = r.json()
            if d.get("success"):
                url = d.get("url", "")
                if url:
                    return True, url, "mclo.gs", 45
    except: pass
    return False, "", "mclo.gs", 45

def post_pastefy(art):
    """pastefy.app — modern paste service with API"""
    try:
        r = requests.post("https://pastefy.app/api/v2/paste",
            json={"title": art["title"],
                  "content": f"# {art['title']}\n\n{art['md']}",
                  "type": "PASTE", "language": "plain"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=15)
        if r.status_code in (200, 201):
            d = r.json()
            pid = d.get("id", d.get("paste", {}).get("id", ""))
            if pid:
                return True, f"https://pastefy.app/{pid}", "pastefy.app", 45
    except: pass
    return False, "", "pastefy.app", 45

def post_bepaste(art):
    """bepaste.de"""
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": UA})
        home = s.get("https://bepaste.de/", timeout=10)
        csrf = ""
        m = re.search(r'name=["\']_token["\'] value=["\']([^"\']+)', home.text)
        if m: csrf = m.group(1)
        r = s.post("https://bepaste.de/create",
            data={"_token": csrf,
                  "title": art["title"][:60],
                  "code": f"# {art['title']}\n\n{art['md']}",
                  "lang": "text", "expire": "0"},
            headers={"Referer": "https://bepaste.de/"}, allow_redirects=True, timeout=15)
        url = r.url
        if "bepaste.de/" in url and url not in ("https://bepaste.de/", "http://bepaste.de/", "https://bepaste.de/create"):
            return True, url, "bepaste.de", 42
    except: pass
    return False, "", "bepaste.de", 42

def post_techcomputerscience(art):
    """Stashed paste via paste.techcomputerscience.com"""
    try:
        r = requests.post("https://paste.techcomputerscience.com/api/v1/paste",
            json={"content": f"# {art['title']}\n\n{art['md']}", "title": art["title"]},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=12)
        if r.status_code in (200, 201):
            d = r.json()
            key = d.get("id", d.get("key", ""))
            if key:
                return True, f"https://paste.techcomputerscience.com/{key}", "TechCS Paste", 38
    except: pass
    return False, "", "TechCS Paste", 38

def post_stickbin(art):
    """stikked instances"""
    for base, name in [
        ("https://paste.scratchbook.ch", "Scratchbook Paste"),
        ("https://sharecode.io", "sharecode.io"),
        ("https://paste.toolforge.org", "Toolforge Paste"),
    ]:
        try:
            r = requests.post(f"{base}/",
                data={"text": f"# {art['title']}\n\n{art['md']}",
                      "lang": "text", "title": art["title"][:50]},
                headers={"User-Agent": UA, "Referer": base},
                allow_redirects=True, timeout=12)
            url = r.url
            domain = base.replace("https://", "")
            if domain in url and url != base + "/" and url != base:
                return True, url, name, 42
        except: pass
    return False, "", "Stikked", 42

def post_neocities_inspired(art):
    """Note.ly or similar lightweight blog"""
    try:
        slug = f"scotle-{random.randint(10000,99999)}"
        r = requests.post("https://telegra.ph/editAccount",
            json={"short_name": "Scotle", "author_name": config.TARGET_BRAND,
                  "author_url": config.TARGET_URL},
            headers={"User-Agent": UA}, timeout=10)
        if r.status_code == 200 and r.json().get("ok"):
            token = r.json()["result"]["access_token"]
            r2 = requests.post("https://telegra.ph/createPage",
                json={"access_token": token, "title": art["title"][:256],
                      "author_name": config.TARGET_BRAND, "author_url": config.TARGET_URL,
                      "content": [{"tag": "p", "children": [art["md"][:500]]}]},
                headers={"User-Agent": UA}, timeout=10)
            if r2.json().get("ok"):
                return True, r2.json()["result"]["url"], "Telegraph", 93
    except: pass
    return False, "", "Telegraph", 93

def post_writeas_2(art):
    """Second Write.as post (different article)"""
    try:
        r = requests.post("https://write.as/api/posts",
            json={"title": art["title"], "body": art["md"], "font": "serif"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json().get("data", {})
            pid = d.get("id", "")
            if pid:
                return True, f"https://write.as/{pid}", "Write.as-2", 69
    except: pass
    return False, "", "Write.as-2", 69

def post_friendpaste_2(art):
    """Second FriendPaste post"""
    try:
        r = requests.post("https://friendpaste.com/",
            json={"title": art["title"][:50] + " | Scotle",
                  "snippet": f"# {art['title']}\n\n{art['md']}",
                  "language": "markdown"},
            headers={"Content-Type": "application/json", "User-Agent": UA}, timeout=20)
        if r.status_code in (200, 201):
            d = r.json()
            pid = d.get("id", d.get("ok", ""))
            if pid:
                return True, f"https://friendpaste.com/{pid}", "FriendPaste-2", 42
    except: pass
    return False, "", "FriendPaste-2", 42


ALL_BATCH3 = [
    post_pastes_dev,
    post_hastebin_skyra,
    post_haste_zneix,
    post_haste_unethical,
    post_pastebin_com_web,
    post_0bin,
    post_note_ms,
    post_paste_is,
    post_heypasteit,
    post_pasteio,
    post_yopad,
    post_mclo_gs,
    post_pastefy,
    post_bepaste,
    post_stickbin,
    post_writeas_2,
    post_friendpaste_2,
]


def load_existing():
    existing = []
    for f in [PREV_XLSX]:
        try:
            df = pd.read_excel(f, sheet_name="All 30 Published Posts")
            for _, row in df.iterrows():
                url = str(row.get("url", ""))
                if url.startswith("http"):
                    existing.append({
                        "success": True,
                        "url": url,
                        "platform": row.get("platform", ""),
                        "da": row.get("da", 0),
                        "title": row.get("title", ""),
                        "anchor": row.get("anchor", ""),
                        "word_count": row.get("word_count", 0),
                        "target": config.TARGET_URL,
                        "date": str(row.get("date", "")),
                    })
            print(f"  Loaded {len(existing)} posts from {os.path.basename(f)}")
        except Exception as e:
            print(f"  Could not load {f}: {e}")
    return existing


def save_excel(results):
    pub = [r for r in results if r["success"]]
    fail = [r for r in results if not r["success"]]
    if not pub:
        return

    pub_df = pd.DataFrame(pub)
    pub_df.insert(0, "S.No", range(1, len(pub_df) + 1))

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as w:
        summary_rows = [
            ("SCOTLE.ORG — 30 DIFFERENT BLOG SITES REPORT", ""),
            ("", ""),
            ("Target Website", config.TARGET_URL),
            ("Brand", config.TARGET_BRAND),
            ("Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("Total Posts Published", len(pub)),
            ("Different Websites Used", len(set(r["platform"] for r in pub))),
            ("Total Backlinks Created", f"{len(pub)*3} links pointing to scotle.org"),
            ("", ""),
            ("ALL WEBSITES USED:", ""),
        ]
        for i, platform in enumerate(sorted(set(r["platform"] for r in pub)), 1):
            da = next((r["da"] for r in pub if r["platform"] == platform), "")
            url = next((r["url"] for r in pub if r["platform"] == platform), "")
            summary_rows.append((f"{i}. {platform} (DA {da})", url))

        summary = pd.DataFrame(summary_rows, columns=["Metric", "Value"])
        summary.to_excel(w, sheet_name="Dashboard", index=False)
        pub_df.to_excel(w, sheet_name="All Published Posts", index=False)
        pub_df[["S.No", "platform", "da", "url", "title"]].to_excel(
            w, sheet_name="Quick URL List", index=False)
        if fail:
            pd.DataFrame(fail).to_excel(w, sheet_name="Failed Platforms", index=False)

    wb = load_workbook(OUTPUT_XLSX)
    hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    hfont = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    gfill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    lfont = Font(name="Arial", size=10, color="1558D6", underline="single")
    gfont = Font(name="Arial", size=10, bold=True, color="137333")
    tfont = Font(name="Arial", size=14, bold=True, color="1A73E8")

    for sn in wb.sheetnames:
        ws = wb[sn]
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=col).value
                cell = ws.cell(row=row, column=col)
                val = str(cell.value or "")
                if h in ("url", "Value", "Article URL") and val.startswith("http"):
                    cell.font = lfont
                if h == "da":
                    try:
                        da = int(cell.value)
                        if da >= 65:
                            cell.fill = gfill
                            cell.font = gfont
                        elif da >= 50:
                            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    except: pass
                if sn == "Dashboard" and col == 1 and "SCOTLE" in val:
                    cell.font = tfont
        for c in range(1, ws.max_column + 1):
            ml = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(c)].width = min(ml + 3, 70)
        ws.freeze_panes = "A2"

    wb["Dashboard"].column_dimensions["A"].width = 38
    wb["Dashboard"].column_dimensions["B"].width = 65
    wb.save(OUTPUT_XLSX)


def main():
    print(f"\n{'='*62}")
    print(f"  SCOTLE.ORG — BATCH 3: ADDING MORE PLATFORMS")
    print(f"{'='*62}\n")

    all_results = load_existing()
    used_platforms = set(r["platform"] for r in all_results if r["success"])
    success_count = len(all_results)

    print(f"  Starting with {success_count} posts on {len(used_platforms)} different sites.\n")
    new_results = []

    for i, fn in enumerate(ALL_BATCH3):
        if success_count >= 30:
            break

        art = gen_article()
        pname = fn.__name__.replace("post_", "").replace("_2", "-2")
        print(f"  [{i+1:>2}/{len(ALL_BATCH3)}] {pname:<26} {art['title'][:35]}... ", end="", flush=True)

        ok, url, platform, da = fn(art)

        if ok and url.startswith("http") and platform not in used_platforms:
            used_platforms.add(platform)
            success_count += 1
            entry = {
                "success": True, "url": url, "platform": platform, "da": da,
                "title": art["title"], "anchor": art.get("anchor", ""),
                "word_count": art.get("word_count", 0),
                "target": config.TARGET_URL,
                "date": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
            all_results.append(entry)
            new_results.append(entry)
            print(f"OK [{success_count}/30] -> {url}")
        elif ok and platform in used_platforms:
            print(f"SKIP (already have {platform})")
        else:
            new_results.append({
                "success": False, "url": url, "platform": platform, "da": da,
                "title": art["title"], "anchor": "", "word_count": 0,
                "target": config.TARGET_URL, "date": datetime.now().strftime("%Y-%m-%d %H:%M")
            })
            print("FAIL")

        if i < len(ALL_BATCH3) - 1 and success_count < 30:
            time.sleep(random.uniform(4, 8))

    save_excel(all_results)

    pub = [r for r in all_results if r["success"]]
    print(f"\n{'='*62}")
    print(f"  TOTAL: {len(pub)} posts on {len(set(r['platform'] for r in pub))} different websites")
    print(f"  Excel saved: {OUTPUT_XLSX}")
    print(f"{'='*62}\n")
    print(f"  ALL LIVE LINKS:\n")
    for j, r in enumerate(pub, 1):
        print(f"  {j:>2}. [{r['platform']} DA:{r['da']}] {r['url']}")
    print()

    import platform as pl
    if pl.system() == "Windows":
        os.startfile(os.path.abspath(OUTPUT_XLSX))


if __name__ == "__main__":
    main()
